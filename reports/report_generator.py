"""
Report Generator - CSV/JSON/PDF/Excel formatida hisobot yaratish (yangi TZ 13-bo'lim).

Qo'llab-quvvatlanadi:
  - Alertlar ro'yxati (CSV, JSON, Excel)
  - PDF "Executive Report" - davr bo'yicha statistika, MITRE ATT&CK
    taqsimoti, eng ko'p ta'sirlangan qurilmalar, batafsil alertlar jadvali
  - Excel - "Summary" varag'i FORMULALAR orqali hisoblaydi (COUNTIF/COUNTA),
    "Alerts" varag'idagi xom ma'lumotdan - qattiq kodlangan raqamlar emas

Ishga tushirish (CLI):
    python -m reports.report_generator --period-days 7 --format csv,json,pdf,excel --output ./report_output
"""
import argparse
import csv
import json
import os
import sys
from collections import Counter
from datetime import datetime, timedelta

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.database import get_session
from db.models import Alert, Device, FileEvent, utcnow


def _alert_row(session, alert: Alert) -> dict:
    device = session.query(Device).filter(Device.id == alert.device_id).first() if alert.device_id else None
    return {
        "id": alert.id,
        "timestamp": alert.timestamp.strftime("%Y-%m-%d %H:%M:%S") if alert.timestamp else "",
        "severity": alert.severity or "",
        "hostname": device.hostname if device else "",
        "ip_address": device.ip_address if device else "",
        "mac_address": device.mac_address if device else "",
        "reason": (alert.reason or "").replace("\n", " | "),
        "action_taken": alert.action_taken or "",
        "notified": "ha" if alert.notified else "yo'q",
        "mitre_technique_id": alert.mitre_technique_id or "",
        "mitre_technique_name": alert.mitre_technique_name or "",
        "mitre_tactic": alert.mitre_tactic or "",
    }


def get_alerts_since(session, since: datetime):
    return (
        session.query(Alert)
        .filter(Alert.timestamp >= since)
        .order_by(Alert.timestamp.desc())
        .all()
    )


def export_alerts_csv(alert_rows: list, filepath: str):
    if not alert_rows:
        fieldnames = ["id", "timestamp", "severity", "hostname", "ip_address", "mac_address",
                      "reason", "action_taken", "notified", "mitre_technique_id",
                      "mitre_technique_name", "mitre_tactic"]
    else:
        fieldnames = list(alert_rows[0].keys())

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(alert_rows)


def export_alerts_json(alert_rows: list, filepath: str):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(alert_rows, f, ensure_ascii=False, indent=2)


def export_summary_pdf(summary: dict, alert_rows: list, filepath: str, title: str = "Xavfsizlik Hisoboti"):
    """Executive Summary + batafsil alertlar jadvali - PDF (reportlab/Platypus)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )

    doc = SimpleDocTemplate(filepath, pagesize=letter,
                             topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(title, styles["Title"]))
    story.append(Paragraph(f"Davr: {summary['period_start']} — {summary['period_end']}", styles["Normal"]))
    story.append(Spacer(1, 16))

    story.append(Paragraph("Umumiy statistika", styles["Heading2"]))
    stat_table = Table(
        [["Ko'rsatkich", "Qiymat"],
         ["Jami alertlar", str(summary["total_alerts"])],
         ["Xabar berilmagan alertlar", str(summary["unnotified_alerts"])],
         ["Zararli fayllar aniqlandi", str(summary["malicious_files_detected"])],
         ["Toza fayllar tekshirildi", str(summary["clean_files_scanned"])]],
        colWidths=[9 * cm, 6 * cm],
    )
    stat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2634")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fa")]),
    ]))
    story.append(stat_table)
    story.append(Spacer(1, 16))

    if summary["severity_breakdown"]:
        story.append(Paragraph("Daraja bo'yicha taqsimot", styles["Heading2"]))
        severity_rows = [["Daraja", "Soni"]] + [[k.upper(), str(v)] for k, v in summary["severity_breakdown"].items()]
        sev_table = Table(severity_rows, colWidths=[9 * cm, 6 * cm])
        sev_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2634")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ]))
        story.append(sev_table)
        story.append(Spacer(1, 16))

    if summary["mitre_technique_breakdown"]:
        story.append(Paragraph("MITRE ATT&CK texnikalari", styles["Heading2"]))
        mitre_rows = [["Texnika ID", "Soni"]] + [[k, str(v)] for k, v in summary["mitre_technique_breakdown"].items()]
        mitre_table = Table(mitre_rows, colWidths=[9 * cm, 6 * cm])
        mitre_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2634")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ]))
        story.append(mitre_table)
        story.append(Spacer(1, 16))

    if summary["top_affected_devices"]:
        story.append(Paragraph("Eng ko'p ta'sirlangan qurilmalar", styles["Heading2"]))
        device_rows = [["Hostname", "IP manzil", "Alertlar soni"]] + [
            [d["hostname"], d["ip_address"], str(d["alert_count"])] for d in summary["top_affected_devices"]
        ]
        device_table = Table(device_rows, colWidths=[6 * cm, 5 * cm, 4 * cm])
        device_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2634")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ]))
        story.append(device_table)

    if alert_rows:
        story.append(PageBreak())
        story.append(Paragraph("Batafsil alertlar ro'yxati", styles["Heading2"]))
        story.append(Spacer(1, 8))

        small_style = ParagraphStyle("small", parent=styles["Normal"], fontSize=7, leading=9)
        detail_rows = [["Vaqt", "Daraja", "Qurilma", "Tafsilot", "MITRE"]]
        for row in alert_rows[:200]:
            reason = (row["reason"] or "")[:90]
            detail_rows.append([
                Paragraph(row["timestamp"], small_style),
                Paragraph(row["severity"].upper(), small_style),
                Paragraph(row["hostname"] or row["ip_address"] or "-", small_style),
                Paragraph(reason, small_style),
                Paragraph(row["mitre_technique_id"] or "-", small_style),
            ])
        detail_table = Table(detail_rows, colWidths=[2.8 * cm, 2 * cm, 3 * cm, 7 * cm, 2.2 * cm], repeatRows=1)
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2634")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fa")]),
        ]))
        story.append(detail_table)

        if len(alert_rows) > 200:
            story.append(Spacer(1, 8))
            story.append(Paragraph(
                f"... va yana {len(alert_rows) - 200} ta alert (to'liq ro'yxat uchun CSV/JSON eksportidan foydalaning).",
                styles["Italic"],
            ))

    doc.build(story)


def export_alerts_excel(summary: dict, alert_rows: list, filepath: str):
    """Excel hisobot - 2 varaq: "Summary" (formulalar bilan) va "Alerts" (batafsil ro'yxat)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_alerts = wb.active
    ws_alerts.title = "Alerts"

    headers = ["ID", "Vaqt", "Daraja", "Hostname", "IP manzil", "MAC manzil",
               "Tafsilot", "Ko'rilgan chora", "Xabar berildi",
               "MITRE Texnika ID", "MITRE Texnika nomi", "MITRE Taktika"]
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2634", end_color="1A2634", fill_type="solid")

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws_alerts.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(alert_rows, start=2):
        values = [
            row["id"], row["timestamp"], row["severity"], row["hostname"],
            row["ip_address"], row["mac_address"], row["reason"], row["action_taken"],
            row["notified"], row["mitre_technique_id"], row["mitre_technique_name"],
            row["mitre_tactic"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_alerts.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)

    widths = [6, 18, 10, 16, 14, 16, 45, 30, 12, 16, 26, 20]
    for i, w in enumerate(widths, start=1):
        ws_alerts.column_dimensions[get_column_letter(i)].width = w

    last_row = len(alert_rows) + 1
    has_data = last_row > 1

    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 16

    ws_summary["A1"] = "Xavfsizlik Monitoring Tizimi - Hisobot"
    ws_summary["A1"].font = Font(name="Arial", bold=True, size=14)
    ws_summary["A2"] = f"Davr: {summary['period_start']} — {summary['period_end']}"
    ws_summary["A2"].font = Font(name="Arial", italic=True, size=10)

    # MUHIM (xlsx skill talabiga ko'ra): raqamlar Python'da hisoblanib
    # yozilmaydi, balki Excel FORMULA orqali "Alerts" varag'idagi xom
    # ma'lumotdan avtomatik hisoblanadi.
    rows = [
        ("Jami alertlar (davr ichida)", f"=COUNTA(Alerts!A2:A{last_row})" if has_data else 0),
        ("  - Critical", f'=COUNTIF(Alerts!C2:C{last_row},"critical")' if has_data else 0),
        ("  - High", f'=COUNTIF(Alerts!C2:C{last_row},"high")' if has_data else 0),
        ("  - Medium", f'=COUNTIF(Alerts!C2:C{last_row},"medium")' if has_data else 0),
        ("  - Low", f'=COUNTIF(Alerts!C2:C{last_row},"low")' if has_data else 0),
        ("Xabar berilmagan alertlar", f'=COUNTIF(Alerts!I2:I{last_row},"yo\'q")' if has_data else 0),
        ("Zararli fayllar aniqlandi", summary["malicious_files_detected"]),
        ("Toza fayllar tekshirildi", summary["clean_files_scanned"]),
    ]

    start_row = 4
    for i, (label, value) in enumerate(rows):
        r = start_row + i
        ws_summary.cell(row=r, column=1, value=label).font = Font(name="Arial", size=11)
        ws_summary.cell(row=r, column=2, value=value).font = Font(name="Arial", size=11, bold=True)

    devices_start = start_row + len(rows) + 2
    ws_summary.cell(row=devices_start, column=1, value="Eng ko'p ta'sirlangan qurilmalar").font = \
        Font(name="Arial", bold=True, size=12)
    ws_summary.cell(row=devices_start + 1, column=1, value="Hostname").font = header_font
    ws_summary.cell(row=devices_start + 1, column=1).fill = header_fill
    ws_summary.cell(row=devices_start + 1, column=2, value="Alertlar soni").font = header_font
    ws_summary.cell(row=devices_start + 1, column=2).fill = header_fill

    for i, d in enumerate(summary["top_affected_devices"]):
        r = devices_start + 2 + i
        ws_summary.cell(row=r, column=1, value=d["hostname"]).font = Font(name="Arial", size=10)
        ws_summary.cell(row=r, column=2, value=d["alert_count"]).font = Font(name="Arial", size=10)

    wb.save(filepath)


def build_summary(session, since: datetime) -> dict:
    alerts = get_alerts_since(session, since)

    severity_counts = Counter(a.severity or "unknown" for a in alerts)
    mitre_counts = Counter(a.mitre_technique_id for a in alerts if a.mitre_technique_id)
    tactic_counts = Counter(a.mitre_tactic for a in alerts if a.mitre_tactic)

    device_alert_counts = Counter(a.device_id for a in alerts if a.device_id)
    top_devices = []
    for device_id, count in device_alert_counts.most_common(10):
        device = session.query(Device).filter(Device.id == device_id).first()
        top_devices.append({
            "hostname": device.hostname if device else "Nomalum",
            "ip_address": device.ip_address if device else "Nomalum",
            "alert_count": count,
        })

    malicious_files = (
        session.query(FileEvent)
        .filter(FileEvent.timestamp >= since, FileEvent.verdict == "malicious")
        .count()
    )
    clean_files = (
        session.query(FileEvent)
        .filter(FileEvent.timestamp >= since, FileEvent.verdict == "clean")
        .count()
    )

    return {
        "period_start": since.strftime("%Y-%m-%d %H:%M:%S"),
        "period_end": utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        "total_alerts": len(alerts),
        "severity_breakdown": dict(severity_counts),
        "mitre_technique_breakdown": dict(mitre_counts),
        "mitre_tactic_breakdown": dict(tactic_counts),
        "top_affected_devices": top_devices,
        "malicious_files_detected": malicious_files,
        "clean_files_scanned": clean_files,
        "unnotified_alerts": sum(1 for a in alerts if not a.notified),
    }


def generate_report(period_days: int, formats: list, output_dir: str) -> dict:
    """
    Asosiy funksiya - berilgan davr uchun hisobotlarni yaratadi.
    Qaytaradi: {"csv": path|None, "json": path|None, "pdf": path|None, "excel": path|None, "summary": dict}
    """
    os.makedirs(output_dir, exist_ok=True)
    since = utcnow() - timedelta(days=period_days)
    timestamp_tag = utcnow().strftime("%Y%m%d_%H%M%S")

    session = get_session()
    try:
        alerts = get_alerts_since(session, since)
        alert_rows = [_alert_row(session, a) for a in alerts]
        summary = build_summary(session, since)

        result = {"csv": None, "json": None, "pdf": None, "excel": None, "summary": summary}

        if "csv" in formats:
            csv_path = os.path.join(output_dir, f"alerts_report_{timestamp_tag}.csv")
            export_alerts_csv(alert_rows, csv_path)
            result["csv"] = csv_path

        if "json" in formats:
            json_path = os.path.join(output_dir, f"alerts_report_{timestamp_tag}.json")
            export_alerts_json({"summary": summary, "alerts": alert_rows}, json_path)
            result["json"] = json_path

        if "pdf" in formats:
            pdf_path = os.path.join(output_dir, f"alerts_report_{timestamp_tag}.pdf")
            export_summary_pdf(summary, alert_rows, pdf_path)
            result["pdf"] = pdf_path

        if "excel" in formats:
            excel_path = os.path.join(output_dir, f"alerts_report_{timestamp_tag}.xlsx")
            export_alerts_excel(summary, alert_rows, excel_path)
            result["excel"] = excel_path

        return result
    finally:
        session.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--period-days", type=int, default=7, help="Necha kunlik hisobot (standart: 7 - haftalik)")
    ap.add_argument("--format", default="csv,json", help="csv,json,pdf,excel (vergul bilan)")
    ap.add_argument("--output", default="./report_output", help="Chiqish papkasi")
    args = ap.parse_args()

    formats = [f.strip() for f in args.format.split(",")]
    result = generate_report(args.period_days, formats, args.output)

    print(f"Hisobot yaratildi ({args.period_days} kunlik davr uchun):")
    if result["csv"]:
        print(f"  CSV:   {result['csv']}")
    if result["json"]:
        print(f"  JSON:  {result['json']}")
    if result["pdf"]:
        print(f"  PDF:   {result['pdf']}")
    if result["excel"]:
        print(f"  Excel: {result['excel']}")
    print(f"\nUmumiy alertlar: {result['summary']['total_alerts']}")
    print(f"Daraja bo'yicha: {result['summary']['severity_breakdown']}")
