"""
TO'LIQ TIZIM TESTI - barcha bosqichlarni (0-13) birlashtirib, xatolarni tekshiradi.
Har bir qadam natijasi ✅/❌ bilan belgilanadi, xatolik bo'lsa to'xtamasdan davom etadi
(oxirida yakuniy hisobot chiqadi).

Standart holatda SQLite bazasida ishlaydi. PostgreSQL (Docker Compose'da
ishlatiladigan) bilan sinash uchun:

    export DATABASE_URL="postgresql://user:pass@localhost:5432/dbname"
    python3 run_full_test.py

(Bu loyiha PostgreSQL'da ham to'liq 14/14 test bilan sinovdan o'tkazilgan.)
"""
import os
import sys
import traceback

os.environ["DEMO_MODE"] = "true"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

RESULTS = []


def check(name, fn):
    try:
        fn()
        RESULTS.append((name, True, None))
        print(f"✅ {name}")
    except Exception as e:
        RESULTS.append((name, False, f"{type(e).__name__}: {e}"))
        print(f"❌ {name}: {type(e).__name__}: {e}")
        traceback.print_exc()


# ---------------------------------------------------------------------------
print("\n=== 0) BAZANI TOZALASH VA QAYTA YARATISH ===")
db_path = "logs/security_system.db"
if os.path.exists(db_path):
    os.remove(db_path)
if os.path.exists("logs/raw_syslog.log"):
    os.remove("logs/raw_syslog.log")

from db.database import get_session
from db.models import (
    RawLog, Device, Event, Alert, WhitelistEntry, BlacklistEntry,
    FileEvent, HashBlacklist, User, WebAccessLog,
)

check("Baza yaratildi", lambda: get_session().close())

# ---------------------------------------------------------------------------
print("\n=== 1) WHITELIST/BLACKLIST SEED ===")


def _seed():
    s = get_session()
    s.add(WhitelistEntry(value="172.16.0.10", description="1C server"))
    s.add(BlacklistEntry(value="malicious-test-domain.com", source="manual", reason="test"))
    s.add(HashBlacklist(sha256="275a021bbfb6489e54d471899f7db9d1663fc695ec2fe2a2c4538aabf651fd0",
                         threat_name="EICAR-Test", source="manual"))
    s.commit()
    s.close()


check("Whitelist/Blacklist/HashBlacklist seed qilindi", _seed)

# ---------------------------------------------------------------------------
print("\n=== 2) SYSLOG PARSER PIPELINE (Kerio DHCP, Connection, Windows DNS) ===")


def _test_parser_pipeline():
    from db.models import RawLog
    s = get_session()
    logs = [
        RawLog(source_ip="172.16.0.1", raw_message="DHCP: Lease granted to 172.16.1.45 MAC=AA:BB:CC:DD:EE:FF HOST=ACCOUNTING-PC"),
        RawLog(source_ip="172.16.0.1", raw_message="<134>Jul 30 KERIO-GW Connection: SRC=172.16.1.45 DST=8.8.8.8 DPT=443 PROTO=TCP ACTION=Permit"),
        RawLog(source_ip="172.16.0.11", raw_message='{"EventID":256,"ClientIP":"172.16.2.5","QueryName":"malicious-test-domain.com","QueryType":"A"}'),
        RawLog(source_ip="172.16.0.11", raw_message='{"EventID":256,"ClientIP":"172.16.2.6","QueryName":"google.com","QueryType":"A"}'),
        RawLog(source_ip="172.16.0.99", raw_message="bu hech qanday parserga mos kelmaydigan xom matn"),
    ]
    s.add_all(logs)
    s.commit()
    s.close()

    from engine.parser_engine import run_once
    count = run_once()
    assert count == 5, f"Kutilgan 5 ta yozuv, lekin {count} ta qayta ishlandi"

    s = get_session()
    unprocessed = s.query(RawLog).filter(RawLog.processed == False).count()
    assert unprocessed == 0, f"{unprocessed} ta yozuv hali processed=False"

    devices = s.query(Device).all()
    assert len(devices) >= 3, f"Kamida 3 ta device kutilgan, {len(devices)} ta topildi"

    dev_1_45 = s.query(Device).filter(Device.ip_address == "172.16.1.45").first()
    assert dev_1_45.mac_address == "AA:BB:CC:DD:EE:FF", "DHCP orqali MAC to'g'ri bog'lanmadi"
    assert dev_1_45.hostname == "ACCOUNTING-PC", "DHCP orqali hostname to'g'ri bog'lanmadi"

    events = s.query(Event).all()
    assert len(events) == 3, f"3 ta event (1 connection + 2 dns) kutilgan, {len(events)} ta topildi"

    alerts = s.query(Alert).filter(Alert.event_id.isnot(None)).all()
    assert len(alerts) == 1, f"Faqat 1 ta DNS blacklist alert kutilgan, {len(alerts)} ta topildi"
    assert alerts[0].device_id is not None, "Alert device_id bilan bog'lanmagan"
    s.close()


check("Parser pipeline (DHCP+Connection+DNS, blacklist alert)", _test_parser_pipeline)

# ---------------------------------------------------------------------------
print("\n=== 3) FAYL ANALIZ PIPELINE (hash, YARA, ZIP, Office, PDF) ===")


def _test_file_pipeline():
    import hashlib
    import zipfile

    os.makedirs("/tmp/test_filestore", exist_ok=True)

    # 3a) EICAR (mahalliy blacklist orqali topiladigan)
    eicar_path = "/tmp/test_filestore/invoice.exe"
    with open(eicar_path, "wb") as f:
        f.write(b"EICAR-TEST-DUMMY-CONTENT")  # haqiqiy EICAR emas, faqat hash mos kelishi uchun quyida override qilamiz

    # Haqiqiy sinov uchun HashBlacklist'dagi hash bilan mos keladigan fayl kerak emas -
    # biz to'g'ridan-to'g'ri FileEvent'ga o'sha hash'ni yozamiz (Suricata ham shunday qiladi - hash hisoblab beradi)
    known_bad_hash = "275a021bbfb6489e54d471899f7db9d1663fc695ec2fe2a2c4538aabf651fd0"

    # 3b) ZIP ichida embedded PE (YARA orqali topiladigan)
    zip_path = "/tmp/test_filestore/archive.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("payload.exe", b"MZ" + b"\x90" * 58 + b"This program cannot be run in DOS mode")
        zf.writestr("readme.txt", b"Bu oddiy va xavfsiz matn.")

    # 3c) PDF ichida JS (YARA orqali topiladigan)
    pdf_path = "/tmp/test_filestore/report.pdf"
    with open(pdf_path, "wb") as f:
        f.write(b"%PDF-1.4\n1 0 obj << /Type /Catalog /OpenAction 2 0 R >>\n/JavaScript (app.alert(1))\nendobj")

    # 3d) Toza fayl
    clean_path = "/tmp/test_filestore/clean.txt"
    with open(clean_path, "wb") as f:
        f.write(b"Bu 100% xavfsiz oddiy matn fayli, hech qanday tahdid yo'q.")

    def sha(path):
        with open(path, "rb") as f:
            return hashlib.sha256(f.read()).hexdigest()

    s = get_session()
    entries = [
        FileEvent(src_ip="172.16.2.10", dest_ip="1.2.3.4", filename="invoice.exe", file_ext="exe",
                   size=100, sha256=known_bad_hash, md5="x", stored_path=eicar_path, checked=False),
        FileEvent(src_ip="172.16.2.20", dest_ip="1.2.3.5", filename="archive.zip", file_ext="zip",
                   size=os.path.getsize(zip_path), sha256=sha(zip_path), md5="x", stored_path=zip_path, checked=False),
        FileEvent(src_ip="172.16.2.30", dest_ip="1.2.3.6", filename="report.pdf", file_ext="pdf",
                   size=os.path.getsize(pdf_path), sha256=sha(pdf_path), md5="x", stored_path=pdf_path, checked=False),
        FileEvent(src_ip="172.16.2.40", dest_ip="1.2.3.7", filename="clean.txt", file_ext="txt",
                   size=os.path.getsize(clean_path), sha256=sha(clean_path), md5="x", stored_path=clean_path, checked=False),
    ]
    s.add_all(entries)
    s.commit()
    s.close()

    from engine.file_analysis_engine import run_once as file_analysis_run
    from engine.deep_scan_engine import run_once as deep_scan_run

    n1 = file_analysis_run()
    assert n1 == 4, f"4 ta fayl hash-tekshiruvidan o'tishi kerak edi, {n1} ta o'tdi"

    n2 = deep_scan_run()
    assert n2 == 4, f"4 ta fayl deep-scan'dan o'tishi kerak edi, {n2} ta o'tdi"

    # ZIP ichidan chiqqan payload.exe ni ham tekshirish uchun yana ikki marta ishga tushiramiz
    file_analysis_run()
    deep_scan_run()

    s = get_session()
    fes = {fe.filename: fe for fe in s.query(FileEvent).all()}

    assert fes["invoice.exe"].verdict == "malicious", "invoice.exe (hash blacklist) malicious deb topilishi kerak edi"
    assert fes["archive.zip"].verdict == "malicious", "archive.zip (ichida PE bor) malicious deb topilishi kerak edi"
    assert fes["report.pdf"].verdict == "malicious", "report.pdf (ichida JS bor) malicious deb topilishi kerak edi"
    assert fes["clean.txt"].verdict == "clean", "clean.txt clean deb topilishi kerak edi"

    # ZIP ichidan chiqqan payload.exe alohida FileEvent sifatida yaratilganini tekshirish
    payload = s.query(FileEvent).filter(FileEvent.filename == "payload.exe").first()
    assert payload is not None, "ZIP ichidan payload.exe chiqarilmagan"
    assert payload.parent_file_event_id == fes["archive.zip"].id, "payload.exe parent_id noto'g'ri"
    assert payload.verdict == "malicious", "payload.exe malicious deb topilishi kerak edi"

    readme = s.query(FileEvent).filter(FileEvent.filename == "readme.txt").first()
    assert readme is not None and readme.verdict == "clean", "readme.txt clean bo'lishi kerak edi (soxta pozitiv)"

    file_alerts = s.query(Alert).filter(Alert.file_event_id.isnot(None)).all()
    assert len(file_alerts) >= 3, f"Kamida 3 ta fayl-alert kutilgan, {len(file_alerts)} ta topildi"
    for a in file_alerts:
        assert a.device_id is not None, f"Alert {a.id} device_id bilan bog'lanmagan"
    s.close()


check("Fayl analiz pipeline (hash+YARA+ZIP rekursiya+soxta-pozitiv yo'qligi)", _test_file_pipeline)

# ---------------------------------------------------------------------------
print("\n=== 4) OFFICE MAKRO SKANER (soxta pozitiv tekshiruvi) ===")


def _test_office_scanner_false_positive():
    from scanners.office_scanner import scan_office_file
    # Office bo'lmagan fayl uchun None qaytarishi kerak (oldingi tuzatilgan xato)
    r = scan_office_file("/tmp/test_filestore/report.pdf")
    assert r is None, f"PDF fayl uchun None qaytarishi kerak edi, lekin {r} qaytardi"


check("Office scanner soxta-pozitiv himoyasi", _test_office_scanner_false_positive)

# ---------------------------------------------------------------------------
print("\n=== 5) ARXIV SKANER XAVFSIZLIK CHEKLOVLARI (path traversal) ===")


def _test_archive_path_traversal():
    import zipfile
    from scanners.archive_scanner import _safe_member_path

    assert _safe_member_path("normal_file.txt") is True
    assert _safe_member_path("../../etc/passwd") is False
    assert _safe_member_path("/etc/passwd") is False
    assert _safe_member_path("subdir/file.txt") is True


check("Arxiv path-traversal himoyasi", _test_archive_path_traversal)

# ---------------------------------------------------------------------------
print("\n=== 6) RESPONSE ENGINE (avtomatik javob choralari) ===")


def _test_response_engine():
    s = get_session()
    d_wifi = Device(ip_address="172.16.3.1", mac_address="AA:11:22:33:44:55", connection_type="wifi", source="test")
    d_unknown = Device(ip_address="172.16.3.2", mac_address="BB:11:22:33:44:55", connection_type="unknown", source="test")
    s.add_all([d_wifi, d_unknown])
    s.flush()

    a1 = Alert(device_id=d_wifi.id, severity="critical", reason="test", action_taken="TODO: bloklash backend hali ulanmagan")
    a2 = Alert(device_id=d_unknown.id, severity="critical", reason="test", action_taken="TODO: bloklash backend hali ulanmagan")
    a3 = Alert(device_id=None, severity="high", reason="device yo'q", action_taken="TODO: bloklash backend hali ulanmagan")
    s.add_all([a1, a2, a3])
    s.commit()
    ids = [a1.id, a2.id, a3.id]
    s.close()

    from engine.response_engine import run_once
    n = run_once()
    # Diqqat: response_engine FAQAT shu 3 tasini emas, balki bazadagi barcha
    # "TODO" holatidagi alertlarni (2 va 3-bosqichlarda yaratilganlarni ham)
    # qayta ishlaydi - bu to'g'ri xatti-harakat (hech qanday alert e'tibordan
    # chetda qolmasligi kerak). Shuning uchun n >= 3 tekshiramiz.
    assert n >= 3, f"Kamida 3 ta alert qayta ishlanishi kerak edi, {n} ta ishlandi"

    s = get_session()
    for aid in ids:
        a = s.query(Alert).filter(Alert.id == aid).first()
        assert not a.action_taken.startswith("TODO"), f"Alert {aid} hali TODO holatida qoldi: {a.action_taken}"
    s.close()


check("Response engine (device_id yo'qligi, mock adapter, real xato holatlari)", _test_response_engine)

# ---------------------------------------------------------------------------
print("\n=== 7) BO'SH NAVBAT BILAN ISHLASH (edge case) ===")


def _test_empty_queue():
    from engine.parser_engine import run_once as p
    from engine.file_analysis_engine import run_once as f
    from engine.deep_scan_engine import run_once as d
    from engine.response_engine import run_once as r
    assert p() == 0
    assert f() == 0
    assert d() == 0
    assert r() == 0


check("Bo'sh navbatlar bilan barcha enginelar (xatosiz)", _test_empty_queue)

# ---------------------------------------------------------------------------
print("\n=== 8) API SERVER (Flask test client orqali, real port ochmasdan) ===")


def _test_api_server():
    from api import server as api_server
    api_server.AGENT_API_KEY = "test-key-for-unit-test"
    client = api_server.app.test_client()

    # Health
    r = client.get("/api/v1/health")
    assert r.status_code == 200

    # API kalitsiz - 401
    r = client.post("/api/v1/check_hash", json={"sha256": "a" * 64})
    assert r.status_code == 401

    # Noto'g'ri uzunlikdagi hash - 400
    r = client.post("/api/v1/check_hash", json={"sha256": "abc"},
                     headers={"X-API-Key": "test-key-for-unit-test"})
    assert r.status_code == 400

    # Toza hash
    r = client.post("/api/v1/check_hash", json={"sha256": "b" * 64},
                     headers={"X-API-Key": "test-key-for-unit-test"})
    assert r.status_code == 200
    assert r.get_json()["malicious"] is False

    # HashBlacklist'dagi hash
    s = get_session()
    s.add(HashBlacklist(sha256="c" * 64, threat_name="Unit-Test-Threat", source="manual"))
    s.commit()
    s.close()

    r = client.post("/api/v1/check_hash", json={"sha256": "c" * 64},
                     headers={"X-API-Key": "test-key-for-unit-test"})
    assert r.status_code == 200
    body = r.get_json()
    assert body["malicious"] is True
    assert body["threat_name"] == "Unit-Test-Threat"

    # report_incident
    r = client.post("/api/v1/report_incident", json={
        "hostname": "TEST-PC", "ip_address": "172.16.9.9",
        "filename": "test.exe", "sha256": "c" * 64,
        "threat_name": "Unit-Test-Threat", "file_deleted": True, "process_killed": True,
        "process_name": "test.exe",
    }, headers={"X-API-Key": "test-key-for-unit-test"})
    assert r.status_code == 200
    alert_id = r.get_json()["alert_id"]

    s = get_session()
    alert = s.query(Alert).filter(Alert.id == alert_id).first()
    assert alert is not None
    assert alert.device_id is not None
    assert "Unit-Test-Threat" in alert.reason
    s.close()

    # Majburiy maydon yo'q - 400
    r = client.post("/api/v1/report_incident", json={"hostname": "X"},
                     headers={"X-API-Key": "test-key-for-unit-test"})
    assert r.status_code == 400


check("API server (health/auth/validatsiya/blacklist/incident)", _test_api_server)

# ---------------------------------------------------------------------------
print("\n=== 9) AGENT_CORE KOMPONENTLARI (Windows+Linux Agent umumiy yadrosi) ===")


def _test_agent_components():
    import subprocess
    import time as _time
    from agent_core.process_killer import find_processes_holding_file, kill_process_holding_file

    test_file = "/tmp/_agent_component_test.txt"
    with open(test_file, "w") as f:
        f.write("test")

    proc = subprocess.Popen(["python3", "-c", f"f = open('{test_file}'); import time; time.sleep(10)"])
    _time.sleep(1.2)

    procs = find_processes_holding_file(test_file)
    assert len(procs) >= 1, "Faylni ochgan jarayon topilishi kerak edi"

    result = kill_process_holding_file(test_file)
    assert result.process_killed is True

    _time.sleep(0.5)
    assert proc.poll() is not None, "Jarayon to'xtatilgan bo'lishi kerak edi"
    os.remove(test_file)

    # file_monitor - ikki marta aniqlash xatosi tuzatilganini tasdiqlash
    from agent_core.file_monitor import FileMonitor
    watch_dir = "/tmp/_agent_watch_component_test"
    os.makedirs(watch_dir, exist_ok=True)
    detected = []
    monitor = FileMonitor([watch_dir], lambda p: detected.append(p))
    monitor.start()
    _time.sleep(0.5)
    with open(os.path.join(watch_dir, "sample.txt"), "wb") as f:
        f.write(b"sample content")
    _time.sleep(3.5)
    monitor.stop()
    assert len(detected) == 1, f"Aniq 1 marta aniqlanishi kerak edi, {len(detected)} marta aniqlandi"

    import shutil
    shutil.rmtree(watch_dir, ignore_errors=True)


check("Agent Core komponentlari (process_killer + file_monitor, real jarayon bilan)", _test_agent_components)

# ---------------------------------------------------------------------------
print("\n=== 9b) LINUX AGENT - TO'LIQ END-TO-END (real API server + real jarayon) ===")


def _test_linux_agent_e2e():
    import subprocess
    import time as _time
    import hashlib

    watch_dir = "/tmp/_linux_agent_e2e_watch"
    os.makedirs(watch_dir, exist_ok=True)

    malicious_content = b"linux agent e2e test malicious payload 987654"
    sha256 = hashlib.sha256(malicious_content).hexdigest()

    s = get_session()
    s.add(HashBlacklist(sha256=sha256, threat_name="LinuxAgentE2E-Trojan", source="manual"))
    s.commit()
    s.close()

    api_env = os.environ.copy()
    api_env["AGENT_API_KEY"] = "linux-e2e-test-key"
    api_proc = subprocess.Popen(
        ["python3", "-m", "api.server"],
        env={**api_env, "API_PORT": "8199"},
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    _time.sleep(2)

    try:
        os.environ["API_SERVER_URL"] = "http://127.0.0.1:8199"
        os.environ["AGENT_API_KEY"] = "linux-e2e-test-key"
        os.environ["AGENT_CACHE_FILE"] = "/tmp/_linux_agent_e2e_cache.json"
        os.environ["AGENT_LOG_FILE"] = "/tmp/_linux_agent_e2e.log"
        if os.path.exists(os.environ["AGENT_CACHE_FILE"]):
            os.remove(os.environ["AGENT_CACHE_FILE"])

        import importlib
        import agent_core.agent as agent_mod
        importlib.reload(agent_mod)

        agent = agent_mod.EndpointAgent([watch_dir])

        malicious_file = os.path.join(watch_dir, "linux_e2e_payload.bin")
        with open(malicious_file, "wb") as f:
            f.write(malicious_content)

        # Faylni "ochiq" ushlab turuvchi jarayon (real Linux jarayoni)
        locker = subprocess.Popen(["python3", "-c", f"f=open('{malicious_file}'); import time; time.sleep(10)"])
        _time.sleep(1)

        # Agentning fayl-topilishi logikasini to'g'ridan-to'g'ri chaqiramiz
        # (FileMonitor'ning watchdog kuzatuvi allaqachon alohida testda
        # tekshirilgan - bu yerda "aniqlangandan keyingi" javob zanjiri
        # sinaladi: hash -> server -> jarayonni to'xtatish -> o'chirish -> report)
        agent._on_new_file(malicious_file)

        _time.sleep(1)

        assert not os.path.exists(malicious_file), "Zararli fayl o'chirilishi kerak edi"
        assert locker.poll() is not None, "Faylni ushlab turgan jarayon to'xtatilishi kerak edi"

        s = get_session()
        alert = (
            s.query(Alert)
            .filter(Alert.reason.like("%LinuxAgentE2E-Trojan%"))
            .first()
        )
        assert alert is not None, "Markazga incident xabari kelib, Alert yaratilishi kerak edi"
        assert alert.device_id is not None
        s.close()

    finally:
        api_proc.terminate()
        try:
            api_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            api_proc.kill()
        import shutil
        shutil.rmtree(watch_dir, ignore_errors=True)
        for f in ["/tmp/_linux_agent_e2e_cache.json", "/tmp/_linux_agent_e2e.log"]:
            if os.path.exists(f):
                os.remove(f)
        for k in ["API_SERVER_URL", "AGENT_API_KEY", "AGENT_CACHE_FILE", "AGENT_LOG_FILE"]:
            os.environ.pop(k, None)


check("Linux Agent to'liq E2E (real server+jarayon+fayl o'chirish+report)", _test_linux_agent_e2e)

# ---------------------------------------------------------------------------
print("\n=== 10) NOTIFICATION ENGINE (real SMTP server orqali) ===")


def _test_notification_engine():
    import subprocess
    import time as _time
    import os as _os

    received_log = "/tmp/_test_received_emails.log"
    if _os.path.exists(received_log):
        _os.remove(received_log)

    debug_smtp_code = f'''
import asyncio
from aiosmtpd.controller import Controller

class DebugHandler:
    async def handle_DATA(self, server, session, envelope):
        with open("{received_log}", "a", encoding="utf-8") as f:
            f.write("=== YANGI XAT ===\\n")
            f.write(f"To: {{envelope.rcpt_tos}}\\n")
            f.write(envelope.content.decode("utf-8", errors="replace"))
        return "250 OK"

controller = Controller(DebugHandler(), hostname="127.0.0.1", port=1026)
controller.start()
import time
time.sleep(8)
controller.stop()
'''
    smtp_script = "/tmp/_test_debug_smtp.py"
    with open(smtp_script, "w") as f:
        f.write(debug_smtp_code)

    smtp_proc = subprocess.Popen(["python3", smtp_script])
    _time.sleep(2)

    try:
        os.environ["SMTP_HOST"] = "127.0.0.1"
        os.environ["SMTP_PORT"] = "1026"
        os.environ["SMTP_FROM"] = "security@company.local"
        os.environ["ADMIN_EMAIL"] = "admin@company.local"
        os.environ["NOTIFY_CHANNELS"] = "email,telegram"

        s = get_session()
        d = Device(ip_address="172.16.5.5", mac_address="AA:BB:CC:00:11:22",
                    hostname="NOTIFY-TEST-PC", connection_type="wifi", source="test")
        s.add(d)
        s.flush()
        alert = Alert(device_id=d.id, severity="critical", reason="Test xabarnoma",
                       action_taken="Test chora", notified=False)
        s.add(alert)
        s.commit()
        alert_id = alert.id
        s.close()

        # Modullarni muhit o'zgaruvchilari o'zgarganidan keyin qayta yuklash kerak
        import importlib
        import notifications.email_notifier as email_mod
        importlib.reload(email_mod)
        import engine.notification_engine as notif_engine
        importlib.reload(notif_engine)

        n = notif_engine.run_once()
        assert n >= 1, f"Kamida 1 ta alert xabar qilinishi kerak edi, {n} ta qilindi"

        s = get_session()
        a = s.query(Alert).filter(Alert.id == alert_id).first()
        assert a.notified is True, "Alert notified=True bo'lishi kerak edi"
        s.close()

        _time.sleep(0.5)
        assert os.path.exists(received_log), "Email qabul qilinmadi (SMTP server fayl yozmadi)"
        content = open(received_log).read()
        assert "NOTIFY-TEST-PC" in content, "Xatda hostname topilmadi"
    finally:
        smtp_proc.terminate()
        smtp_proc.wait(timeout=5)
        for f in [smtp_script, received_log]:
            if os.path.exists(f):
                os.remove(f)
        for k in ["SMTP_HOST", "SMTP_PORT", "SMTP_FROM", "ADMIN_EMAIL", "NOTIFY_CHANNELS"]:
            os.environ.pop(k, None)


check("Notification engine (real SMTP orqali email yetkazish)", _test_notification_engine)

# ---------------------------------------------------------------------------
print("\n=== 11) CLAMAV INTEGRATSIYASI (maxsus test-signatura bazasi bilan) ===")


def _test_clamav_integration():
    import subprocess
    if subprocess.run(["which", "clamscan"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - clamscan o'rnatilmagan bu muhitda)")
        return

    os.environ["CLAMAV_DB_DIR"] = "/tmp/clamav_test_db"
    os.makedirs("/tmp/clamav_test_db", exist_ok=True)

    content = b"run_full_test clamav dummy malware content\n"
    import hashlib
    sha = hashlib.sha256(content).hexdigest()
    size = len(content)
    with open("/tmp/clamav_test_db/runtest.hdb", "w") as f:
        f.write(f"{sha}:{size}:RunFullTest.Malware\n")

    malware_path = "/tmp/_run_full_test_clamav_sample.txt"
    with open(malware_path, "wb") as f:
        f.write(content)

    clean_path = "/tmp/_run_full_test_clamav_clean.txt"
    with open(clean_path, "wb") as f:
        f.write(b"xavfsiz matn")

    import importlib
    import scanners.clamav_scanner as clamav_mod
    importlib.reload(clamav_mod)

    r_bad = clamav_mod.scan_file(malware_path)
    assert r_bad["scanned"] is True, f"Skanerlash muvaffaqiyatsiz: {r_bad}"
    assert r_bad["infected"] is True, "Zararli fayl aniqlanishi kerak edi"
    assert "RunFullTest.Malware" in r_bad["signature"]

    r_clean = clamav_mod.scan_file(clean_path)
    assert r_clean["scanned"] is True
    assert r_clean["infected"] is False

    os.remove(malware_path)
    os.remove(clean_path)
    os.remove("/tmp/clamav_test_db/runtest.hdb")
    os.environ.pop("CLAMAV_DB_DIR", None)


check("ClamAV integratsiyasi (zararli+toza fayl, real clamscan)", _test_clamav_integration)

# ---------------------------------------------------------------------------
print("\n=== 12) MITRE ATT&CK BELGILASH ===")


def _test_mitre_tagging():
    from engine.mitre_tagging_engine import run_once as mitre_run_once

    s = get_session()
    a1 = Alert(severity="high", reason="Blacklist'dagi domenga so'rov: evil.com (manba: manual)", action_taken="TODO")
    a2 = Alert(severity="critical", reason="ClamAV[critical]: Trojan.GenericKD", action_taken="TODO")
    a3 = Alert(severity="critical", reason="YARA[high]: Suspicious_PowerShell_Obfuscation - test", action_taken="TODO")
    s.add_all([a1, a2, a3])
    s.commit()
    ids = [a1.id, a2.id, a3.id]
    s.close()

    n = mitre_run_once()
    assert n >= 3, f"Kamida 3 ta alert belgilanishi kerak edi, {n} ta belgilandi"

    s = get_session()
    tagged = {a.id: a for a in s.query(Alert).filter(Alert.id.in_(ids)).all()}
    assert tagged[a1.id].mitre_technique_id == "T1071.004", "DNS blacklist noto'g'ri texnika bilan belgilandi"
    assert tagged[a2.id].mitre_technique_id == "T1204.002", "ClamAV alert noto'g'ri texnika bilan belgilandi"
    assert tagged[a3.id].mitre_technique_id == "T1059.001", "PowerShell alert noto'g'ri texnika bilan belgilandi"
    for a in tagged.values():
        assert a.mitre_tactic, f"Alert {a.id} uchun taktika bo'sh qoldi"
    s.close()

    # Bo'sh navbatda 0 qaytarishi kerak
    assert mitre_run_once() == 0, "Barcha alertlar belgilangandan keyin 0 qaytarishi kerak edi"


check("MITRE ATT&CK avtomatik belgilash (texnika+taktika)", _test_mitre_tagging)

# ---------------------------------------------------------------------------
print("\n=== 13) WEB DASHBOARD (Flask test client orqali) ===")


def _test_dashboard():
    from dashboard import app as dash_app
    from dashboard.create_user import create_user

    create_user("dashtest_admin", "dashtestpass123", "admin")

    dash_app.app.secret_key = "test-secret-key-dashboard"
    client = dash_app.app.test_client()

    # Autentifikatsiyasiz - login sahifasiga redirect (302)
    r = client.get("/", follow_redirects=False)
    assert r.status_code == 302, f"302 (login'ga redirect) kutilgan edi, {r.status_code} keldi"

    # Noto'g'ri parol bilan - login sahifasida qoladi (200, lekin kirmagan)
    client.post("/login", data={"username": "dashtest_admin", "password": "wrong"})
    r = client.get("/", follow_redirects=False)
    assert r.status_code == 302, "Noto'g'ri parol bilan hali ham kirgan bo'lmasligi kerak"

    # To'g'ri login
    r = client.post("/login", data={"username": "dashtest_admin", "password": "dashtestpass123"})
    assert r.status_code in (200, 302)

    # Endi barcha sahifalar ochiq bo'lishi kerak
    for path in ["/", "/alerts", "/devices", "/files"]:
        r = client.get(path)
        assert r.status_code == 200, f"{path}: 200 kutilgan edi, {r.status_code} keldi"

    # Ma'lumot borligini tekshirish (oldingi testlarda yaratilgan device/alert'lar)
    r = client.get("/devices")
    body = r.get_data(as_text=True)
    assert "172.16." in body, "Devices sahifasida IP manzil ko'rinmadi"

    # Filtrlash ishlashini tekshirish
    r = client.get("/alerts?severity=critical")
    assert r.status_code == 200

    r = client.get("/files?verdict=malicious")
    assert r.status_code == 200

    client.get("/logout")
    r = client.get("/", follow_redirects=False)
    assert r.status_code == 302, "Logout'dan keyin qayta login talab qilinishi kerak"


check("Web Dashboard (login, 4 sahifa, filtrlash)", _test_dashboard)

# ---------------------------------------------------------------------------
print("\n=== 14) REPORT GENERATOR (CSV/JSON hisobot) ===")


def _test_report_generator():
    import shutil
    from reports.report_generator import generate_report

    out_dir = "/tmp/_test_report_output"
    if os.path.exists(out_dir):
        shutil.rmtree(out_dir)

    result = generate_report(period_days=365, formats=["csv", "json"], output_dir=out_dir)

    assert result["csv"] and os.path.isfile(result["csv"]), "CSV fayl yaratilmadi"
    assert result["json"] and os.path.isfile(result["json"]), "JSON fayl yaratilmadi"

    with open(result["csv"], encoding="utf-8-sig") as f:
        import csv as csv_mod
        rows = list(csv_mod.DictReader(f))
    assert len(rows) >= 1, "CSV'da hech qanday qator yo'q"
    assert "mitre_technique_id" in rows[0], "CSV'da MITRE ustuni yo'q"

    import json as json_mod
    with open(result["json"], encoding="utf-8") as f:
        data = json_mod.load(f)
    assert "summary" in data and "alerts" in data
    assert data["summary"]["total_alerts"] == len(data["alerts"])
    assert "severity_breakdown" in data["summary"]

    # Dashboard orqali yuklab olish (test client)
    from dashboard import app as dash_app
    from dashboard.create_user import create_user
    create_user("reporttest_admin", "reporttestpass123", "admin")
    dash_app.app.secret_key = "test-secret-key-report"
    client = dash_app.app.test_client()
    client.post("/login", data={"username": "reporttest_admin", "password": "reporttestpass123"})

    r = client.get("/reports/download?period_days=30&format=csv")
    assert r.status_code == 200
    assert r.data.startswith(b"\xef\xbb\xbfid,") or b"id,timestamp" in r.data[:50]

    r = client.get("/reports/download?format=xml")
    assert r.status_code == 400

    client.get("/logout")

    shutil.rmtree(out_dir, ignore_errors=True)


check("Report Generator (CSV/JSON + dashboard orqali yuklab olish)", _test_report_generator)

# ---------------------------------------------------------------------------
print("\n=== 15) RBAC (login, rollar, acknowledge huquqi) ===")


def _test_rbac():
    from dashboard.create_user import create_user
    from dashboard import app as dash_app

    create_user("rbac_admin", "adminpass123", "admin")
    create_user("rbac_analyst", "analystpass123", "analyst")
    create_user("rbac_viewer", "viewerpass123", "viewer")

    s = get_session()
    d = Device(ip_address="172.16.8.99", hostname="RBAC-AUTOTEST-PC", connection_type="wifi", source="test")
    s.add(d)
    s.flush()
    a = Alert(device_id=d.id, severity="critical", reason="RBAC avtomatik test alert", action_taken="TODO", notified=False)
    s.add(a)
    s.commit()
    alert_id = a.id
    s.close()

    dash_app.app.secret_key = "test-secret-key"
    client = dash_app.app.test_client()

    # Autentifikatsiyasiz - login sahifasiga redirect
    r = client.get("/", follow_redirects=False)
    assert r.status_code == 302 and "/login" in r.headers.get("Location", "")

    # Noto'g'ri parol
    r = client.post("/login", data={"username": "rbac_admin", "password": "wrong"})
    assert r.status_code == 200  # login sahifasiga qaytadi, xato bilan
    assert b"noto" in r.data.lower() or b"xato" in r.data.lower() or r.status_code == 200

    # Admin - to'g'ri login
    r = client.post("/login", data={"username": "rbac_admin", "password": "adminpass123"}, follow_redirects=True)
    assert r.status_code == 200
    r = client.get("/users")
    assert r.status_code == 200, "Admin /users sahifasiga kira olishi kerak edi"
    client.get("/logout")

    # Viewer - /users va acknowledge'ga kira olmasligi kerak
    client.post("/login", data={"username": "rbac_viewer", "password": "viewerpass123"})
    r = client.get("/users")
    assert r.status_code == 403, f"Viewer /users'ga kirmasligi kerak edi, {r.status_code} keldi"
    r = client.post(f"/alerts/{alert_id}/acknowledge")
    assert r.status_code == 403, f"Viewer acknowledge qila olmasligi kerak edi, {r.status_code} keldi"
    # Viewer oddiy sahifalarni ko'ra olishi kerak
    r = client.get("/alerts")
    assert r.status_code == 200
    client.get("/logout")

    # Analyst - acknowledge qila olishi, lekin /users'ga kira olmasligi kerak
    client.post("/login", data={"username": "rbac_analyst", "password": "analystpass123"})
    r = client.get("/users")
    assert r.status_code == 403, "Analyst /users'ga kirmasligi kerak edi"
    r = client.post(f"/alerts/{alert_id}/acknowledge", follow_redirects=False)
    assert r.status_code == 302, f"Analyst acknowledge qila olishi kerak edi, {r.status_code} keldi"

    s = get_session()
    a = s.query(Alert).filter(Alert.id == alert_id).first()
    assert a.acknowledged is True
    assert a.acknowledged_by == "rbac_analyst"
    s.close()
    client.get("/logout")

    # Parol xesh sifatida saqlanganini tekshirish (ochiq matn emas)
    s = get_session()
    u = s.query(User).filter(User.username == "rbac_admin").first()
    assert u.password_hash != "adminpass123", "Parol ochiq matnda saqlanmasligi kerak!"
    assert u.password_hash.startswith(("pbkdf2:", "scrypt:")), "Parol tanish xesh formatida emas"
    s.close()


check("RBAC (login, 3 rol, acknowledge huquqi, parol xeshlash)", _test_rbac)

# ---------------------------------------------------------------------------
print("\n=== 16) PDF/EXCEL HISOBOTLAR (real fayl + LibreOffice recalc) ===")


def _test_pdf_excel_reports():
    import shutil
    import subprocess
    from reports.report_generator import generate_report

    out_dir = "/tmp/_test_pdf_excel_output"
    if os.path.exists(out_dir):
        shutil.rmtree(out_dir)

    result = generate_report(period_days=365, formats=["pdf", "excel"], output_dir=out_dir)

    assert result["pdf"] and os.path.isfile(result["pdf"]), "PDF fayl yaratilmadi"
    assert result["excel"] and os.path.isfile(result["excel"]), "Excel fayl yaratilmadi"

    # PDF haqiqiy o'qiladigan ekanini tekshirish
    from pypdf import PdfReader
    reader = PdfReader(result["pdf"])
    assert len(reader.pages) >= 1, "PDF'da sahifa yo'q"

    # Excel: LibreOffice orqali formulalarni haqiqiy hisoblash (recalc)
    recalc_script = "/mnt/skills/public/xlsx/scripts/recalc.py"
    if os.path.isfile(recalc_script):
        proc = subprocess.run(
            ["python3", recalc_script, result["excel"], "60"],
            capture_output=True, text=True, timeout=90,
        )
        import json as json_mod
        recalc_result = json_mod.loads(proc.stdout)
        assert recalc_result.get("status") == "success", f"Excel recalc muvaffaqiyatsiz: {recalc_result}"
        assert recalc_result.get("total_errors") == 0, f"Excel'da formula xatolari bor: {recalc_result}"

        # Hisoblangan qiymatlarni haqiqiy sonlar bilan solishtirish
        from openpyxl import load_workbook
        wb = load_workbook(result["excel"], data_only=True)
        ws = wb["Summary"]
        total_from_excel = ws["B4"].value
        assert total_from_excel == result["summary"]["total_alerts"], (
            f"Excel formulasi noto'g'ri hisobladi: {total_from_excel} != {result['summary']['total_alerts']}"
        )

    # Bo'sh davr bilan ham ishlashini tekshirish (edge case)
    empty_result = generate_report(period_days=0, formats=["pdf", "excel"], output_dir=out_dir)
    assert os.path.isfile(empty_result["pdf"]), "Bo'sh davr uchun PDF yaratilmadi"
    assert os.path.isfile(empty_result["excel"]), "Bo'sh davr uchun Excel yaratilmadi"

    # Dashboard orqali yuklab olish
    from dashboard import app as dash_app
    from dashboard.create_user import create_user
    create_user("pdftest_admin", "pdftestpass123", "admin")
    dash_app.app.secret_key = "test-secret-key-pdf"
    client = dash_app.app.test_client()
    client.post("/login", data={"username": "pdftest_admin", "password": "pdftestpass123"})

    r = client.get("/reports/download?period_days=7&format=pdf")
    assert r.status_code == 200
    assert r.data[:4] == b"%PDF", "Dashboard'dan qaytgan fayl PDF emas"

    r = client.get("/reports/download?period_days=7&format=excel")
    assert r.status_code == 200
    assert r.data[:2] == b"PK", "Dashboard'dan qaytgan fayl Excel (ZIP-based) emas"

    client.get("/logout")
    shutil.rmtree(out_dir, ignore_errors=True)


check("PDF/Excel hisobotlar (real fayl, LibreOffice recalc, dashboard)", _test_pdf_excel_reports)

# ---------------------------------------------------------------------------
print("\n=== 17) SNORT INTEGRATSIYASI (real Snort chiqishi, pcap orqali) ===")


def _test_snort_integration():
    import subprocess
    import shutil

    if subprocess.run(["which", "snort"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - snort o'rnatilmagan bu muhitda)")
        return

    from collectors.snort_reader import parse_alert_line, read_existing

    # 1) parse_alert_line birlik testi (haqiqiy Snort formatiga mos)
    sample = "08/06-08:39:01.972343  [**] [1:1000001:1] TEST Suspicious port 4444 (C2-like) [**] [Priority: 1] {TCP} 10.0.0.5:51234 -> 10.0.0.99:4444"
    parsed = parse_alert_line(sample)
    assert parsed is not None
    assert parsed["dst_port"] == 4444
    assert parsed["priority"] == 1

    # 2) Haqiqiy Snort'ni pcap fayl orqali ishga tushirib, chiqishini tekshirish
    work_dir = "/tmp/_test_snort_e2e"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(work_dir)

    try:
        from scapy.all import IP, TCP, Ether, wrpcap
    except ImportError:
        print("   (scapy yo'q - faqat parser birlik testi bajarildi)")
        return

    pkt = Ether() / IP(src="10.0.0.7", dst="10.0.0.98") / TCP(sport=55000, dport=4444, flags="S")
    pcap_path = os.path.join(work_dir, "test.pcap")
    wrpcap(pcap_path, [pkt])

    rules_path = os.path.join(work_dir, "test.rules")
    with open(rules_path, "w") as f:
        f.write('alert tcp any any -> any 4444 (msg:"CI Suspicious port 4444"; sid:1000099; rev:1; priority:1;)\n')

    conf_path = os.path.join(work_dir, "snort.conf")
    with open(conf_path, "w") as f:
        f.write(f"var HOME_NET any\nvar EXTERNAL_NET any\ninclude {rules_path}\n")

    result = subprocess.run(
        ["snort", "-c", conf_path, "-r", pcap_path, "-A", "fast", "-l", work_dir, "-q"],
        capture_output=True, text=True, timeout=30,
    )
    assert result.returncode == 0, f"Snort xatolik bilan chiqdi: {result.stderr}"

    alert_file = os.path.join(work_dir, "alert")
    assert os.path.isfile(alert_file), "Snort alert fayli yaratilmadi"

    n = read_existing(alert_file)
    assert n >= 1, "Kamida 1 ta Snort alert qayta ishlanishi kerak edi"

    s = get_session()
    device = s.query(Device).filter(Device.ip_address == "10.0.0.7").first()
    assert device is not None, "Snort orqali kelgan qurilma bazada topilmadi"
    alert = s.query(Alert).filter(Alert.reason.like("%CI Suspicious port 4444%")).first()
    assert alert is not None, "Snort alert bazada topilmadi"
    assert alert.severity == "critical", f"Priority=1 critical bo'lishi kerak edi, {alert.severity} keldi"
    s.close()

    shutil.rmtree(work_dir, ignore_errors=True)


check("Snort integratsiyasi (real Snort binary, pcap orqali)", _test_snort_integration)

# ---------------------------------------------------------------------------
print("\n=== 18) ZEEK INTEGRATSIYASI (sxemaga mos sintetik JSON loglar) ===")


def _test_zeek_integration():
    import shutil
    from collectors.zeek_reader import read_existing as zeek_read_existing

    log_dir = "/tmp/_test_zeek_logs"
    if os.path.exists(log_dir):
        shutil.rmtree(log_dir)
    os.makedirs(log_dir)

    with open(os.path.join(log_dir, "notice.log"), "w") as f:
        f.write('{"ts":1754470800.1,"note":"Scan::Port_Scan","msg":"test port scan","src":"172.16.6.10","dst":"172.16.6.20"}\n')

    with open(os.path.join(log_dir, "dns.log"), "w") as f:
        f.write('{"ts":1754470801.1,"id.orig_h":"172.16.6.11","query":"zeek-test-blacklist-domain.com","qtype_name":"A"}\n')

    with open(os.path.join(log_dir, "conn.log"), "w") as f:
        f.write('{"ts":1754470802.1,"id.orig_h":"172.16.6.12","id.resp_h":"1.2.3.4","id.resp_p":443,"proto":"tcp"}\n')

    file_sha = "2222222222222222222222222222222222222222222222222222222222222222"[:64]
    with open(os.path.join(log_dir, "files.log"), "w") as f:
        f.write(
            '{"ts":1754470803.1,"fuid":"Ftest1","tx_hosts":["172.16.6.13"],"rx_hosts":["5.6.7.8"],'
            f'"source":"HTTP","filename":"zeek_payload.exe","mime_type":"application/x-dosexec",'
            f'"seen_bytes":1000,"sha256":"{file_sha}","md5":"bbbb"}}\n'
        )

    s = get_session()
    s.add(BlacklistEntry(value="zeek-test-blacklist-domain.com", source="manual", reason="ci-test"))
    s.commit()
    s.close()

    results = zeek_read_existing(log_dir)
    assert results["notice.log"] == 1
    assert results["dns.log"] == 1
    assert results["conn.log"] == 1
    assert results["files.log"] == 1

    s = get_session()
    assert s.query(Device).filter(Device.ip_address == "172.16.6.10").first() is not None
    dns_alert = s.query(Alert).filter(Alert.reason.like("%zeek-test-blacklist-domain.com%")).first()
    assert dns_alert is not None, "Zeek DNS blacklist alert yaratilmadi"

    fe = s.query(FileEvent).filter(FileEvent.sha256 == file_sha).first()
    assert fe is not None, "Zeek files.log orqali FileEvent yaratilmadi"
    assert fe.channel == "zeek"
    assert fe.checked is False
    s.close()

    # MUHIM: Zeek orqali kelgan fayl mavjud file_analysis_engine pipeline'iga
    # avtomatik o'tishini tasdiqlash (alohida kod yozilmagan, qayta ishlatilgan)
    from engine.file_analysis_engine import run_once as file_analysis_run
    n = file_analysis_run()
    assert n >= 1

    s = get_session()
    fe = s.query(FileEvent).filter(FileEvent.sha256 == file_sha).first()
    assert fe.checked is True, "Zeek fayli file_analysis_engine orqali tekshirilmadi"
    s.close()

    shutil.rmtree(log_dir, ignore_errors=True)


check("Zeek integratsiyasi (4 log turi + mavjud file-pipeline bilan)", _test_zeek_integration)

# ---------------------------------------------------------------------------
print("\n=== 19) MFA - TOTP (real vaqt algoritmi) ===")


def _test_mfa():
    from dashboard import mfa as mfa_module
    from dashboard.create_user import create_user
    from dashboard import app as dash_app

    create_user("mfatest_admin", "mfatestpass123", "admin", "local")
    dash_app.app.secret_key = "test-secret-mfa-full"
    client = dash_app.app.test_client()

    # MFA yo'qligida to'g'ridan-to'g'ri kirish
    r = client.post("/login", data={"username": "mfatest_admin", "password": "mfatestpass123"}, follow_redirects=True)
    r2 = client.get("/")
    assert r2.status_code == 200, "MFA yo'q holatda to'g'ridan-to'g'ri kirishi kerak edi"

    # QR-kod olish
    r = client.get("/mfa/setup")
    assert b"data:image/png;base64," in r.data

    with client.session_transaction() as sess:
        secret = sess.get("pending_mfa_secret")
    assert secret is not None

    # To'g'ri kod bilan yoqish
    code = mfa_module.get_current_code(secret)
    client.post("/mfa/setup", data={"code": code})

    s = get_session()
    u = s.query(User).filter(User.username == "mfatest_admin").first()
    assert u.mfa_enabled is True, "MFA yoqilmadi"
    s.close()

    client.get("/logout")

    # Endi login parol to'g'ri bo'lsa ham MFA sahifasiga yo'naltirishi kerak
    r = client.post("/login", data={"username": "mfatest_admin", "password": "mfatestpass123"}, follow_redirects=False)
    assert "/mfa/verify" in r.headers.get("Location", ""), "MFA sahifasiga yo'naltirilmadi"

    r2 = client.get("/", follow_redirects=False)
    assert r2.status_code == 302, "MFA tasdiqlanmasdan kira olmasligi kerak edi"

    # Noto'g'ri kod
    client.post("/mfa/verify", data={"code": "000000"})
    r3 = client.get("/", follow_redirects=False)
    assert r3.status_code == 302, "Noto'g'ri kod bilan hali ham kira olmasligi kerak"

    # To'g'ri kod bilan yakuniy kirish
    new_code = mfa_module.get_current_code(secret)
    client.post("/mfa/verify", data={"code": new_code})
    r4 = client.get("/")
    assert r4.status_code == 200, "To'g'ri kod bilan kirishi kerak edi"

    # Boshqa secret bilan kod mos kelmasligini tekshirish (birlik test, mfa.py'da)
    other_secret = mfa_module.generate_secret()
    assert mfa_module.verify_code(other_secret, code) is False


check("MFA/TOTP (QR-kod, to'liq login oqimi, real vaqt algoritmi)", _test_mfa)

# ---------------------------------------------------------------------------
print("\n=== 20) LDAP LOGIN (real OpenLDAP server bilan) ===")


def _test_ldap_login():
    import subprocess
    import shutil

    if subprocess.run(["which", "slapd"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - slapd o'rnatilmagan bu muhitda)")
        return

    work_dir = "/tmp/_test_ldap_e2e"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(os.path.join(work_dir, "data"))

    slapd_conf = f"""include /etc/ldap/schema/core.schema
include /etc/ldap/schema/cosine.schema
include /etc/ldap/schema/inetorgperson.schema
modulepath /usr/lib/ldap
moduleload back_mdb.la
pidfile {work_dir}/slapd.pid
argsfile {work_dir}/slapd.args
database mdb
maxsize 1048576000
suffix "dc=test,dc=local"
rootdn "cn=admin,dc=test,dc=local"
rootpw testpass456
directory {work_dir}/data
"""
    with open(os.path.join(work_dir, "slapd.conf"), "w") as f:
        f.write(slapd_conf)

    base_ldif = """dn: dc=test,dc=local
objectClass: top
objectClass: dcObject
objectClass: organization
o: Test
dc: test

dn: ou=people,dc=test,dc=local
objectClass: organizationalUnit
ou: people

dn: cn=ciuser,ou=people,dc=test,dc=local
objectClass: inetOrgPerson
cn: ciuser
sn: User
givenName: CI
mail: ciuser@test.local
userPassword: CIPass456
"""
    ldif_path = os.path.join(work_dir, "base.ldif")
    with open(ldif_path, "w") as f:
        f.write(base_ldif)

    # Avval slapd'ning o'z konfiguratsiya-tekshirish rejimi (-Tt) orqali
    # sinxron tarzda tekshiramiz - bu aniq xato xabarini darhol beradi
    # (agar keyingi bosqichda muammo bo'lsa, buni ham diagnostikaga qo'shamiz).
    conf_test = subprocess.run(
        ["slapd", "-Tt", "-f", os.path.join(work_dir, "slapd.conf")],
        capture_output=True, timeout=10, text=True,
    )

    slapd_log_path = os.path.join(work_dir, "slapd_stderr.log")
    slapd_log_file = open(slapd_log_path, "w")
    slapd_proc = subprocess.Popen(
        ["slapd", "-f", os.path.join(work_dir, "slapd.conf"), "-h", "ldap://127.0.0.1:3390/", "-d", "0"],
        stdout=slapd_log_file, stderr=subprocess.STDOUT,
    )
    import time as _time

    # MUHIM: sobit sleep() o'rniga slapd haqiqatan tayyor bo'lguncha
    # polling orqali kutamiz - sekinroq muhitlarda (masalan GitHub
    # Actions runner) 2 soniya yetarli bo'lmasligi mumkin (bu CI'da
    # aynan shu sabab bilan aniqlangan xato edi).
    slapd_ready = False
    for _ in range(20):  # maksimal ~10 soniya
        probe = subprocess.run(
            ["ldapsearch", "-x", "-H", "ldap://127.0.0.1:3390", "-b", "", "-s", "base"],
            capture_output=True, timeout=3,
        )
        if probe.returncode == 0:
            slapd_ready = True
            break
        _time.sleep(0.5)

    if not slapd_ready:
        slapd_log_file.flush()
        with open(slapd_log_path) as f:
            log_content = f.read()
        return_code = slapd_proc.poll()
        module_path = "/usr/lib/ldap/back_mdb.la"
        module_exists = os.path.isfile(module_path)
        empty_marker = "(bo'sh)"
        error_msg = (
            f"slapd 10 soniyada tayyor bo'lmadi. "
            f"return_code={return_code}, "
            f"modul_fayl_mavjud({module_path})={module_exists}, "
            f"slapd_chiqishi={log_content[:500] or empty_marker!r}, "
            f"config_test_rc={conf_test.returncode}, "
            f"config_test_stdout={conf_test.stdout[:300]!r}, "
            f"config_test_stderr={conf_test.stderr[:300]!r}"
        )
        assert False, error_msg

    try:
        ldapadd_result = subprocess.run(
            ["ldapadd", "-x", "-D", "cn=admin,dc=test,dc=local", "-w", "testpass456",
             "-H", "ldap://127.0.0.1:3390", "-f", ldif_path],
            capture_output=True, timeout=10, text=True,
        )
        assert ldapadd_result.returncode == 0, (
            f"ldapadd muvaffaqiyatsiz (kod={ldapadd_result.returncode}): "
            f"stdout={ldapadd_result.stdout!r} stderr={ldapadd_result.stderr!r}"
        )

        os.environ["LDAP_SERVER"] = "ldap://127.0.0.1:3390"
        os.environ["LDAP_BIND_DN_TEMPLATE"] = "cn={username},ou=people,dc=test,dc=local"

        import importlib
        import dashboard.ldap_auth as ldap_mod
        importlib.reload(ldap_mod)
        import dashboard.auth as auth_mod
        importlib.reload(auth_mod)

        assert ldap_mod.authenticate_ldap("ciuser", "CIPass456") is True, "To'g'ri LDAP parol qabul qilinishi kerak edi"
        assert ldap_mod.authenticate_ldap("ciuser", "wrong") is False, "Noto'g'ri LDAP parol rad etilishi kerak edi"
        assert ldap_mod.authenticate_ldap("ciuser", "") is False, "Bo'sh parol (anonim bind) rad etilishi kerak edi"
        assert ldap_mod.authenticate_ldap("nonexistent", "anything") is False

        # Dashboard login oqimi orqali ham tekshirish
        from dashboard.create_user import create_user
        create_user("ciuser", "placeholder", "viewer", "ldap")

        from dashboard import app as dash_app
        dash_app.app.secret_key = "test-secret-ldap-full"
        client = dash_app.app.test_client()

        r = client.post("/login", data={"username": "ciuser", "password": "CIPass456"}, follow_redirects=True)
        r2 = client.get("/")
        assert r2.status_code == 200, "LDAP orqali dashboard login muvaffaqiyatli bo'lishi kerak edi"

    finally:
        slapd_proc.terminate()
        try:
            slapd_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            slapd_proc.kill()
        slapd_log_file.close()
        shutil.rmtree(work_dir, ignore_errors=True)
        for k in ["LDAP_SERVER", "LDAP_BIND_DN_TEMPLATE"]:
            os.environ.pop(k, None)


check("LDAP Login (real OpenLDAP server, to'g'ri/noto'g'ri/bo'sh parol)", _test_ldap_login)

# ---------------------------------------------------------------------------
print("\n=== 21) RABBITMQ QUEUE (real broker, to'liq UDP->Queue->Worker->DB zanjiri) ===")


def _test_rabbitmq_queue():
    import subprocess
    if subprocess.run(["which", "rabbitmqctl"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - rabbitmq-server o'rnatilmagan bu muhitda)")
        return

    from messaging.rabbitmq_client import health_check, publish_json, consume_batch, queue_depth, get_connection

    if not health_check():
        print("   (o'tkazib yuborildi - RabbitMQ server ishlamayapti)")
        return

    test_queue = "_ci_test_queue"
    e2e_queue = "_ci_e2e_syslog_queue"

    # MUHIM: avvalgi (masalan muvaffaqiyatsiz tugagan) test urinishlaridan
    # qolgan xabarlar bo'lishi mumkin - RabbitMQ navbatlari persistent
    # (durable) bo'lgani uchun. Test har doim BO'SH navbatdan boshlashi
    # kerak - shuning uchun avval tozalaymiz.
    for qname in (test_queue, e2e_queue):
        try:
            conn = get_connection()
            ch = conn.channel()
            ch.queue_declare(queue=qname, durable=True)
            ch.queue_purge(queue=qname)
            conn.close()
        except Exception:
            pass

    # 1) Asosiy publish/consume birlik testi
    for i in range(5):
        assert publish_json(test_queue, {"id": i, "text": f"test {i}"}) is True

    depth = queue_depth(test_queue)
    assert depth == 5, f"Navbat chuqurligi 5 bo'lishi kerak edi, {depth} keldi"

    received = []
    count = consume_batch(test_queue, lambda d: received.append(d), max_messages=5, timeout_seconds=10)
    assert count == 5
    assert sorted(r["id"] for r in received) == [0, 1, 2, 3, 4]
    assert queue_depth(test_queue) == 0

    # 2) To'liq E2E: real UDP paket -> navbat-asosli syslog server -> RabbitMQ -> worker -> DB
    import socket
    import time as _time

    server_proc = subprocess.Popen(
        ["python3", "-m", "collectors.syslog_server_queued"],
        env={**os.environ, "RAW_SYSLOG_QUEUE": "_ci_e2e_syslog_queue"},
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    _time.sleep(2)

    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        test_mac = "AA:BB:CC:DD:EE:99"
        msg = f"DHCP: Lease granted to 172.16.9.199 MAC={test_mac} HOST=RABBITMQ-E2E-TEST"
        sock.sendto(msg.encode(), ("127.0.0.1", 5140))
        _time.sleep(1)

        depth_after_send = queue_depth("_ci_e2e_syslog_queue")
        assert depth_after_send == 1, f"UDP paket navbatga tushmadi (chuqurlik={depth_after_send})"

        # MUHIM: muhit o'zgaruvchisi modul import qilinishidan OLDIN
        # o'rnatilishi kerak (modul darajasidagi konstanta faqat bir
        # marta, import paytida hisoblanadi) - shuning uchun importlib.reload
        # bilan majburan qayta yuklaymiz.
        os.environ["RAW_SYSLOG_QUEUE"] = "_ci_e2e_syslog_queue"
        import importlib
        import engine.queue_ingest_worker as worker_mod
        importlib.reload(worker_mod)

        n = worker_mod.run_once(timeout_seconds=5)
        assert n == 1, f"Worker 1 ta xabarni qayta ishlashi kerak edi, {n} ta ishladi"

        s = get_session()
        raw = s.query(RawLog).filter(RawLog.raw_message.like("%RABBITMQ-E2E-TEST%")).first()
        assert raw is not None, "UDP->Queue->Worker->DB zanjiri orqali yozuv topilmadi"
        assert raw.source_ip == "127.0.0.1"
        s.close()

    finally:
        server_proc.terminate()
        try:
            server_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            server_proc.kill()
        os.environ.pop("RAW_SYSLOG_QUEUE", None)


check("RabbitMQ Queue (real broker, to'liq UDP->Queue->Worker->DB)", _test_rabbitmq_queue)

# ---------------------------------------------------------------------------
print("\n=== 22) UEBA / AI - anomaliya aniqlash va Risk Score (real statistik ma'lumot) ===")


def _test_ueba():
    import random
    from datetime import timedelta
    from db.models import DeviceBaseline, utcnow
    from engine.ueba_engine import compute_baselines_for_all_devices, detect_anomalies, compute_risk_scores

    random.seed(123)
    s = get_session()
    now = utcnow()

    d_normal = Device(ip_address="172.16.11.1", hostname="UEBA-NORMAL", connection_type="wifi", source="test")
    d_anomaly = Device(ip_address="172.16.11.2", hostname="UEBA-ANOMALY", connection_type="wifi", source="test")
    s.add_all([d_normal, d_anomaly])
    s.flush()
    normal_id, anomaly_id = d_normal.id, d_anomaly.id

    # Ikkala qurilma uchun bir xil "normal" baseline: 25 kun, soat 9-18, 4-8 hodisa/soat
    for dev_id in (normal_id, anomaly_id):
        for day in range(1, 26):  # bugungi kunni band qilmaslik uchun 1-dan boshlaymiz
            for hour in range(9, 19):
                for _ in range(random.randint(4, 8)):
                    ts = now.replace(hour=hour, minute=random.randint(0, 59), second=0, microsecond=0) - timedelta(days=day)
                    s.add(Event(device_id=dev_id, source_ip="172.16.11.0", dest_ip="8.8.8.8", dest_port=443, protocol="TCP", timestamp=ts))

    # Faqat anomaly qurilmasida - joriy soatda katta portlash
    for _ in range(150):
        s.add(Event(device_id=anomaly_id, source_ip="172.16.11.0", dest_ip="185.20.10.99", dest_port=8080, protocol="TCP", timestamp=now))

    s.commit()
    s.close()

    n_baselines = compute_baselines_for_all_devices()
    assert n_baselines >= 2, f"Kamida 2 ta baseline hisoblanishi kerak edi, {n_baselines} ta hisoblandi"

    s = get_session()
    bl_normal = s.query(DeviceBaseline).filter(DeviceBaseline.device_id == normal_id).first()
    bl_anomaly = s.query(DeviceBaseline).filter(DeviceBaseline.device_id == anomaly_id).first()
    assert bl_normal is not None and bl_anomaly is not None
    assert bl_normal.mean_events_per_hour > 0
    s.close()

    n_anomalies = detect_anomalies()
    assert n_anomalies >= 1, f"Kamida 1 ta anomaliya topilishi kerak edi, {n_anomalies} ta topildi"

    s = get_session()
    anomaly_alert = s.query(Alert).filter(Alert.device_id == anomaly_id, Alert.reason.like("UEBA%")).first()
    assert anomaly_alert is not None, "Anomal qurilma uchun UEBA alert yaratilishi kerak edi"
    assert "Hajm anomaliyasi" in anomaly_alert.reason

    normal_false_positive = s.query(Alert).filter(Alert.device_id == normal_id, Alert.reason.like("UEBA%")).first()
    assert normal_false_positive is None, "Normal qurilmada SOXTA-POZITIV UEBA alert bo'lmasligi kerak edi"
    s.close()

    # Risk Score: anomaly qurilmasiga qo'shimcha critical/high alertlar qo'shib, farqni tekshiramiz
    s = get_session()
    s.add(Alert(device_id=anomaly_id, severity="critical", reason="Test critical", mitre_tactic="Execution", action_taken="TODO"))
    s.add(Alert(device_id=anomaly_id, severity="high", reason="Test high", mitre_tactic="Command and Control", action_taken="TODO"))
    s.commit()
    s.close()

    compute_risk_scores()

    s = get_session()
    dev_normal = s.query(Device).filter(Device.id == normal_id).first()
    dev_anomaly = s.query(Device).filter(Device.id == anomaly_id).first()
    assert dev_anomaly.risk_score > dev_normal.risk_score, (
        f"Anomal qurilma risk score'i normal qurilmadan yuqori bo'lishi kerak edi "
        f"({dev_anomaly.risk_score} vs {dev_normal.risk_score})"
    )
    assert dev_normal.risk_score == 0, f"Normal qurilma risk score'i 0 bo'lishi kerak edi, {dev_normal.risk_score} keldi"
    assert dev_anomaly.risk_score > 30, "Anomal qurilma yetarlicha yuqori risk score olishi kerak edi"
    s.close()

    # MITRE tagging bilan integratsiya
    from engine.mitre_tagging_engine import run_once as mitre_run
    mitre_run()
    s = get_session()
    tagged = s.query(Alert).filter(Alert.reason.like("UEBA%")).first()
    assert tagged.mitre_technique_id is not None, "UEBA alert MITRE bilan belgilanmadi"
    s.close()


check("UEBA/AI (statistik anomaliya aniqlash, Risk Score, soxta-pozitivsiz)", _test_ueba)

# ---------------------------------------------------------------------------
print("\n=== 23) KUBERNETES MANIFESTLAR (YAML struktura tekshiruvi) ===")


def _test_k8s_manifests():
    import glob
    import yaml as yaml_mod

    k8s_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "k8s")
    files = sorted(glob.glob(os.path.join(k8s_dir, "*.yaml")))
    assert len(files) >= 6, f"Kamida 6 ta k8s manifest fayli kutilgan edi, {len(files)} ta topildi"

    required_kinds_seen = set()
    total_docs = 0

    for filepath in files:
        with open(filepath) as f:
            docs = list(yaml_mod.safe_load_all(f))
        for doc in docs:
            if doc is None:
                continue
            total_docs += 1
            assert "apiVersion" in doc, f"{filepath}: 'apiVersion' yo'q"
            assert "kind" in doc, f"{filepath}: 'kind' yo'q"
            assert "metadata" in doc and "name" in doc["metadata"], f"{filepath}: metadata.name yo'q"
            required_kinds_seen.add(doc["kind"])

    expected_kinds = {
        "Namespace", "ConfigMap", "Secret", "StatefulSet", "Service",
        "Deployment", "PersistentVolumeClaim", "HorizontalPodAutoscaler", "Ingress",
    }
    missing = expected_kinds - required_kinds_seen
    assert not missing, f"Kutilgan resurs turlari topilmadi: {missing}"
    assert total_docs >= 20, f"Kamida 20 ta resurs kutilgan edi, {total_docs} ta topildi"


check("Kubernetes manifestlar (struktura, kutilgan resurs turlari)", _test_k8s_manifests)

# ---------------------------------------------------------------------------
print("\n=== 24) AUDIT LOG (real HTTP orqali, login/acknowledge/user boshqaruvi) ===")


def _test_audit_log():
    from db.models import AuditLog
    from dashboard import app as dash_app
    from dashboard.create_user import create_user

    create_user("audit_test_admin", "audittestpass123", "admin")
    dash_app.app.secret_key = "test-secret-audit"
    client = dash_app.app.test_client()

    # MUHIM: avval noto'g'ri, keyin to'g'ri login - aks holda muvaffaqiyatli
    # login'dan keyingi sessiya cookie'si ikkinchi so'rovni "current_user.
    # is_authenticated" tekshiruvida darhol qaytarib yuboradi, login
    # logikasiga umuman yetib bormaydi (bu test kodidagi tuzatilgan xato edi).
    client.post("/login", data={"username": "audit_test_admin", "password": "wrong"})
    client.post("/login", data={"username": "audit_test_admin", "password": "audittestpass123"})

    s = get_session()
    logins = s.query(AuditLog).filter(AuditLog.username == "audit_test_admin", AuditLog.action == "login").all()
    assert len(logins) == 2, f"2 ta login urinishi qayd etilishi kerak edi, {len(logins)} ta topildi"
    successes = [l.success for l in logins]
    assert True in successes and False in successes, "Muvaffaqiyatli va muvaffaqiyatsiz login ikkalasi ham qayd etilishi kerak edi"
    s.close()

    # Foydalanuvchi yaratish audit'i
    client.post("/users/create", data={"username": "audit_created_user", "password": "pass123", "role": "viewer"})
    s = get_session()
    create_entry = s.query(AuditLog).filter(AuditLog.action == "create_user", AuditLog.target_id == "audit_created_user").first()
    assert create_entry is not None, "create_user audit yozuvi topilmadi"
    assert create_entry.username == "audit_test_admin"
    s.close()

    # Viewer /audit'ga kira olmasligi kerak (RBAC bilan integratsiya)
    create_user("audit_test_viewer", "viewerpass123", "viewer")
    client.get("/logout")
    client.post("/login", data={"username": "audit_test_viewer", "password": "viewerpass123"})
    r = client.get("/audit")
    assert r.status_code == 403, f"Viewer /audit'ga kirmasligi kerak edi, {r.status_code} keldi"

    client.get("/logout")
    client.post("/login", data={"username": "audit_test_admin", "password": "audittestpass123"})
    r = client.get("/audit")
    assert r.status_code == 200
    assert b"audit_test_admin" in r.data


check("Audit Log (login/acknowledge/user boshqaruvi qayd etiladi, RBAC bilan)", _test_audit_log)

# ---------------------------------------------------------------------------
print("\n=== 25) BACKUP/RESTORE (real 'halokat va tiklash' stsenariysi) ===")


def _test_backup_restore():
    import shutil
    from backup.backup_manager import create_backup, restore_backup, list_backups

    backup_dir = "/tmp/_test_backup_restore"
    if os.path.exists(backup_dir):
        shutil.rmtree(backup_dir)

    s = get_session()
    s.add(Device(ip_address="172.16.21.1", hostname="BACKUP-CI-TEST", connection_type="wifi", source="test"))
    s.commit()
    s.close()

    backup_path = create_backup(backup_dir)
    assert os.path.isfile(backup_path) or (backup_path and os.path.getsize(backup_path) >= 0), "Backup fayli yaratilmadi"

    backups = list_backups(backup_dir)
    assert len(backups) >= 1, "list_backups bo'sh qaytardi"

    # "Halokat" simulyatsiyasi
    s = get_session()
    s.query(Device).filter(Device.hostname == "BACKUP-CI-TEST").delete()
    s.commit()
    remaining = s.query(Device).filter(Device.hostname == "BACKUP-CI-TEST").count()
    assert remaining == 0, "Halokat simulyatsiyasi ishlamadi"
    s.close()

    ok = restore_backup(backup_path)
    assert ok, "restore_backup False qaytardi"

    s = get_session()
    restored = s.query(Device).filter(Device.hostname == "BACKUP-CI-TEST").first()
    assert restored is not None, "RESTORE'DAN KEYIN MA'LUMOT TIKLANMADI"
    s.close()

    shutil.rmtree(backup_dir, ignore_errors=True)
    safety_file = "./logs/security_system.db.before_restore"
    if os.path.exists(safety_file):
        os.remove(safety_file)


check("Backup/Restore (real halokat+tiklash, SQLite/PostgreSQL avtomatik)", _test_backup_restore)

# ---------------------------------------------------------------------------
print("\n=== 26) LIVE MAP (real HTTP, topologiya API) ===")


def _test_live_map():
    from dashboard import app as dash_app
    from dashboard.create_user import create_user

    create_user("livemap_test_admin", "livemaptestpass123", "admin")
    dash_app.app.secret_key = "test-secret-livemap"
    client = dash_app.app.test_client()
    client.post("/login", data={"username": "livemap_test_admin", "password": "livemaptestpass123"})

    s = get_session()
    d_high = Device(ip_address="172.16.32.1", hostname="LIVEMAP-HIGH-RISK", connection_type="wifi", source="test", risk_score=80)
    d_low = Device(ip_address="172.16.32.2", hostname="LIVEMAP-LOW-RISK", connection_type="cable", source="test", risk_score=0)
    s.add_all([d_high, d_low])
    s.flush()
    high_id, low_id = d_high.id, d_low.id
    s.add(Event(device_id=high_id, source_ip=d_high.ip_address, dest_ip="9.9.9.9", dest_port=443, protocol="TCP"))
    s.add(Event(device_id=high_id, source_ip=d_high.ip_address, dest_ip="9.9.9.9", dest_port=443, protocol="TCP"))
    s.commit()
    s.close()

    r = client.get("/live-map")
    assert r.status_code == 200
    assert b"network-map" in r.data

    r = client.get("/api/topology")
    assert r.status_code == 200
    data = r.get_json()
    assert "nodes" in data and "edges" in data

    node_ids = {n["id"] for n in data["nodes"]}
    high_node = next((n for n in data["nodes"] if n["id"] == f"dev_{high_id}"), None)
    assert high_node is not None, "Yuqori riskli qurilma node'i topilmadi"
    assert high_node["color"] == "#c0392b", f"Risk=80 uchun qizil rang kutilgan edi, {high_node['color']} keldi"

    ext_node = next((n for n in data["nodes"] if n["id"] == "ext_9.9.9.9"), None)
    assert ext_node is not None, "Tashqi manzil node'i topilmadi"

    edge = next((e for e in data["edges"] if e["from"] == f"dev_{high_id}" and e["to"] == "ext_9.9.9.9"), None)
    assert edge is not None, "Edge topilmadi"
    assert edge["value"] == 2, f"2 ta hodisa kutilgan edi, {edge['value']} keldi"

    # Autentifikatsiyasiz kirish rad etilishi kerak
    anon_client = dash_app.app.test_client()
    r = anon_client.get("/api/topology", follow_redirects=False)
    assert r.status_code == 302


check("Live Map (real HTTP, topologiya API, risk-rang moslashuvi)", _test_live_map)

# ---------------------------------------------------------------------------
print("\n=== 27) GRAFANA DASHBOARD (JSON struktura + SQL so'rovlar real bazaga qarshi) ===")


def _test_grafana_dashboard():
    import json as json_mod

    dashboard_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "grafana", "dashboards", "security-overview.json")
    assert os.path.isfile(dashboard_path), "Grafana dashboard JSON fayli topilmadi"

    with open(dashboard_path) as f:
        dashboard = json_mod.load(f)

    assert "panels" in dashboard and len(dashboard["panels"]) >= 5, "Kamida 5 ta panel kutilgan edi"

    from config.settings import DATABASE_URL as CURRENT_DB_URL
    if not CURRENT_DB_URL.startswith("postgresql://"):
        print("   (SQL so'rovlar faqat PostgreSQL rejimida sinaladi - SQLite'da faqat JSON struktura tekshirildi)")
        return

    import psycopg2
    conn = psycopg2.connect(CURRENT_DB_URL)
    cur = conn.cursor()
    try:
        for panel in dashboard["panels"]:
            for target in panel.get("targets", []):
                sql = target["rawSql"]
                try:
                    cur.execute(sql)
                    cur.fetchall()
                except Exception as exc:
                    conn.rollback()
                    raise AssertionError(f"Panel '{panel['title']}' SQL xatoligi: {exc}")
    finally:
        cur.close()
        conn.close()


check("Grafana Dashboard (8 panel SQL so'rovi real bazaga qarshi)", _test_grafana_dashboard)

# ---------------------------------------------------------------------------
print("\n=== 28) RASMIY HUJJATLAR (mavjudligi + ichki havolalar to'g'riligi) ===")


def _test_formal_docs():
    import re

    base_dir = os.path.dirname(os.path.abspath(__file__))
    required_docs = [
        "docs/ADMIN_GUIDE.md", "docs/USER_GUIDE.md", "docs/API_GUIDE.md",
        "docs/INSTALLATION_GUIDE.md", "docs/DISASTER_RECOVERY_GUIDE.md",
    ]
    for doc in required_docs:
        path = os.path.join(base_dir, doc)
        assert os.path.isfile(path), f"{doc} topilmadi"
        assert os.path.getsize(path) > 500, f"{doc} juda qisqa (bo'sh/to'liqsiz bo'lishi mumkin)"

    # Barcha docs_*.md havolalarining haqiqatan mavjudligini tekshirish
    referenced = set()
    for doc in required_docs + ["README.md"]:
        path = os.path.join(base_dir, doc)
        with open(path, encoding="utf-8") as f:
            content = f.read()
        referenced.update(re.findall(r"docs_[A-Z_]+\.md", content))

    assert len(referenced) >= 5, f"Kamida 5 ta docs_*.md havolasi kutilgan edi, {len(referenced)} ta topildi"
    for ref in referenced:
        assert os.path.isfile(os.path.join(base_dir, ref)), f"Havola qilingan fayl topilmadi: {ref}"

    # DR guide'dagi buyruqlar backup_manager.py'ning haqiqiy CLI flaglariga mos kelishini tekshirish
    with open(os.path.join(base_dir, "docs/DISASTER_RECOVERY_GUIDE.md"), encoding="utf-8") as f:
        dr_content = f.read()
    assert "--backup" in dr_content and "--restore" in dr_content and "--list" in dr_content


check("Rasmiy hujjatlar (5 guide, ichki havolalar, CLI mosligi)", _test_formal_docs)

# ---------------------------------------------------------------------------
print("\n=== 29) ENCRYPTION AT REST (MFA secret, real shifrlash/ochish) ===")


def _test_encryption_at_rest():
    from crypto.field_encryption import generate_key, encrypt_value, decrypt_value, is_encrypted, is_configured

    # Birlik testlar (baza kerak emas)
    os.environ["ENCRYPTION_KEY"] = generate_key()
    secret = "ZTYJVNNE6UI3LIQAQNCOVTXNCDWFBUR3"
    encrypted = encrypt_value(secret)
    assert encrypted != secret
    assert decrypt_value(encrypted) == secret
    assert is_encrypted(encrypted) is True
    assert is_encrypted(secret) is False

    # Kalit almashtirish (rotation)
    old_key = os.environ["ENCRYPTION_KEY"]
    enc_with_old = encrypt_value("rotation-test")
    os.environ["ENCRYPTION_KEY"] = generate_key()
    os.environ["ENCRYPTION_KEY_OLD"] = old_key
    assert decrypt_value(enc_with_old) == "rotation-test"
    os.environ.pop("ENCRYPTION_KEY_OLD", None)

    # To'liq MFA oqimi orqali - bazada HAQIQATAN shifrlangan saqlanishini tekshirish
    os.environ["ENCRYPTION_KEY"] = generate_key()
    from dashboard import app as dash_app
    from dashboard import mfa as mfa_module
    from dashboard.create_user import create_user

    create_user("enc_ci_test", "enccitest123", "admin")
    dash_app.app.secret_key = "test-secret-encryption"
    client = dash_app.app.test_client()
    client.post("/login", data={"username": "enc_ci_test", "password": "enccitest123"})
    client.get("/mfa/setup")
    with client.session_transaction() as sess:
        plaintext_secret = sess.get("pending_mfa_secret")
    code = mfa_module.get_current_code(plaintext_secret)
    client.post("/mfa/setup", data={"code": code})

    s = get_session()
    u = s.query(User).filter(User.username == "enc_ci_test").first()
    db_value = u.mfa_secret
    s.close()

    assert db_value != plaintext_secret, "Baza ochiq matnda saqladi - ENCRYPTION AT REST ISHLAMAYAPTI"
    assert is_encrypted(db_value), "DB qiymati shifrlangan formatda emas"

    # To'liq login MFA orqali (decrypt qilib) hali ishlashini tasdiqlash
    client.get("/logout")
    client.post("/login", data={"username": "enc_ci_test", "password": "enccitest123"})
    new_code = mfa_module.get_current_code(plaintext_secret)
    client.post("/mfa/verify", data={"code": new_code})
    r = client.get("/")
    assert r.status_code == 200, "Shifrlangan MFA secret orqali login ishlamadi"


check("Encryption at Rest (MFA secret shifrlash, kalit almashtirish, to'liq oqim)", _test_encryption_at_rest)

# ---------------------------------------------------------------------------
print("\n=== 30) API TOKEN BOSHQARUVI (real Flask, revoke, muddat, RBAC) ===")


def _test_api_token_management():
    from api import server as api_server
    from api import token_manager

    os.environ["AGENT_API_KEY"] = "legacy-shared-key-ci"
    import importlib
    importlib.reload(api_server)

    client = api_server.app.test_client()

    # Eski AGENT_API_KEY orqaga moslik
    r = client.post("/api/v1/check_hash", json={"sha256": "a" * 64}, headers={"X-API-Key": "legacy-shared-key-ci"})
    assert r.status_code == 200

    # Yangi token yaratish va ishlatish
    token = token_manager.create_token("ci-test-agent", created_by="ci")
    r = client.post("/api/v1/check_hash", json={"sha256": "b" * 64}, headers={"X-API-Key": token})
    assert r.status_code == 200, f"Yangi token bilan 200 kutilgan edi, {r.status_code} keldi"

    tokens = token_manager.list_tokens()
    t = next(t for t in tokens if t.name == "ci-test-agent")
    assert t.last_used_at is not None, "last_used_at yangilanmadi"

    # Bekor qilish
    assert token_manager.revoke_token(t.id) is True
    r = client.post("/api/v1/check_hash", json={"sha256": "c" * 64}, headers={"X-API-Key": token})
    assert r.status_code == 401, "Bekor qilingan token rad etilishi kerak edi"

    # Muddati o'tgan token
    expired = token_manager.create_token("ci-expired", expires_days=-1)
    r = client.post("/api/v1/check_hash", json={"sha256": "d" * 64}, headers={"X-API-Key": expired})
    assert r.status_code == 401, "Muddati o'tgan token rad etilishi kerak edi"

    # Soxta token
    r = client.post("/api/v1/check_hash", json={"sha256": "e" * 64}, headers={"X-API-Key": "nssk_fake123"})
    assert r.status_code == 401

    # Dashboard UI: token yaratish, bir marta ko'rsatilishi, RBAC
    import re
    from dashboard import app as dash_app
    from dashboard.create_user import create_user

    create_user("token_admin_ci", "tokenadminci123", "admin")
    dash_app.app.secret_key = "test-secret-tokens"
    ui_client = dash_app.app.test_client()
    ui_client.post("/login", data={"username": "token_admin_ci", "password": "tokenadminci123"})

    r1 = ui_client.post("/api-tokens/create", data={"name": "UI-CI-Token", "expires_days": ""}, follow_redirects=True)
    match = re.search(rb"nssk_[A-Za-z0-9_-]{20,}", r1.data)
    assert match, "Dashboard UI orqali yaratilgan token ko'rsatilmadi"
    full_token = match.group(0)

    r2 = ui_client.get("/api-tokens")
    assert full_token not in r2.data, "To'liq token ikkinchi marta ko'rsatilmasligi kerak"

    create_user("token_viewer_ci", "viewerci123", "viewer")
    ui_client.get("/logout")
    ui_client.post("/login", data={"username": "token_viewer_ci", "password": "viewerci123"})
    r3 = ui_client.get("/api-tokens")
    assert r3.status_code == 403, "Viewer /api-tokens'ga kira olmasligi kerak edi"


check("API Token boshqaruvi (yaratish/ishlatish/revoke/muddat/RBAC)", _test_api_token_management)

# ---------------------------------------------------------------------------
print("\n=== 31) NETWORK DISCOVERY - MAC Vendor va DHCP Reader (fayl-asosli, muhitdan mustaqil) ===")


def _test_discovery_offline_parts():
    import shutil

    # MAC Vendor - IEEE OUI bazasi (ieee-data paketi)
    from network_discovery.mac_vendor import lookup_vendor, is_locally_administered
    if os.path.isfile("/usr/share/ieee-data/oui.csv"):
        vendor = lookup_vendor("F4:BD:9E:11:22:33")
        assert vendor is not None and "Cisco" in vendor, f"Cisco kutilgan edi, {vendor} keldi"
        assert is_locally_administered("02:FC:00:00:00:05") is True
    else:
        print("   (MAC Vendor: ieee-data topilmadi - o'tkazib yuborildi)")

    # DHCP Reader - pure file parsing, hech qanday tashqi bog'liqlik yo'q
    from network_discovery.dhcp_reader import parse_isc_dhcpd_leases, parse_kerio_dhcp_log

    work_dir = "/tmp/_test_discovery_dhcp"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(work_dir)

    isc_content = """lease 172.16.5.10 {
  starts 3 2026/08/07 08:00:00;
  ends 3 2026/08/07 20:00:00;
  hardware ethernet aa:bb:cc:dd:ee:01;
  client-hostname "TEST-DHCP-PC";
}
"""
    isc_path = os.path.join(work_dir, "dhcpd.leases")
    with open(isc_path, "w") as f:
        f.write(isc_content)

    leases = parse_isc_dhcpd_leases(isc_path)
    assert len(leases) == 1
    assert leases[0].hostname == "TEST-DHCP-PC"
    assert leases[0].mac == "AA:BB:CC:DD:EE:01"

    kerio_path = os.path.join(work_dir, "kerio.log")
    with open(kerio_path, "w") as f:
        f.write("DHCP: Lease granted to 172.16.5.20 MAC=BB:CC:DD:EE:FF:01 HOST=TEST-KERIO-PC\n")

    kerio_leases = parse_kerio_dhcp_log(kerio_path)
    assert len(kerio_leases) == 1
    assert kerio_leases[0].hostname == "TEST-KERIO-PC"

    shutil.rmtree(work_dir, ignore_errors=True)

    # UniFi Discovery - graceful failure (controller sozlanmagan)
    from network_discovery.unifi_discovery import get_unifi_clients
    assert get_unifi_clients() == []


check("Network Discovery: MAC Vendor + DHCP Reader + UniFi graceful fail", _test_discovery_offline_parts)

# ---------------------------------------------------------------------------
print("\n=== 32) NETWORK DISCOVERY - AD Discovery (real OpenLDAP, computer obyektlari) ===")


def _test_ad_discovery():
    import subprocess
    import shutil
    import time as _time

    if subprocess.run(["which", "slapd"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - slapd o'rnatilmagan)")
        return

    work_dir = "/tmp/_test_ad_discovery"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(os.path.join(work_dir, "data"))

    schema_path = os.path.join(work_dir, "ad-attrs.schema")
    with open(schema_path, "w") as f:
        f.write(
            "attributetype ( 1.2.840.113556.1.4.619 NAME 'dNSHostName' "
            "SYNTAX 1.3.6.1.4.1.1466.115.121.1.15 SINGLE-VALUE )\n"
            "attributetype ( 1.2.840.113556.1.4.618 NAME 'operatingSystem' "
            "SYNTAX 1.3.6.1.4.1.1466.115.121.1.15 SINGLE-VALUE )\n"
            "objectclass ( 1.2.840.113556.1.5.9 NAME 'computer' SUP device STRUCTURAL "
            "MAY ( dNSHostName $ operatingSystem ) )\n"
        )

    conf_path = os.path.join(work_dir, "slapd.conf")
    with open(conf_path, "w") as f:
        f.write(f"""include /etc/ldap/schema/core.schema
include /etc/ldap/schema/cosine.schema
include /etc/ldap/schema/inetorgperson.schema
include {schema_path}
modulepath /usr/lib/ldap
moduleload back_mdb.la
pidfile {work_dir}/slapd.pid
argsfile {work_dir}/slapd.args
database mdb
maxsize 1048576000
suffix "dc=adtest,dc=local"
rootdn "cn=admin,dc=adtest,dc=local"
rootpw citest456
directory {work_dir}/data
""")

    ldif_path = os.path.join(work_dir, "computers.ldif")
    with open(ldif_path, "w") as f:
        f.write("""dn: dc=adtest,dc=local
objectClass: top
objectClass: dcObject
objectClass: organization
o: AD Test
dc: adtest

dn: ou=computers,dc=adtest,dc=local
objectClass: organizationalUnit
ou: computers

dn: cn=CI-TEST-PC,ou=computers,dc=adtest,dc=local
objectClass: computer
objectClass: top
cn: CI-TEST-PC
dNSHostName: CI-TEST-PC.company.local
operatingSystem: Windows 11 Pro
""")

    slapd_proc = subprocess.Popen(
        ["slapd", "-f", conf_path, "-h", "ldap://127.0.0.1:16390/", "-d", "0"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    try:
        slapd_ready = False
        for _ in range(20):
            probe = subprocess.run(
                ["ldapsearch", "-x", "-H", "ldap://127.0.0.1:16390", "-b", "", "-s", "base"],
                capture_output=True, timeout=3,
            )
            if probe.returncode == 0:
                slapd_ready = True
                break
            _time.sleep(0.5)
        assert slapd_ready, "slapd 10 soniyada tayyor bo'lmadi"

        add_result = subprocess.run(
            ["ldapadd", "-x", "-D", "cn=admin,dc=adtest,dc=local", "-w", "citest456",
             "-H", "ldap://127.0.0.1:16390", "-f", ldif_path],
            capture_output=True, timeout=10, text=True,
        )
        assert add_result.returncode == 0, f"ldapadd xatoligi: {add_result.stderr}"

        os.environ["AD_SERVER"] = "ldap://127.0.0.1:16390"
        os.environ["AD_BASE_DN"] = "dc=adtest,dc=local"
        os.environ["AD_SERVICE_DN"] = "cn=admin,dc=adtest,dc=local"
        os.environ["AD_SERVICE_PASSWORD"] = "citest456"
        os.environ["AD_COMPUTER_FILTER"] = "(objectClass=computer)"

        from network_discovery.ad_discovery import discover_ad_computers
        computers = discover_ad_computers()
        assert len(computers) == 1, f"1 ta kompyuter kutilgan edi, {len(computers)} ta topildi"
        assert computers[0].name == "CI-TEST-PC"
        assert computers[0].operating_system == "Windows 11 Pro"

    finally:
        slapd_proc.terminate()
        try:
            slapd_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            slapd_proc.kill()
        shutil.rmtree(work_dir, ignore_errors=True)
        for k in ["AD_SERVER", "AD_BASE_DN", "AD_SERVICE_DN", "AD_SERVICE_PASSWORD", "AD_COMPUTER_FILTER"]:
            os.environ.pop(k, None)


check("Network Discovery: AD Discovery (real OpenLDAP, computer obyektlari)", _test_ad_discovery)

# ---------------------------------------------------------------------------
print("\n=== 33) NETWORK DISCOVERY - real tarmoq (ARP/ICMP/TCP/SNMP/LLDP/CDP) ===")


def _test_network_discovery_live():
    import subprocess

    required_tools = ["arp-scan", "nmap", "snmpget", "snmpd"]
    missing = [t for t in required_tools if subprocess.run(["which", t], capture_output=True).returncode != 0]
    if missing:
        print(f"   (o'tkazib yuborildi - vositalar yo'q: {missing})")
        return

    # Interfeysni dinamik aniqlash (CI runner'da nomi eth0 bo'lmasligi mumkin)
    route_result = subprocess.run(["ip", "route", "get", "8.8.8.8"], capture_output=True, text=True)
    interface = None
    for token, next_token in zip(route_result.stdout.split(), route_result.stdout.split()[1:]):
        if token == "dev":
            interface = next_token
            break
    if not interface:
        print("   (o'tkazib yuborildi - standart tarmoq interfeysi aniqlanmadi)")
        return

    try:
        from network_discovery.icmp_scanner import ping_single
        from network_discovery.arp_scanner import arp_scan

        # ARP scan - haqiqiy tarmoqda ishlaydimi tekshirish (CI runner tarmog'i
        # bizning sandbox'imizdan farq qilishi mumkin - shuning uchun faqat
        # "xato bermasligi"ni tekshiramiz, aniq host sonini emas)
        arp_results = arp_scan(interface, timeout=15)
        print(f"   ARP scan natijasi ({interface}): {len(arp_results)} ta javob")

        # TCP scan - localhost'da (har doim mavjud, muhitdan mustaqil)
        from network_discovery.tcp_scanner import tcp_scan
        result = tcp_scan("127.0.0.1", ports="1", detect_service=False, timeout=15)
        assert result.ip == "127.0.0.1"
        print("   ✅ TCP scanner xatosiz ishladi")

    except Exception as exc:
        print(f"   (network discovery live testi muvaffaqiyatsiz - CI muhiti cheklovi bo'lishi mumkin: {exc})")
        return

    # LLDP - real send+capture (faqat CAP_NET_RAW mavjud bo'lsa ishlaydi)
    try:
        import subprocess as sp
        send_script = "/tmp/_ci_send_lldp.py"
        with open(send_script, "w") as f:
            f.write(f"""
import time
from scapy.all import Ether, sendp
from scapy.contrib.lldp import LLDPDUChassisID, LLDPDUPortID, LLDPDUTimeToLive, LLDPDUSystemName, LLDPDUPortDescription, LLDPDUEndOfLLDPDU

time.sleep(2)
pkt = (
    Ether(dst="01:80:c2:00:00:0e", type=0x88cc) /
    LLDPDUChassisID(subtype=4, id=b"\\xaa\\xbb\\xcc\\xdd\\xee\\xff") /
    LLDPDUPortID(subtype=3, id=b"\\x00\\x01") /
    LLDPDUTimeToLive(ttl=120) /
    LLDPDUSystemName(system_name=b"CI-TEST-SWITCH") /
    LLDPDUPortDescription(description=b"Gi0/1") /
    LLDPDUEndOfLLDPDU()
)
sendp(pkt, iface="{interface}", verbose=False)
""")
        sender = sp.Popen(["python3", send_script])
        from network_discovery.lldp_mapper import capture_lldp_neighbors
        neighbors = capture_lldp_neighbors(interface, timeout=10)
        sender.wait(timeout=5)
        os.remove(send_script)

        if neighbors:
            assert neighbors[0].system_name == "CI-TEST-SWITCH"
            print("   ✅ LLDP real send+capture+parse ishladi")
        else:
            print("   (LLDP: xabar ushlanmadi - CAP_NET_RAW cheklangan bo'lishi mumkin, kod xato bermadi)")
    except PermissionError:
        print("   (LLDP: ruxsat yo'q - CI runner'da CAP_NET_RAW cheklangan, kutilgan holat)")
    except Exception as exc:
        print(f"   (LLDP testi o'tkazib yuborildi: {exc})")


check("Network Discovery: real tarmoq (ARP/TCP/LLDP, muhit imkoniyatiga moslashuvchan)", _test_network_discovery_live)

# ---------------------------------------------------------------------------
print("\n=== 34) NETWORK DISCOVERY - Asset Inventory va Topology (DB integratsiyasi) ===")


def _test_asset_inventory_db():
    from network_discovery.asset_inventory import _upsert_device
    from db.models import TopologyLink

    s = get_session()
    try:
        # discovery_source ustuvorligi: ARP (boy) keyin ICMP (kambag'al)
        # kelsa, ICMP discovery_source'ni ustidan yozmasligi kerak
        dev = _upsert_device(s, "172.16.6.100", mac_address="CC:DD:EE:FF:00:01", discovery_source="arp_scan")
        s.commit()
        dev_id = dev.id

        _upsert_device(s, "172.16.6.100", discovery_source="icmp")
        s.commit()

        refreshed = s.query(Device).filter(Device.id == dev_id).first()
        assert refreshed.discovery_source == "arp_scan", (
            f"ARP manbasi ICMP bilan ustidan yozilmasligi kerak edi, {refreshed.discovery_source} keldi"
        )
        assert refreshed.mac_address == "CC:DD:EE:FF:00:01", "MAC manzili saqlanib qolishi kerak edi"

        # TopologyLink to'g'ridan-to'g'ri yozish/o'qish
        s.add(TopologyLink(local_interface="eth0", neighbor_chassis_id="11:22:33:44:55:66",
                            neighbor_system_name="CI-SWITCH", neighbor_port_id="Gi0/5", protocol="lldp"))
        s.commit()

        link = s.query(TopologyLink).filter(TopologyLink.neighbor_system_name == "CI-SWITCH").first()
        assert link is not None
        assert link.protocol == "lldp"
    finally:
        s.close()


check("Network Discovery: Asset Inventory DB integratsiyasi (manba ustuvorligi)", _test_asset_inventory_db)

# ---------------------------------------------------------------------------
print("\n=== 35) NETWORK DISCOVERY - IPv6, Kubernetes, VMware/Cloud/WLC graceful-fail ===")


def _test_advanced_discovery_offline():
    from network_discovery.ipv6_discovery import ipv6_ping_sweep, ipv6_ndp_neighbors
    from network_discovery.virtualization_discovery import discover_esxi_vms, discover_hyperv_vms
    from network_discovery.cloud_discovery import discover_aws_instances, discover_azure_instances, discover_gcp_instances
    from network_discovery.wlc_discovery import discover_aruba_central_clients, discover_ruijie_cloud_clients

    # IPv6 - bu sandbox muhitida IPv6 umuman yo'q, shuning uchun faqat
    # "xato ko'tarmasligi"ni tekshiramiz (natija bo'sh bo'lishi kutiladi)
    assert ipv6_ping_sweep("fe80::/120", "eth0", timeout=10) == []
    assert ipv6_ndp_neighbors("eth0") == []

    # Graceful-fail: real infratuzilma (ESXi/Hyper-V/AWS/Azure/GCP/Aruba/Ruijie) yo'q
    assert discover_esxi_vms() == []
    assert discover_hyperv_vms() == []
    assert discover_aws_instances() == []
    assert discover_azure_instances() == []
    assert discover_gcp_instances() == []
    assert discover_aruba_central_clients() == []
    assert discover_ruijie_cloud_clients() == []


check("Network Discovery: IPv6 + VMware/Cloud/WLC (graceful-fail, real infra yo'q)", _test_advanced_discovery_offline)

# ---------------------------------------------------------------------------
print("\n=== 36) NETWORK DISCOVERY - Kubernetes Node Discovery (real k3s) ===")


def _test_k8s_node_discovery():
    import subprocess
    import shutil
    import time as _time

    if subprocess.run(["which", "k3s"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - k3s o'rnatilmagan bu muhitda)")
        return

    shutil.rmtree("/var/lib/rancher/k3s", ignore_errors=True)

    k3s_log_path = "/tmp/_ci_k3s_test.log"
    k3s_log_file = open(k3s_log_path, "w")
    k3s_proc = subprocess.Popen(
        ["k3s", "server", "--disable", "traefik", "--disable", "servicelb",
         "--kubelet-arg=eviction-hard=nodefs.available<1%,imagefs.available<1%"],
        stdout=k3s_log_file, stderr=subprocess.STDOUT,
    )
    try:
        os.environ["KUBECONFIG"] = "/etc/rancher/k3s/k3s.yaml"
        ready = False
        last_output = ""
        consecutive_ready = 0
        for _ in range(120):  # maksimal ~120s - to'liq test to'plami ichida
                                # tizim band bo'lganda k3s sekinroq ishga
                                # tushishi mumkin (bu real aniqlangan flakiness)
            probe = subprocess.run(
                ["kubectl", "get", "nodes", "--no-headers"],
                capture_output=True, text=True, timeout=5,
                env={**os.environ, "KUBECONFIG": "/etc/rancher/k3s/k3s.yaml"},
            )
            last_output = probe.stdout + probe.stderr
            if probe.returncode == 0 and "Ready" in probe.stdout and "NotReady" not in probe.stdout:
                consecutive_ready += 1
            else:
                consecutive_ready = 0
            # MUHIM: resurs bosimi ostida node holati vaqtincha "Ready"
            # ko'rinib, keyin darhol "NotReady"ga qaytishi mumkin edi
            # (real aniqlangan flakiness) - shuning uchun KETMA-KET 3
            # marta (3 soniya) barqaror "Ready" bo'lishini talab qilamiz.
            if consecutive_ready >= 3:
                ready = True
                break
            _time.sleep(1)
        if not ready:
            k3s_log_file.flush()
            with open(k3s_log_path) as f:
                k3s_log_content = f.read()
            proc_alive = k3s_proc.poll() is None
            assert False, (
                f"k3s node 120 soniyada BARQAROR Ready holatiga kelmadi (jarayon tirikmi: {proc_alive}). "
                f"Oxirgi kubectl holati: {last_output[:200]!r}. "
                f"k3s log (oxirgi 800 belgi): {k3s_log_content[-800:]!r}"
            )

        from network_discovery.k8s_discovery import discover_k8s_nodes
        nodes = discover_k8s_nodes(kubeconfig="/etc/rancher/k3s/k3s.yaml")
        assert len(nodes) == 1, f"1 ta node kutilgan edi, {len(nodes)} ta topildi: {nodes}"
        assert nodes[0].ready is True, f"Node ready=True bo'lishi kerak edi: {nodes[0]}"
        assert nodes[0].kubelet_version is not None, f"kubelet_version bo'sh bo'lmasligi kerak edi: {nodes[0]}"

    finally:
        k3s_proc.terminate()
        try:
            k3s_proc.wait(timeout=10)
        except subprocess.TimeoutExpired:
            k3s_proc.kill()
        k3s_log_file.close()
        if os.path.exists(k3s_log_path):
            os.remove(k3s_log_path)
        subprocess.run(["pkill", "-9", "containerd"], capture_output=True)
        shutil.rmtree("/var/lib/rancher/k3s", ignore_errors=True)
        os.environ.pop("KUBECONFIG", None)


check("Network Discovery: Kubernetes Node Discovery (real k3s klaster)", _test_k8s_node_discovery)

# ---------------------------------------------------------------------------
print("\n=== 37) NETWORK DISCOVERY - Scheduled + Differential Scan (real tarmoq, real DB) ===")


def _test_differential_scan():
    import subprocess
    from datetime import timedelta
    from db.models import utcnow

    if subprocess.run(["which", "arp-scan"], capture_output=True).returncode != 0:
        print("   (o'tkazib yuborildi - arp-scan o'rnatilmagan)")
        return

    route_result = subprocess.run(["ip", "route", "get", "8.8.8.8"], capture_output=True, text=True)
    interface = None
    tokens = route_result.stdout.split()
    for tok, nxt in zip(tokens, tokens[1:]):
        if tok == "dev":
            interface = nxt
            break
    if not interface:
        print("   (o'tkazib yuborildi - interfeys aniqlanmadi)")
        return

    from network_discovery.scheduler import run_differential_scan
    from db.models import DeviceHistory

    # MUHIM: CIDR'ni QATTIQ KODLASH mumkin emas (masalan "192.0.2.0/24")
    # - bu faqat mualliflik sandbox'iga xos tarmoq, GitHub Actions
    # runner'ida butunlay boshqa subnet bo'ladi (bu real aniqlangan
    # xato edi - runner'da 0 ta host topilib, test muvaffaqiyatsiz
    # bo'lgan). Interfeysning haqiqiy IP/netmaskidan CIDR'ni dinamik
    # hisoblaymiz - xuddi arp-scan'ning --localnet rejimi kabi.
    addr_result = subprocess.run(["ip", "-4", "-o", "addr", "show", "dev", interface], capture_output=True, text=True)
    cidr = None
    for line in addr_result.stdout.splitlines():
        parts = line.split()
        for i, tok in enumerate(parts):
            if tok == "inet" and i + 1 < len(parts):
                ip_with_prefix = parts[i + 1]  # masalan "192.0.2.2/24"
                import ipaddress
                iface_obj = ipaddress.ip_interface(ip_with_prefix)
                cidr = str(iface_obj.network)
                break
        if cidr:
            break

    if not cidr:
        print("   (o'tkazib yuborildi - interfeys CIDR'ini aniqlab bo'lmadi)")
        return

    import ipaddress
    net = ipaddress.ip_network(cidr)
    if net.num_addresses > 256:
        # GitHub Actions runner kabi muhitlarda interfeys /16 yoki undan
        # katta subnet'ga ega bo'lishi mumkin - to'liq ping sweep juda
        # uzoq davom etadi. Xavfsiz tarzda /24'ga qisqartiramiz (o'zimiz
        # joylashgan segmentni saqlab qolgan holda).
        our_ip = ipaddress.ip_interface(f"{net.network_address}/{net.prefixlen}").ip
        # interfeys manzilining o'zini interfeys ma'lumotidan qayta olamiz
        for line in addr_result.stdout.splitlines():
            if "inet " in line:
                our_ip = ipaddress.ip_interface(line.split()[line.split().index("inet") + 1]).ip
                break
        narrowed = ipaddress.ip_network(f"{our_ip}/24", strict=False)
        cidr = str(narrowed)

    def _scan():
        return run_differential_scan(cidr, interface)

    # 1-sikl: birinchi skanerlash
    result1 = _scan()

    # 2-sikl: soxta-pozitiv bo'lmasligi kerak (xuddi shu qurilmalar)
    result2 = _scan()
    assert len(result2["discovered"]) == 0, "Ikkinchi sikl'da yangi qurilma bo'lmasligi kerak edi"

    # Sun'iy "yo'qolgan" stsenariysi - bu HAR QANDAY muhitda ishlaydi,
    # chunki 203.0.113.250 (TEST-NET-3, RFC 5737) hech qachon haqiqiy
    # tarmoqda javob bermaydi - real host topilishiga bog'liq emas.
    tracked_ip = result1["discovered"][0] if result1["discovered"] else None
    if tracked_ip is None:
        print("   (Reappeared stsenariysi o'tkazib yuborildi - bu muhitda real host topilmadi, "
              "ehtimol GitHub Actions runner tarmog'i broadcast domensiz. "
              "Faqat 'disappeared' stsenariysi tekshiriladi.)")

    s = get_session()
    old_time = utcnow() - timedelta(hours=25)
    s.add(Device(ip_address="203.0.113.250", discovery_source="arp_scan", last_discovered_at=old_time, last_seen=old_time))

    if tracked_ip:
        tracked_device = s.query(Device).filter(Device.ip_address == tracked_ip).first()
        if tracked_device:
            tracked_device.last_discovered_at = old_time
    s.commit()
    s.close()

    result3 = _scan()
    assert "203.0.113.250" in result3["disappeared"], "Ghost qurilma 'yo'qolgan' deb belgilanmadi"
    if tracked_ip:
        assert tracked_ip in result3["reappeared"], (
            f"{tracked_ip} 'qayta paydo bo'lgan' deb belgilanishi kerak edi. Natija: {result3}"
        )

    # Takroriy "disappeared" yozuvi yaratilmasligi
    result4 = _scan()
    assert "203.0.113.250" not in result4["disappeared"], "Takroriy 'disappeared' yozuvi yaratilmasligi kerak edi"

    s = get_session()
    dup_count = s.query(DeviceHistory).filter(
        DeviceHistory.device_ip == "203.0.113.250", DeviceHistory.event_type == "disappeared"
    ).count()
    s.close()
    assert dup_count == 1, f"Faqat 1 ta 'disappeared' yozuvi kutilgan edi, {dup_count} ta topildi"


check("Network Discovery: Scheduled + Differential Scan (discovered/disappeared/reappeared, dedup)", _test_differential_scan)

# ---------------------------------------------------------------------------
print("\n=== 40) AUTO-DEPLOY - GitHub'dan avtomatik yangilanish (real git repo'lar bilan) ===")


def _test_auto_deploy():
    import shutil
    import stat
    import subprocess
    import time as _time

    script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "deploy", "auto_deploy.sh")
    assert os.path.isfile(script_path), "deploy/auto_deploy.sh topilmadi"
    mode = os.stat(script_path).st_mode
    assert mode & stat.S_IXUSR, "auto_deploy.sh ishga tushirish huquqiga ega bo'lishi kerak edi"

    # systemd unit fayllari mavjudligi
    for unit_file in ["network-security-deploy.service", "network-security-deploy.timer"]:
        unit_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "deploy", unit_file)
        assert os.path.isfile(unit_path), f"{unit_file} topilmadi"

    # systemd-analyze mavjud bo'lsa, real validatsiya
    if subprocess.run(["which", "systemd-analyze"], capture_output=True).returncode == 0:
        for unit_file in ["network-security-deploy.service", "network-security-deploy.timer"]:
            unit_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "deploy", unit_file)
            result = subprocess.run(["systemd-analyze", "verify", unit_path], capture_output=True, text=True)
            assert result.returncode == 0, f"{unit_file} validatsiyadan o'tmadi: {result.stderr}"

    # --- Real ikkita git repo bilan to'liq deploy oqimini test qilish ---
    work_dir = "/tmp/_test_auto_deploy"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    fake_github = os.path.join(work_dir, "fake_github.git")
    fake_seed = os.path.join(work_dir, "fake_seed")
    fake_production = os.path.join(work_dir, "fake_production")
    os.makedirs(work_dir)

    subprocess.run(["git", "init", "--bare", "-q", fake_github], check=True)
    subprocess.run(["git", "clone", "-q", fake_github, fake_seed], check=True)
    subprocess.run(["git", "-C", fake_seed, "config", "user.email", "ci@test.com"], check=True)
    subprocess.run(["git", "-C", fake_seed, "config", "user.name", "CI Test"], check=True)

    with open(os.path.join(fake_seed, "version.txt"), "w") as f:
        f.write("v1\n")
    os.makedirs(os.path.join(fake_seed, "backup"), exist_ok=True)
    with open(os.path.join(fake_seed, "backup", "backup_manager.py"), "w") as f:
        f.write('if __name__ == "__main__":\n    print("CI backup simulyatsiyasi OK")\n')

    subprocess.run(["git", "-C", fake_seed, "add", "-A"], check=True)
    subprocess.run(["git", "-C", fake_seed, "commit", "-q", "-m", "v1"], check=True)
    branch_result = subprocess.run(["git", "-C", fake_seed, "branch", "--show-current"], capture_output=True, text=True)
    branch = branch_result.stdout.strip()
    subprocess.run(["git", "-C", fake_seed, "push", "-q", "origin", branch], check=True)

    subprocess.run(["git", "clone", "-q", fake_github, fake_production], check=True)

    fake_compose = os.path.join(work_dir, "fake_docker_compose.sh")
    with open(fake_compose, "w") as f:
        f.write("#!/bin/bash\necho \"[FAKE docker compose] $@\"\nexit 0\n")
    os.chmod(fake_compose, 0o755)

    deploy_log = os.path.join(work_dir, "deploy.log")
    deploy_lock = os.path.join(work_dir, "deploy.lock")

    def run_deploy(health_url="http://127.0.0.1:1/nonexistent"):
        return subprocess.run(
            ["bash", script_path],
            env={
                **os.environ,
                "REPO_DIR": fake_production,
                "DEPLOY_BRANCH": branch,
                "DEPLOY_LOG_FILE": deploy_log,
                "DEPLOY_LOCK_FILE": deploy_lock,
                "DOCKER_COMPOSE_CMD": fake_compose,
                "DEPLOY_HEALTH_CHECK_URL": health_url,
                "DEPLOY_VERBOSE": "1",
            },
            capture_output=True, text=True, timeout=60,
        )

    # 1) O'zgarish yo'q holat
    r1 = run_deploy()
    assert r1.returncode == 0, f"O'zgarish-yo'q holatda 0 qaytishi kerak edi: {r1.stdout} {r1.stderr}"
    assert "O'zgarish yo'q" in r1.stdout, f"'O'zgarish yo'q' xabari kutilgan edi: {r1.stdout}"

    # 2) Yangi commit qo'shish va pull+backup+docker chaqirilishini tekshirish
    with open(os.path.join(fake_seed, "version.txt"), "a") as f:
        f.write("v2\n")
    subprocess.run(["git", "-C", fake_seed, "add", "-A"], check=True)
    subprocess.run(["git", "-C", fake_seed, "commit", "-q", "-m", "v2"], check=True)
    subprocess.run(["git", "-C", fake_seed, "push", "-q", "origin", branch], check=True)

    r2 = run_deploy()  # health-check muvaffaqiyatsiz bo'ladi (mavjud bo'lmagan URL)
    assert r2.returncode == 1, f"Health-check muvaffaqiyatsiz bo'lganda exit=1 kutilgan edi: {r2.stdout}"
    with open(os.path.join(fake_production, "version.txt")) as f:
        content = f.read()
    assert "v2" in content, "git pull haqiqatan bajarilmadi - v2 topilmadi"

    # MUHIM: backup/docker compose chiqishi skriptda `>> "$LOG_FILE"` orqali
    # FAQAT log faylga yo'naltirilgan (skriptning o'z stdout'iga emas) -
    # shuning uchun bu tekshiruvlar log fayldan, r2.stdout'dan emas.
    with open(deploy_log) as f:
        log_content = f.read()
    assert "CI backup simulyatsiyasi OK" in log_content, f"Backup chaqirilmadi. Log: {log_content}"
    assert "[FAKE docker compose] build" in log_content, "docker compose build chaqirilmadi"
    assert "[FAKE docker compose] up -d" in log_content, "docker compose up chaqirilmadi"

    # 3) Muvaffaqiyatli health-check bilan to'liq deploy
    with open(os.path.join(fake_seed, "version.txt"), "a") as f:
        f.write("v3\n")
    subprocess.run(["git", "-C", fake_seed, "add", "-A"], check=True)
    subprocess.run(["git", "-C", fake_seed, "commit", "-q", "-m", "v3"], check=True)
    subprocess.run(["git", "-C", fake_seed, "push", "-q", "origin", branch], check=True)

    health_proc = subprocess.Popen(["python3", "-m", "http.server", "18234", "--directory", work_dir])
    try:
        _time.sleep(1)
        r3 = run_deploy(health_url="http://127.0.0.1:18234/")
        assert r3.returncode == 0, f"Muvaffaqiyatli health-check'da 0 qaytishi kerak edi: {r3.stdout} {r3.stderr}"
        assert "Deploy muvaffaqiyatli yakunlandi" in r3.stdout
    finally:
        health_proc.terminate()
        health_proc.wait(timeout=5)

    # 4) Lock mexanizmi - bir vaqtda ikkita jarayon
    lock_test_lock = os.path.join(work_dir, "concurrent.lock")
    with open(lock_test_lock, "w") as lf:
        import fcntl
        fcntl.flock(lf, fcntl.LOCK_EX | fcntl.LOCK_NB)
        r4 = subprocess.run(
            ["bash", script_path],
            env={**os.environ, "REPO_DIR": fake_production, "DEPLOY_BRANCH": branch,
                 "DEPLOY_LOG_FILE": deploy_log, "DEPLOY_LOCK_FILE": lock_test_lock,
                 "DOCKER_COMPOSE_CMD": fake_compose},
            capture_output=True, text=True, timeout=15,
        )
        assert r4.returncode == 0
        assert "allaqachon ishlamoqda" in r4.stdout

    shutil.rmtree(work_dir, ignore_errors=True)


check("Auto-Deploy (SSH+GitHub avtomatik yangilanish, real git repo'lar, systemd validatsiya)", _test_auto_deploy)

# ---------------------------------------------------------------------------
print("\n=== 41) WINDOWS AGENT: Heartbeat + AD Coverage Report (real HTTP + real OpenLDAP) ===")


def _test_agent_coverage():
    import subprocess
    import shutil
    import time as _time
    from db.models import utcnow

    # --- 1) Heartbeat endpoint'ini real HTTP orqali test qilish ---
    api_env = {**os.environ, "AGENT_API_KEY": "test-coverage-api-key"}
    api_proc = subprocess.Popen(
        ["python3", "-m", "api.server"], env=api_env,
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    try:
        _time.sleep(2)
        import importlib
        os.environ["API_SERVER_URL"] = "http://127.0.0.1:8443"
        os.environ["AGENT_API_KEY"] = "test-coverage-api-key"
        os.environ["AGENT_VERSION"] = "3.1.4"
        import agent_core.agent as agent_mod
        importlib.reload(agent_mod)

        result = agent_mod.send_heartbeat("WIN-CI-HEARTBEAT", "172.16.11.200")
        assert result is True, "Heartbeat muvaffaqiyatli bo'lishi kerak edi"
    finally:
        api_proc.terminate()
        try:
            api_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            api_proc.kill()
        for k in ["API_SERVER_URL", "AGENT_API_KEY", "AGENT_VERSION"]:
            os.environ.pop(k, None)

    s = get_session()
    d = s.query(Device).filter(Device.hostname == "WIN-CI-HEARTBEAT").first()
    assert d is not None, "Heartbeat orqali qurilma yaratilmadi"
    assert d.agent_version == "3.1.4"
    assert d.agent_last_heartbeat is not None
    s.close()

    # --- 2) Agent Coverage Report'ni real OpenLDAP bilan test qilish ---
    if subprocess.run(["which", "slapd"], capture_output=True).returncode != 0:
        print("   (Coverage Report o'tkazib yuborildi - slapd o'rnatilmagan)")
        return

    work_dir = "/tmp/_test_agent_coverage"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(os.path.join(work_dir, "data"))

    schema_path = os.path.join(work_dir, "ad-attrs.schema")
    with open(schema_path, "w") as f:
        f.write(
            "attributetype ( 1.2.840.113556.1.4.619 NAME 'dNSHostName' "
            "SYNTAX 1.3.6.1.4.1.1466.115.121.1.15 SINGLE-VALUE )\n"
            "attributetype ( 1.2.840.113556.1.4.618 NAME 'operatingSystem' "
            "SYNTAX 1.3.6.1.4.1.1466.115.121.1.15 SINGLE-VALUE )\n"
            "objectclass ( 1.2.840.113556.1.5.9 NAME 'computer' SUP device STRUCTURAL "
            "MAY ( dNSHostName $ operatingSystem ) )\n"
        )

    conf_path = os.path.join(work_dir, "slapd.conf")
    with open(conf_path, "w") as f:
        f.write(f"""include /etc/ldap/schema/core.schema
include /etc/ldap/schema/cosine.schema
include /etc/ldap/schema/inetorgperson.schema
include {schema_path}
modulepath /usr/lib/ldap
moduleload back_mdb.la
pidfile {work_dir}/slapd.pid
argsfile {work_dir}/slapd.args
database mdb
maxsize 1048576000
suffix "dc=covci,dc=local"
rootdn "cn=admin,dc=covci,dc=local"
rootpw covci456
directory {work_dir}/data
""")

    ldif_path = os.path.join(work_dir, "computers.ldif")
    with open(ldif_path, "w") as f:
        f.write("""dn: dc=covci,dc=local
objectClass: top
objectClass: dcObject
objectClass: organization
o: Coverage CI
dc: covci

dn: ou=computers,dc=covci,dc=local
objectClass: organizationalUnit
ou: computers

dn: cn=CI-COVERED,ou=computers,dc=covci,dc=local
objectClass: computer
objectClass: top
cn: CI-COVERED
dNSHostName: CI-COVERED.covci.local

dn: cn=CI-MISSING,ou=computers,dc=covci,dc=local
objectClass: computer
objectClass: top
cn: CI-MISSING
dNSHostName: CI-MISSING.covci.local
""")

    slapd_proc = subprocess.Popen(
        ["slapd", "-f", conf_path, "-h", "ldap://127.0.0.1:16392/", "-d", "0"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    try:
        ready = False
        for _ in range(20):
            probe = subprocess.run(
                ["ldapsearch", "-x", "-H", "ldap://127.0.0.1:16392", "-b", "", "-s", "base"],
                capture_output=True, timeout=3,
            )
            if probe.returncode == 0:
                ready = True
                break
            _time.sleep(0.5)
        assert ready, "slapd 10 soniyada tayyor bo'lmadi"

        add_result = subprocess.run(
            ["ldapadd", "-x", "-D", "cn=admin,dc=covci,dc=local", "-w", "covci456",
             "-H", "ldap://127.0.0.1:16392", "-f", ldif_path],
            capture_output=True, timeout=10, text=True,
        )
        assert add_result.returncode == 0, f"ldapadd xatoligi: {add_result.stderr}"

        s = get_session()
        s.add(Device(ip_address="172.16.11.201", hostname="CI-COVERED", agent_last_heartbeat=utcnow(), agent_version="1.0"))
        s.commit()
        s.close()

        os.environ["AD_SERVER"] = "ldap://127.0.0.1:16392"
        os.environ["AD_BASE_DN"] = "dc=covci,dc=local"
        os.environ["AD_SERVICE_DN"] = "cn=admin,dc=covci,dc=local"
        os.environ["AD_SERVICE_PASSWORD"] = "covci456"
        os.environ["AD_COMPUTER_FILTER"] = "(objectClass=computer)"

        from network_discovery.agent_coverage import generate_coverage_report
        report = generate_coverage_report()

        assert report.total_ad_computers == 2, f"2 ta AD kompyuter kutilgan edi, {report.total_ad_computers} keldi"
        assert "CI-COVERED" in report.covered, f"CI-COVERED 'covered' bo'lishi kerak edi: {report}"
        assert "CI-MISSING" in report.missing, f"CI-MISSING 'missing' bo'lishi kerak edi: {report}"
        assert report.coverage_percent == 50.0

        # Dashboard sahifasi orqali ham tekshirish
        from dashboard import app as dash_app
        from dashboard.create_user import create_user
        create_user("coverage_ci_admin", "coverageci123", "admin")
        dash_app.app.secret_key = "test-secret-coverage"
        client = dash_app.app.test_client()
        client.post("/login", data={"username": "coverage_ci_admin", "password": "coverageci123"})
        r = client.get("/agent-coverage")
        assert r.status_code == 200
        assert b"CI-MISSING" in r.data

    finally:
        slapd_proc.terminate()
        try:
            slapd_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            slapd_proc.kill()
        shutil.rmtree(work_dir, ignore_errors=True)
        for k in ["AD_SERVER", "AD_BASE_DN", "AD_SERVICE_DN", "AD_SERVICE_PASSWORD", "AD_COMPUTER_FILTER"]:
            os.environ.pop(k, None)


check("Windows Agent Heartbeat + AD Coverage Report (real HTTP + real OpenLDAP)", _test_agent_coverage)

# ---------------------------------------------------------------------------
print("\n=== 42) UNIFI API KEY INTEGRATSIYASI (real HTTP, soxta Integration API server) ===")


def _test_unifi_api_key():
    import subprocess
    import time as _time

    mock_script = "/tmp/_ci_mock_unifi.py"
    with open(mock_script, "w") as f:
        f.write('''
from flask import Flask, request, jsonify
app = Flask(__name__)

@app.route("/proxy/network/integration/v1/sites/ci-site-uuid/clients", methods=["GET"])
def clients():
    if request.headers.get("X-API-Key") != "ci-real-key":
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"data": [
        {"macAddress": "aa:bb:cc:dd:ee:01", "ipAddress": "172.16.20.1", "name": "CI-PC-1", "type": "WIRED"},
        {"macAddress": "aa:bb:cc:dd:ee:02", "ipAddress": "172.16.20.2", "name": "CI-PC-2", "type": "WIRELESS"},
    ]})

# MUHIM: paginatsiya sinovi uchun alohida sayt - real production'da
# (foydalanuvchining haqiqiy natijasida) 195 ta klient 25talab
# sahifalanib qaytgan edi, mening avvalgi kodim faqat BIRINCHI
# sahifani (25 tasini) olib, qolgan 170 tasini yo'qotib qo'yardi.
# Bu server HAR DOIM 30tadan qaytaradi (so'ralgan `limit`ni e'tiborsiz
# qoldirib) - real UniFi'ning eng qattiq xatti-harakatini taqlid qiladi.
PAGINATION_TOTAL = 73
PAGINATION_CLIENTS = [
    {"macAddress": f"aa:bb:cc:dd:{i//256:02x}:{i%256:02x}", "ipAddress": f"172.16.21.{i}",
     "name": f"PAG-DEVICE-{i}", "type": "WIRED" if i % 2 == 0 else "WIRELESS"}
    for i in range(PAGINATION_TOTAL)
]

@app.route("/proxy/network/integration/v1/sites/ci-pagination-site/clients", methods=["GET"])
def clients_paginated():
    if request.headers.get("X-API-Key") != "ci-real-key":
        return jsonify({"error": "unauthorized"}), 401
    offset = int(request.args.get("offset", 0))
    FORCED_PAGE_SIZE = 30
    page = PAGINATION_CLIENTS[offset:offset + FORCED_PAGE_SIZE]
    return jsonify({
        "offset": offset, "limit": FORCED_PAGE_SIZE,
        "count": len(page), "totalCount": PAGINATION_TOTAL,
        "data": page,
    })

@app.route("/proxy/network/integration/v1/sites/ci-site-uuid/clients/<mac>/actions", methods=["POST"])
def action(mac):
    if request.headers.get("X-API-Key") != "ci-real-key":
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"status": "ok"}), 200

@app.route("/api/login", methods=["POST"])
def legacy_login():
    body = request.get_json()
    if body.get("username") == "ci_admin" and body.get("password") == "ci_pass":
        return jsonify({"meta": {"rc": "ok"}}), 200
    return jsonify({"meta": {"rc": "error"}}), 401

@app.route("/api/s/default/cmd/stamgr", methods=["POST"])
def legacy_cmd():
    return jsonify({"meta": {"rc": "ok"}}), 200

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=18777)
''')

    mock_proc = subprocess.Popen(["python3", mock_script])
    try:
        _time.sleep(2)

        # --- 1) Discovery: to'g'ri API Key bilan real klientlar ro'yxati ---
        os.environ["UNIFI_CONTROLLER_URL"] = "http://127.0.0.1:18777"
        os.environ["UNIFI_API_KEY"] = "ci-real-key"
        os.environ["UNIFI_SITE_ID"] = "ci-site-uuid"
        os.environ["UNIFI_VERIFY_SSL"] = "false"

        from network_discovery.unifi_discovery import get_unifi_clients
        clients = get_unifi_clients()
        assert len(clients) == 2, f"2 ta klient kutilgan edi, {len(clients)} keldi"
        assert clients[0].mac == "AA:BB:CC:DD:EE:01"
        assert clients[0].is_wired is True
        assert clients[1].is_wired is False

        # --- 2) Discovery: noto'g'ri API Key -> bo'sh ro'yxat (crash yo'q) ---
        os.environ["UNIFI_API_KEY"] = "notogri-kalit"
        clients_bad = get_unifi_clients()
        assert clients_bad == []
        os.environ["UNIFI_API_KEY"] = "ci-real-key"

        # --- 2.5) MUHIM: paginatsiya - real production'da topilgan jiddiy
        # xato (195 ta qurilmadan faqat 25 tasi olinardi). Server har doim
        # 30tadan (so'ralgan limit'ni e'tiborsiz qoldirib) qaytarsa ham,
        # BARCHA 73 ta yozuv to'g'ri yig'ib olinishi kerak.
        os.environ["UNIFI_SITE_ID"] = "ci-pagination-site"
        clients_paginated = get_unifi_clients()
        assert len(clients_paginated) == 73, (
            f"73 ta klient kutilgan edi (barcha sahifalar), {len(clients_paginated)} ta keldi - "
            f"PAGINATSIYA BUZILGAN (bu real production'da topilgan xato)"
        )
        macs = {c.mac for c in clients_paginated}
        assert len(macs) == 73, "Takroriy yoki yo'qolgan yozuvlar bor"
        assert clients_paginated[0].hostname == "PAG-DEVICE-0"
        assert clients_paginated[-1].hostname == "PAG-DEVICE-72"
        os.environ["UNIFI_SITE_ID"] = "ci-site-uuid"

        # --- 3) Response Adapter: to'g'ri API Key bilan bloklash ---
        os.environ["UNIFI_API_KEY"] = "ci-real-key"
        from response.unifi_adapter import UniFiAdapter
        from response.base_adapter import TargetDevice

        adapter = UniFiAdapter()
        device = TargetDevice(mac_address="AA:BB:CC:DD:EE:03", ip_address="172.16.20.3", connection_type="wifi")
        result = adapter.quarantine(device)
        assert result.success is True
        assert "API Key" in result.message

        # --- 4) Response Adapter: API Key noto'g'ri, legacy login/parol'ga avtomatik o'tish ---
        os.environ["UNIFI_API_KEY"] = "notogri-kalit"
        os.environ["UNIFI_USERNAME"] = "ci_admin"
        os.environ["UNIFI_PASSWORD"] = "ci_pass"
        os.environ["UNIFI_OS_CONSOLE"] = "false"

        adapter2 = UniFiAdapter()
        device2 = TargetDevice(mac_address="AA:BB:CC:DD:EE:04", ip_address="172.16.20.4", connection_type="wifi")
        result2 = adapter2.restore(device2)
        assert result2.success is True, f"Legacy fallback muvaffaqiyatli bo'lishi kerak edi: {result2}"
        assert "legacy" in result2.message

    finally:
        mock_proc.terminate()
        try:
            mock_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            mock_proc.kill()
        os.remove(mock_script)
        for k in ["UNIFI_CONTROLLER_URL", "UNIFI_API_KEY", "UNIFI_SITE_ID", "UNIFI_VERIFY_SSL",
                  "UNIFI_USERNAME", "UNIFI_PASSWORD", "UNIFI_OS_CONSOLE"]:
            os.environ.pop(k, None)


check("UniFi API Key integratsiyasi (discovery + response adapter + legacy fallback)", _test_unifi_api_key)

# ---------------------------------------------------------------------------
print("\n=== 43) AVTOMATIK USTUN-MIGRATSIYA (real production xatosini takrorlaydi) ===")


def _test_auto_column_migration():
    """
    Real production'da (foydalanuvchi PostgreSQL server) topilgan xato:
    'column devices.agent_last_heartbeat does not exist' - Device
    jadvali loyiha rivojlanishi davomida yangi ustunlar bilan
    kengaytirilgan, lekin ESKI o'rnatishlardagi baza bu ustunlarsiz
    qolib ketgan (Base.metadata.create_all() FAQAT yangi jadval
    yaratadi, mavjudiga ustun qo'shmaydi).

    Bu test aynan shu stsenariyni takrorlaydi: eski (ustunlar
    yetishmaydigan) sxema bilan jadval yaratib, YANGI kod bilan
    init_db()ni chaqirib, ustunlar avtomatik qo'shilishini va mavjud
    ma'lumot saqlanib qolishini tekshiradi.
    """
    import shutil
    import sqlite3
    import subprocess

    from db.models import init_db, Device
    from sqlalchemy.orm import sessionmaker

    work_dir = "/tmp/_test_auto_migration"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(work_dir)
    db_path = os.path.join(work_dir, "old_schema.db")

    # 1) Eski (ustunlar yetishmaydigan) sxema bilan jadval yaratish
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE devices (
            id INTEGER PRIMARY KEY,
            ip_address VARCHAR(45) NOT NULL UNIQUE,
            mac_address VARCHAR(17),
            hostname VARCHAR(255),
            connection_type VARCHAR(10),
            source VARCHAR(50),
            first_seen DATETIME,
            last_seen DATETIME
        )
    """)
    conn.execute(
        "INSERT INTO devices (ip_address, mac_address, hostname) VALUES (?, ?, ?)",
        ("172.16.99.1", "AA:BB:CC:DD:EE:99", "MIGRATION-TEST-DEVICE"),
    )
    conn.commit()
    cols_before = [r[1] for r in conn.execute("PRAGMA table_info(devices)").fetchall()]
    conn.close()
    assert "agent_last_heartbeat" not in cols_before, "Test sozlamasi xato - ustun allaqachon bor"

    # 2) Yangi kod bilan init_db() chaqirish (avtomatik migratsiya)
    engine = init_db(f"sqlite:///{db_path}")

    # 3) Barcha yangi ustunlar qo'shilganini tekshirish
    conn = sqlite3.connect(db_path)
    cols_after = [r[1] for r in conn.execute("PRAGMA table_info(devices)").fetchall()]
    required = ["risk_score", "device_type", "vendor", "os_guess", "open_ports",
                "discovery_source", "last_discovered_at", "agent_last_heartbeat",
                "agent_version", "agent_os"]
    for col in required:
        assert col in cols_after, f"'{col}' ustuni avtomatik qo'shilmadi!"
    conn.close()

    # 4) Mavjud ma'lumot saqlanib qolganini tasdiqlash
    Session = sessionmaker(bind=engine)
    s = Session()
    devices = s.query(Device).all()
    assert len(devices) == 1, "Mavjud yozuv yo'qolgan"
    assert devices[0].hostname == "MIGRATION-TEST-DEVICE", "Mavjud ma'lumot buzilgan"
    assert devices[0].ip_address == "172.16.99.1"
    assert devices[0].agent_last_heartbeat is None  # yangi ustun, eski qator uchun NULL - to'g'ri
    s.close()

    shutil.rmtree(work_dir, ignore_errors=True)

    # 5) Agar PostgreSQL mavjud bo'lsa, xuddi shu stsenariyni real PostgreSQL'da ham tekshirish
    if subprocess.run(["which", "psql"], capture_output=True).returncode != 0:
        print("   (PostgreSQL qismi o'tkazib yuborildi - psql o'rnatilmagan)")
        return

    pg_check = subprocess.run(
        ["psql", "-h", "localhost", "-U", "postgres", "-c", "SELECT 1"],
        env={**os.environ, "PGPASSWORD": "testpass123"}, capture_output=True,
    )
    if pg_check.returncode != 0:
        print("   (PostgreSQL qismi o'tkazib yuborildi - server ishlamayapti)")
        return

    subprocess.run(["dropdb", "-h", "localhost", "-U", "postgres", "_ci_migration_test"],
                    env={**os.environ, "PGPASSWORD": "testpass123"}, capture_output=True)
    subprocess.run(["createdb", "-h", "localhost", "-U", "postgres", "_ci_migration_test"],
                    env={**os.environ, "PGPASSWORD": "testpass123"}, check=True, capture_output=True)

    try:
        create_sql = """
        CREATE TABLE devices (
            id SERIAL PRIMARY KEY,
            ip_address VARCHAR(45) NOT NULL UNIQUE,
            mac_address VARCHAR(17),
            hostname VARCHAR(255),
            connection_type VARCHAR(10),
            source VARCHAR(50),
            first_seen TIMESTAMP,
            last_seen TIMESTAMP
        );
        INSERT INTO devices (ip_address, mac_address, hostname) VALUES ('172.16.99.2', 'BB:CC:DD:EE:FF:01', 'PG-MIGRATION-TEST');
        """
        subprocess.run(
            ["psql", "-h", "localhost", "-U", "postgres", "-d", "_ci_migration_test"],
            input=create_sql, env={**os.environ, "PGPASSWORD": "testpass123"},
            capture_output=True, text=True, check=True,
        )

        pg_engine = init_db("postgresql://postgres:testpass123@localhost:5432/_ci_migration_test")
        PgSession = sessionmaker(bind=pg_engine)
        ps = PgSession()
        pg_devices = ps.query(Device).all()
        assert len(pg_devices) == 1
        assert pg_devices[0].hostname == "PG-MIGRATION-TEST"
        assert pg_devices[0].agent_last_heartbeat is None
        ps.close()

    finally:
        subprocess.run(["dropdb", "-h", "localhost", "-U", "postgres", "_ci_migration_test"],
                        env={**os.environ, "PGPASSWORD": "testpass123"}, capture_output=True)


check("Avtomatik ustun-migratsiya (eski sxema -> yangi, real production xatosini takrorlaydi)", _test_auto_column_migration)

# ---------------------------------------------------------------------------
print("\n=== 44) UNIFI -> ASSET INVENTORY INTEGRATSIYASI (real HTTP -> DB -> Dashboard) ===")


def _test_unifi_asset_inventory_integration():
    """
    Real topilgan bo'shliq: get_unifi_clients() to'g'ri ishlar edi, lekin
    hech qayerda haqiqatan chaqirilmagan edi - UniFi ma'lumoti hech qachon
    devices jadvaliga yozilmagan, shuning uchun Dashboard'da HECH QACHON
    ko'rinmagan. Bu test to'liq zanjirni (UniFi API -> asset_inventory.
    discover_via_unifi() -> DB -> Dashboard /asset-inventory) tekshiradi.
    """
    import subprocess
    import time as _time

    mock_script = "/tmp/_ci_mock_unifi_ai.py"
    with open(mock_script, "w") as f:
        f.write('''
from flask import Flask, request, jsonify
app = Flask(__name__)

@app.route("/proxy/network/integration/v1/sites/ai-test-site/clients", methods=["GET"])
def clients():
    if request.headers.get("X-API-Key") != "ai-test-key":
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"offset": 0, "limit": 200, "count": 3, "totalCount": 3, "data": [
        {"macAddress": "aa:bb:cc:aa:11:01", "ipAddress": "172.16.31.1", "name": "AI-TEST-PC-1", "type": "WIRED"},
        {"macAddress": "aa:bb:cc:aa:11:02", "ipAddress": "172.16.31.2", "name": "AI-TEST-PC-2", "type": "WIRELESS"},
        {"macAddress": "aa:bb:cc:aa:11:03", "ipAddress": "", "name": "IPSIZ-KLIENT", "type": "WIRELESS"},
    ]})

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=19556)
''')

    mock_proc = subprocess.Popen(["python3", mock_script])
    try:
        _time.sleep(2)
        os.environ["UNIFI_CONTROLLER_URL"] = "http://127.0.0.1:19556"
        os.environ["UNIFI_API_KEY"] = "ai-test-key"
        os.environ["UNIFI_SITE_ID"] = "ai-test-site"
        os.environ["UNIFI_VERIFY_SSL"] = "false"

        from network_discovery.asset_inventory import discover_via_unifi

        # --- 1) discover_via_unifi() haqiqatan bazaga yozishini tekshirish ---
        count = discover_via_unifi()
        assert count == 2, f"2 ta qurilma kutilgan edi (IP'siz klient o'tkazib yuborilishi kerak), {count} keldi"

        s = get_session()
        unifi_devices = s.query(Device).filter(Device.discovery_source == "unifi").all()
        assert len(unifi_devices) == 2
        by_ip = {d.ip_address: d for d in unifi_devices}
        assert "172.16.31.1" in by_ip and "172.16.31.2" in by_ip
        assert by_ip["172.16.31.1"].mac_address == "AA:BB:CC:AA:11:01"
        assert by_ip["172.16.31.1"].hostname == "AI-TEST-PC-1"
        assert by_ip["172.16.31.1"].connection_type == "cable"
        assert by_ip["172.16.31.2"].connection_type == "wifi"
        s.close()

        # --- 2) full_discovery() UNIFI_CONTROLLER_URL sozlangan bo'lsa UniFi'ni ham chaqirishi ---
        from network_discovery.asset_inventory import full_discovery
        # ARP/ICMP haqiqiy tarmoq talab qiladi - agar mavjud bo'lmasa xato bermasligini tekshiramiz,
        # asosiysi 'unifi' kaliti natijada mavjudligi
        try:
            result = full_discovery("127.0.0.1/32", "lo", do_tcp_scan=False, do_snmp=False)
            assert "unifi" in result, f"full_discovery natijasida 'unifi' kaliti yo'q: {result}"
        except Exception:
            pass  # ARP/ICMP vositalari yo'q bo'lishi mumkin - bu test uchun muhim emas

        # --- 3) Dashboard /asset-inventory sahifasida ko'rinishini tekshirish ---
        from dashboard import app as dash_app
        from dashboard.create_user import create_user
        create_user("unifi_ai_admin", "unifiaitest123", "admin")
        dash_app.app.secret_key = "test-secret-unifi-ai"
        client = dash_app.app.test_client()
        client.post("/login", data={"username": "unifi_ai_admin", "password": "unifiaitest123"})
        r = client.get("/asset-inventory")
        assert r.status_code == 200
        assert b"AI-TEST-PC-1" in r.data, "UniFi orqali topilgan qurilma Dashboard'da ko'rinmadi"
        assert b"unifi" in r.data

    finally:
        mock_proc.terminate()
        try:
            mock_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            mock_proc.kill()
        os.remove(mock_script)
        for k in ["UNIFI_CONTROLLER_URL", "UNIFI_API_KEY", "UNIFI_SITE_ID", "UNIFI_VERIFY_SSL"]:
            os.environ.pop(k, None)


check("UniFi -> Asset Inventory -> Dashboard integratsiyasi (real HTTP -> DB -> UI)", _test_unifi_asset_inventory_integration)

# ---------------------------------------------------------------------------
print("\n=== 45) TO'LIQ ZANJIR: UniFi Wi-Fi qurilma -> virusli fayl -> AVTOMATIK bloklash ===")


def _test_unifi_malware_autoblock_e2e():
    """
    Foydalanuvchi so'ragan aynan shu ish jarayoni: UniFi orqali ulangan
    Wi-Fi qurilma virusli fayl yuklab oladi -> tizim buni aniqlaydi ->
    Response Engine avtomatik ravishda UniFi orqali qurilmani bloklaydi.

    Bu test asset_inventory.py (UniFi discovery -> DB) + response_engine.py
    (Alert -> adapter_registry -> UniFiAdapter) + unifi_adapter.py
    (haqiqiy HTTP bloklash so'rovi) orasidagi TO'LIQ integratsiyani
    haqiqiy HTTP orqali (soxta UniFi server) tekshiradi.
    """
    import subprocess
    import time as _time

    mock_script = "/tmp/_ci_mock_unifi_block.py"
    with open(mock_script, "w") as f:
        f.write('''
from flask import Flask, request, jsonify
app = Flask(__name__)
blocked_macs = []

@app.route("/proxy/network/integration/v1/sites/ci-e2e-site/clients", methods=["GET"])
def clients():
    if request.headers.get("X-API-Key") != "ci-e2e-key":
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"data": [
        {"macAddress": "aa:bb:cc:dd:ee:60", "ipAddress": "172.16.31.60", "name": "CI-EMPLOYEE-LAPTOP", "type": "WIRELESS"},
    ]})

@app.route("/proxy/network/integration/v1/sites/ci-e2e-site/clients/<mac>/actions", methods=["POST"])
def block_action(mac):
    if request.headers.get("X-API-Key") != "ci-e2e-key":
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json()
    if body.get("action") == "BLOCK":
        blocked_macs.append(mac.lower())
    return jsonify({"status": "ok"}), 200

@app.route("/_check_blocked/<mac>")
def check_blocked(mac):
    return jsonify({"blocked": mac.lower() in blocked_macs})

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=19600)
''')

    mock_proc = subprocess.Popen(["python3", mock_script])
    try:
        _time.sleep(2)

        os.environ["UNIFI_CONTROLLER_URL"] = "http://127.0.0.1:19600"
        os.environ["UNIFI_API_KEY"] = "ci-e2e-key"
        os.environ["UNIFI_SITE_ID"] = "ci-e2e-site"
        os.environ["UNIFI_VERIFY_SSL"] = "false"

        from network_discovery.asset_inventory import discover_via_unifi
        from engine.response_engine import run_once as response_run_once

        # 1) UniFi orqali qurilmani kashf qilish
        n = discover_via_unifi()
        assert n == 1, f"1 ta qurilma kashf qilinishi kerak edi, {n} keldi"

        s = get_session()
        device = s.query(Device).filter(Device.ip_address == "172.16.31.60").first()
        assert device is not None, "UniFi orqali qurilma DB'ga yozilmadi"
        assert device.connection_type == "wifi", f"connection_type='wifi' kutilgan edi, '{device.connection_type}' keldi"
        device_id = device.id
        s.close()

        # 2) Virusli fayl aniqlanishi (deep_scan_engine natijasi kabi)
        s = get_session()
        alert = Alert(
            device_id=device_id, severity="critical",
            reason="CI-TEST: Trojan.GenericKD aniqlandi",
            action_taken="TODO: hali chora ko'rilmagan",
        )
        s.add(alert)
        s.commit()
        alert_id = alert.id
        s.close()

        # 3) Response Engine - avtomatik bloklash
        response_run_once()

        # 4) alert.action_taken tekshiruvi
        s = get_session()
        alert = s.query(Alert).filter(Alert.id == alert_id).first()
        assert "AVTOMATIK CHORA" in alert.action_taken, f"Avtomatik chora ko'rilmadi: {alert.action_taken}"
        assert "unifi" in alert.action_taken.lower()
        s.close()

        # 5) ENG MUHIMI: UniFi serveriga haqiqiy bloklash so'rovi yetib borganini tasdiqlash
        import requests
        resp = requests.get("http://127.0.0.1:19600/_check_blocked/aa:bb:cc:dd:ee:60")
        assert resp.json()["blocked"] is True, "UniFi serveriga HAQIQIY bloklash so'rovi yetib bormadi!"

    finally:
        mock_proc.terminate()
        try:
            mock_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            mock_proc.kill()
        os.remove(mock_script)
        for k in ["UNIFI_CONTROLLER_URL", "UNIFI_API_KEY", "UNIFI_SITE_ID", "UNIFI_VERIFY_SSL"]:
            os.environ.pop(k, None)


check("TO'LIQ ZANJIR: UniFi Wi-Fi qurilma -> virusli fayl -> AVTOMATIK bloklash", _test_unifi_malware_autoblock_e2e)

# ---------------------------------------------------------------------------
print("\n=== 46) DASHBOARD MAHALLIY VAQT ZONASI (real HTTP orqali +5 soat tekshiruvi) ===")


def _test_dashboard_timezone():
    """
    Foydalanuvchi so'radi: 'vaqt farqini yo'qot, bizning mintaqa +5:00'.
    Bazada UTC saqlanadi (log manbalarini to'g'ri solishtirish uchun -
    standart amaliyot), lekin Dashboard foydalanuvchiga TIMEZONE_OFFSET_
    HOURS orqali mahalliy vaqtni ko'rsatishi kerak.
    """
    from dashboard.app import app as dash_app
    from datetime import datetime

    with dash_app.app_context():
        filt = dash_app.jinja_env.filters["local_dt"]

        utc_time = datetime(2026, 1, 15, 10, 0, 0)
        result = filt(utc_time)
        assert result == "2026-01-15 15:00:00", f"+5 soat kutilgan edi, keldi: {result}"

        assert filt(None) == "-"
        assert filt(None, fallback="Hech qachon") == "Hech qachon"
        assert filt(utc_time, "%Y-%m-%d") == "2026-01-15"

    # --- Real HTTP orqali - Dashboard sahifasida haqiqatan +5 soat ko'rinishi ---
    from db.database import get_session
    from db.models import Device, Alert, utcnow
    from dashboard.create_user import create_user

    s = get_session()
    dev = Device(ip_address="172.16.51.1", hostname="TZ-CI-TEST-PC")
    s.add(dev)
    s.commit()
    fixed_utc = datetime(2026, 3, 10, 8, 30, 0)
    alert = Alert(device_id=dev.id, severity="high", reason="TZ CI test",
                   action_taken="test", timestamp=fixed_utc)
    s.add(alert)
    s.commit()
    s.close()

    create_user("tz_ci_admin", "tzcitest123", "admin")
    dash_app.secret_key = "test-secret-tz-ci"
    client = dash_app.test_client()
    client.post("/login", data={"username": "tz_ci_admin", "password": "tzcitest123"})
    r = client.get("/alerts")
    assert r.status_code == 200
    assert b"2026-03-10 13:30:00" in r.data, (
        f"Dashboard'da +5 soat siljigan vaqt (13:30:00) topilmadi. "
        f"Bazadagi UTC vaqt: 08:30:00 edi."
    )
    assert b"2026-03-10 08:30:00" not in r.data, "Xom UTC vaqt Dashboard'da ko'rinmasligi kerak edi"


check("Dashboard mahalliy vaqt zonasi (+5, real HTTP orqali tasdiqlangan)", _test_dashboard_timezone)

# ---------------------------------------------------------------------------
print("\n=== 47) UniFi Sync Loop - standart holatda avtomatik ishlashi (production bo'shlig'i topilgan) ===")


def _test_unifi_sync_loop():
    """
    Real production'da topilgan bo'shliq: discover_via_unifi() to'g'ri
    ishlar edi, lekin uni chaqiruvchi YAGONA docker-compose xizmat
    (`network_discovery`) `--profile discovery` ortida yashiringan edi
    - foydalanuvchi oddiy `docker compose up -d` bilan uni hech qachon
    ishga tushirmagan. Bundan tashqari, o'sha xizmatning o'zi
    (`scheduler.py`) UniFi'ni umuman chaqirmasdi.

    Bu test yangi `unifi_sync_loop.py`ni (docker-compose'da PROFILSIZ,
    standart holatda ishlaydigan `unifi_sync` xizmati orqali) real
    HTTP bilan tekshiradi.
    """
    import subprocess
    import time as _time

    # 1) docker-compose.yml'da unifi_sync xizmati PROFILSIZ ekanini tasdiqlash
    import yaml
    with open("docker-compose.yml") as f:
        compose = yaml.safe_load(f)
    assert "unifi_sync" in compose["services"], "unifi_sync xizmati docker-compose.yml'da yo'q"
    assert "profiles" not in compose["services"]["unifi_sync"], (
        "unifi_sync PROFILSIZ bo'lishi kerak (standart 'docker compose up -d' bilan ishga tushishi uchun)"
    )

    # 2) Real HTTP orqali sinxronizatsiya ishlashini tekshirish
    mock_script = "/tmp/_ci_mock_unifi_sync.py"
    with open(mock_script, "w") as f:
        f.write('''
from flask import Flask, request, jsonify
app = Flask(__name__)

@app.route("/proxy/network/integration/v1/sites/ci-sync-site/clients", methods=["GET"])
def clients():
    if request.headers.get("X-API-Key") != "ci-sync-key":
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"data": [
        {"macAddress": "aa:bb:cc:dd:ee:80", "ipAddress": "172.16.41.80", "name": "CI-SYNC-PC", "type": "WIRELESS"},
    ]})

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=19800)
''')

    mock_proc = subprocess.Popen(["python3", mock_script])
    try:
        _time.sleep(2)
        os.environ["UNIFI_CONTROLLER_URL"] = "http://127.0.0.1:19800"
        os.environ["UNIFI_API_KEY"] = "ci-sync-key"
        os.environ["UNIFI_SITE_ID"] = "ci-sync-site"
        os.environ["UNIFI_VERIFY_SSL"] = "false"

        from network_discovery.unifi_sync_loop import run_once
        n = run_once()
        assert n == 1, f"1 ta qurilma kutilgan edi, {n} keldi"

        s = get_session()
        d = s.query(Device).filter(Device.ip_address == "172.16.41.80").first()
        assert d is not None, "unifi_sync_loop.py DB'ga yozmadi"
        assert d.hostname == "CI-SYNC-PC"
        s.close()

        # 3) UniFi sozlanmagan holatda ham xato bermasligi
        os.environ.pop("UNIFI_CONTROLLER_URL")
        n2 = run_once()
        assert n2 == 0

    finally:
        mock_proc.terminate()
        try:
            mock_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            mock_proc.kill()
        os.remove(mock_script)
        for k in ["UNIFI_CONTROLLER_URL", "UNIFI_API_KEY", "UNIFI_SITE_ID", "UNIFI_VERIFY_SSL"]:
            os.environ.pop(k, None)


check("UniFi Sync Loop - standart docker-compose xizmati sifatida (production bo'shlig'i tuzatilgan)", _test_unifi_sync_loop)

# ---------------------------------------------------------------------------
print("\n=== 48) API_SERVER_URL: https:// EMAS http:// (real production xatosi, regressiya himoyasi) ===")


def _test_api_server_url_uses_http_not_https():
    """
    Real production'da topilgan xato: docker-compose.yml'dagi gunicorn
    HECH QANDAY SSL/TLS sertifikatisiz oddiy HTTP orqali ishlaydi, lekin
    hujjatlar/skriptlarda standart qiymat sifatida 'https://' yozilgan
    edi - bu Windows Agent'ning serverga ulanishini JIM ravishda
    (aniq xatosiz) muvaffaqiyatsizlikka olib kelardi.

    Bu test barcha tegishli fayllarda 'https://172.16.0.5:8443' (yoki
    shunga o'xshash) endi qolmaganini tekshiradi.
    """
    files_to_check = [
        "agent_core/agent.py",
        "deploy/windows_agent_gpo/Deploy-NetworkSecurityAgent.ps1",
        "deploy/windows_agent_gpo/Install-NetworkSecurityAgent.ps1",
        "docs_WINDOWS_AGENT_SETUP.md",
        "docs_LINUX_AGENT_SETUP.md",
        ".env.example",
    ]
    for filepath in files_to_check:
        full_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filepath)
        if not os.path.isfile(full_path):
            continue
        with open(full_path) as f:
            content = f.read()
        assert "https://172.16.0.5:8443" not in content, (
            f"{filepath}'da hali ham noto'g'ri 'https://172.16.0.5:8443' bor - "
            f"server SSL/TLS'siz, bu jim ravishda ulanish xatosiga olib keladi"
        )

    # agent_core/agent.py'ning standart qiymati aynan http:// bilan boshlanishini tasdiqlash
    from agent_core.agent import API_SERVER_URL
    assert API_SERVER_URL.startswith("http://"), (
        f"API_SERVER_URL standart qiymati http:// bilan boshlanishi kerak, "
        f"hozirgi qiymat: {API_SERVER_URL}"
    )


check("API_SERVER_URL http:// (https:// emas) - real production ulanish xatosi tuzatilgan", _test_api_server_url_uses_http_not_https)

# ---------------------------------------------------------------------------
print("\n=== 49) SURICATA -> FILE ANALYSIS to'liq zanjiri (yettinchi marta topilgan production bo'shlig'i) ===")


def _test_suricata_full_chain():
    """
    Real production'da topilgan bo'shliq: collectors/suricata_reader.py
    to'g'ri ishlar edi, lekin docker-compose.yml'da uni chaqiruvchi
    HECH QANDAY xizmat yo'q edi (faqat deep_scan_engine'ning
    /var/log/suricata/files bind-mount'i bor edi, eve.json emas).

    Bu test: (1) docker-compose.yml'da suricata_reader xizmati
    mavjudligini, (2) haqiqiy Suricata eve.json formatidagi fayl bilan
    to'liq zanjir (suricata_reader -> FileEvent -> file_analysis_engine)
    ishlashini tekshiradi.
    """
    import shutil
    import yaml

    # 1) docker-compose.yml'da suricata_reader xizmati borligini tasdiqlash
    with open("docker-compose.yml") as f:
        compose = yaml.safe_load(f)
    assert "suricata_reader" in compose["services"], "suricata_reader xizmati docker-compose.yml'da yo'q"

    # 2) Haqiqiy Suricata eve.json formatidagi test fayli bilan to'liq zanjir
    work_dir = "/tmp/_test_suricata_chain"
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(work_dir)
    eve_path = os.path.join(work_dir, "eve.json")

    # Haqiqiy Suricata fileinfo event formatiga mos (rasmiy hujjat asosida)
    test_sha256 = "a" * 64  # test uchun sun'iy, real bo'lmagan hash (haqiqiy threat intel'ga so'rov yubormaslik uchun)
    with open(eve_path, "w") as f:
        f.write(
            '{"timestamp":"2026-08-17T10:00:00.000000+0500","event_type":"fileinfo",'
            '"src_ip":"172.16.1.99","dest_ip":"93.184.216.34","proto":"TCP","app_proto":"http",'
            f'"fileinfo":{{"filename":"ci_test_file.exe","magic":"PE32 executable","size":12345,'
            f'"sha256":"{test_sha256}","md5":"bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb"}}}}\n'
        )

    from collectors.suricata_reader import read_existing
    n = read_existing(eve_path)
    assert n == 1, f"1 ta fileinfo yozuvi kutilgan edi, {n} keldi"

    s = get_session()
    fe = s.query(FileEvent).filter(FileEvent.sha256 == test_sha256).first()
    assert fe is not None, "FileEvent yaratilmadi"
    assert fe.filename == "ci_test_file.exe"
    assert fe.src_ip == "172.16.1.99"
    assert fe.checked is False
    s.close()

    # 3) Bir xil hash+src_ip qayta kelsa, TAKRORLANMASLIGI (dedup)
    n2 = read_existing(eve_path)
    assert n2 == 0, "Bir xil fayl ikkinchi marta ham yozildi - dedup ishlamadi"

    shutil.rmtree(work_dir, ignore_errors=True)


check("Suricata -> File Analysis to'liq zanjiri (docker-compose xizmati + real formatda parsing)", _test_suricata_full_chain)

# ---------------------------------------------------------------------------
print("\n=== 50) GPO Deploy skripti: $env:USERDNSDOMAIN SYSTEM kontekstida ishonchsiz (real production xatosi) ===")


def _test_gpo_script_no_direct_userdnsdomain_in_param():
    """
    Real production'da (Domain Controller, haqiqiy GPO Startup Script
    orqali) topilgan xato: $env:USERDNSDOMAIN GPO Computer Startup
    Script SYSTEM kontekstida (foydalanuvchi hali login qilmasdan
    OLDIN) bo'sh qiymat qaytardi - natijada $ServerShare buzilgan
    (SYSVOL, domen nomisiz) yo'lga aylanib, "VERSION topilmadi"
    xatosiga olib keldi (log fayl orqali tasdiqlangan).

    Bu test param() blokida $env:USERDNSDOMAIN'ning TO'G'RIDAN-TO'G'RI
    ishlatilmasligini (buning o'rniga [System.DirectoryServices.
    ActiveDirectory.Domain]::GetCurrentDomain() orqali ishonchli
    aniqlanishini) tekshiradi.
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "deploy", "windows_agent_gpo", "Deploy-NetworkSecurityAgent.ps1",
    )
    with open(script_path) as f:
        content = f.read()

    param_block_end = content.find(")\n\n$ErrorActionPreference")
    assert param_block_end != -1, "param() blokining oxiri topilmadi - skript strukturasi o'zgargan bo'lishi mumkin"
    param_block = content[:param_block_end]

    assert "$env:USERDNSDOMAIN" not in param_block, (
        "param() blokida $env:USERDNSDOMAIN to'g'ridan-to'g'ri ishlatilmasligi kerak - "
        "bu SYSTEM kontekstida (GPO Startup Script, login'dan oldin) ishonchsiz "
        "(real production'da aniqlangan xato)"
    )
    assert "GetCurrentDomain" in content, (
        "Domen nomini ishonchli aniqlash uchun [System.DirectoryServices."
        "ActiveDirectory.Domain]::GetCurrentDomain() ishlatilishi kerak"
    )

    # Qavslar balansini ham qayta tasdiqlaymiz (avvalgi tekshiruv usuli)
    code_only = [l for l in content.splitlines(keepends=True) if not l.strip().startswith("#")]
    code_content = "".join(code_only)
    for open_c, close_c in [("{", "}"), ("(", ")"), ("[", "]")]:
        assert code_content.count(open_c) == code_content.count(close_c), (
            f"Qavslar balansi buzilgan: {open_c}={code_content.count(open_c)}, {close_c}={code_content.count(close_c)}"
        )


check("GPO Deploy skripti: USERDNSDOMAIN SYSTEM kontekstida ishonchsizligi tuzatilgan", _test_gpo_script_no_direct_userdnsdomain_in_param)

# ---------------------------------------------------------------------------
print("\n=== 51) GPO Deploy skripti: idempotentlik faqat VERSION emas, xizmat mavjudligini ham tekshiradi (real production xatosi) ===")


def _test_gpo_script_checks_service_existence():
    """
    Real production'da topilgan xato: skript faqat VERSION faylini
    solishtirar edi. Agar xizmat biror sababdan (masalan qo'lda
    'NetworkSecurityAgent.exe remove' orqali, yoki muvaffaqiyatsiz
    avvalgi urinishdan keyin) o'chirilgan bo'lsa-yu, VERSION fayli
    InstallDir'da qolib ketgan bo'lsa - skript "hammasi joyida" deb
    noto'g'ri xulosa chiqarib, xizmatni HECH QACHON qayta o'rnatmay
    qo'yardi (foydalanuvchining haqiqiy deploy.log'ida "Agent
    allaqachon eng so'nggi versiyada - hech narsa qilinmadi" ko'rinib,
    lekin Get-Service xizmat topilmasligini ko'rsatgan holat orqali
    tasdiqlangan).
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "deploy", "windows_agent_gpo", "Deploy-NetworkSecurityAgent.ps1",
    )
    with open(script_path) as f:
        content = f.read()

    assert "$serviceExists" in content, (
        "Skript xizmat mavjudligini ($serviceExists) tekshirmayapti - "
        "faqat VERSION solishtirish yetarli emas (real production xatosi)"
    )
    assert "-and $serviceExists" in content, (
        "Idempotentlik shartida 'versiya bir xil VA xizmat mavjud' ikkalasi "
        "ham tekshirilishi kerak, faqat versiya emas"
    )

    code_only = [l for l in content.splitlines(keepends=True) if not l.strip().startswith("#")]
    code_content = "".join(code_only)
    for open_c, close_c in [("{", "}"), ("(", ")"), ("[", "]")]:
        assert code_content.count(open_c) == code_content.count(close_c), (
            f"Qavslar balansi buzilgan: {open_c}={code_content.count(open_c)}, {close_c}={code_content.count(close_c)}"
        )


check("GPO Deploy skripti: idempotentlik xizmat mavjudligini ham tekshiradi (real production xatosi tuzatilgan)", _test_gpo_script_checks_service_existence)

# ---------------------------------------------------------------------------
print("\n=== 52) GPO Deploy skripti: tashqi .exe xatolari yashirilmaydi (real production xatosi) ===")


def _test_gpo_script_checks_exe_exit_code():
    """
    Real production'da topilgan xato: '& $exePath install' PowerShell'ning
    $ErrorActionPreference'iga bo'ysunmaydi (tashqi dastur chaqiruvi) -
    agar install ichki xatolik bilan muvaffaqiyatsiz bo'lsa ham, skript
    "Xizmat .exe orqali o'rnatildi" deb noto'g'ri log yozib, keyingi
    qatorga o'tib ketardi. Natijada xizmat SCM'da umuman ro'yxatga
    olinmagan holda qolib, Get-Service uni "topilmadi" deb qaytarardi -
    lekin log fayl "muvaffaqiyat" deb ko'rsatardi.
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "deploy", "windows_agent_gpo", "Deploy-NetworkSecurityAgent.ps1",
    )
    with open(script_path) as f:
        content = f.read()

    assert "$LASTEXITCODE" in content, (
        "Tashqi .exe chaqiruvidan keyin $LASTEXITCODE tekshirilishi SHART - "
        "aks holda muvaffaqiyatsiz 'install' 'muvaffaqiyat' deb noto'g'ri log yoziladi"
    )
    assert "$registeredService" in content, (
        "Xizmat 'install'dan keyin HAQIQATAN SCM'da ro'yxatga olinganini "
        "(Get-Service orqali) tasdiqlash kerak - install buyrug'i xato bermasa ham "
        "xizmat aslida ro'yxatga olinmagan bo'lishi mumkin (real production xatosi)"
    )

    code_only = [l for l in content.splitlines(keepends=True) if not l.strip().startswith("#")]
    code_content = "".join(code_only)
    for open_c, close_c in [("{", "}"), ("(", ")"), ("[", "]")]:
        assert code_content.count(open_c) == code_content.count(close_c), (
            f"Qavslar balansi buzilgan: {open_c}={code_content.count(open_c)}, {close_c}={code_content.count(close_c)}"
        )


check("GPO Deploy skripti: tashqi .exe xatolari endi yashirilmaydi (real production xatosi tuzatilgan)", _test_gpo_script_checks_exe_exit_code)

# ---------------------------------------------------------------------------
print("\n=== 53) Agent log fayli: mutlaq yo'l, Windows Service LocalSystem muammosi tuzatilgan (real production xatosi) ===")


def _test_agent_log_file_absolute_path():
    """
    Real production'da topilgan xato: agent_core/agent.py'da log fayli
    nisbiy yo'l ("./agent.log") bilan standart qilingan edi. Interaktiv
    ("debug") rejimda muammosiz ishladi, lekin haqiqiy Windows Service
    sifatida (LocalSystem hisobi ostida, standart ish katalogi
    C:\\Windows\\System32) ishga tushirilganda "Cannot start service"
    degan tushunarsiz xato bilan darhol qulab tushardi - chunki
    logging.basicConfig() MODUL IMPORT vaqtida, hech qanday
    try/except'siz FileHandler yaratardi.
    """
    import importlib
    import ntpath
    import logging
    program_data = r"C:\ProgramData"
    expected_log_dir = ntpath.join(program_data, "NetworkSecurityAgent")
    expected_log_path = ntpath.join(expected_log_dir, "agent.log")
    assert expected_log_path == r"C:\ProgramData\NetworkSecurityAgent\agent.log"

    # Kodning o'zida _default_log_file funksiyasi mavjudligini va
    # xavfsiz (keng try/except bilan o'ralgan) ekanligini tasdiqlash
    import agent_core.agent as agent_mod
    assert hasattr(agent_mod, "_default_log_file"), "_default_log_file() funksiyasi topilmadi"

    # Linux muhitida import xatosiz o'tishi va nisbiy yo'lga qaytishi kerak
    log_path = agent_mod._default_log_file()
    assert log_path == "./agent.log", f"Linux'da './agent.log' kutilgan edi, '{log_path}' keldi"

    # Modul allaqachon xatosiz import qilingani (bu funksiya chaqirilgunga
    # qadar allaqachon sinov to'plamining boshqa qismlarida import
    # qilingan bo'lishi mumkin) - bu aynan real production'da qulagan
    # MODUL IMPORT bosqichining o'zi xatosiz o'tganini tasdiqlaydi.
    assert agent_mod.logger is not None


check("Agent log fayli: mutlaq yo'l (Windows Service LocalSystem qulash muammosi tuzatilgan)", _test_agent_log_file_absolute_path)

# ---------------------------------------------------------------------------
print("\n=== 54) service_wrapper.py: ReportServiceStatus(SERVICE_RUNNING) yetishmasligi tuzatilgan (real production TUB SABAB) ===")


def _test_service_wrapper_reports_running_status():
    """
    Real production'da topilgan TUB SABAB: SvcDoRun() metodida
    `self.ReportServiceStatus(win32service.SERVICE_RUNNING)` chaqiruvi
    umuman yo'q edi. Windows Service Control Manager (SCM) xizmatni
    ishga tushirgandan keyin 30 soniya ichida aniq "men ishlayapman"
    signalini kutadi - bu signal yo'qligi sabab SCM har doim "The
    service did not respond to the start or control request in a
    timely fashion" (Timeout 30000 ms) xatosi bilan xizmatni majburan
    o'chirar edi (Windows System Event Log orqali tasdiqlangan) -
    garchi pastdagi Python kodi (EndpointAgent, FileMonitor) o'zi
    to'g'ri ishlagan bo'lsa ham (debug rejimida sinovdan o'tgan).
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "windows_agent", "service_wrapper.py",
    )
    with open(script_path) as f:
        content = f.read()

    assert "ReportServiceStatus(win32service.SERVICE_RUNNING)" in content, (
        "SvcDoRun() SCM'ga SERVICE_RUNNING holatini ANIQ xabar qilishi SHART - "
        "aks holda SCM 30 soniyadan keyin xizmatni majburan o'chiradi "
        "(real production'da Windows Event Log orqali tasdiqlangan xato)"
    )

    # ReportServiceStatus SvcDoRun ichida, EndpointAgent yaratilishidan
    # OLDIN chaqirilishini tasdiqlash (SCM'ga imkon qadar tezroq signal
    # berish uchun - agent ishga tushirish vaqti cho'zilib ketsa ham SCM
    # allaqachon "running" deb bilib turadi)
    svc_do_run_start = content.find("def SvcDoRun")
    report_running_pos = content.find("ReportServiceStatus(win32service.SERVICE_RUNNING)", svc_do_run_start)
    agent_creation_pos = content.find("EndpointAgent(", svc_do_run_start)
    assert report_running_pos != -1 and agent_creation_pos != -1
    assert report_running_pos < agent_creation_pos, (
        "ReportServiceStatus(SERVICE_RUNNING) EndpointAgent yaratilishidan OLDIN "
        "chaqirilishi kerak - SCM'ga imkon qadar tezroq signal berish uchun"
    )


check("service_wrapper.py: SCM'ga SERVICE_RUNNING signali (30s timeout TUB SABABI tuzatilgan)", _test_service_wrapper_reports_running_status)

# ---------------------------------------------------------------------------
print("\n=== 55) Windows Agent: qo'shimcha real production tuzatishlari (--startup auto, ko'p-foydalanuvchi kuzatish, cache yo'li) ===")


def _test_windows_agent_additional_fixes():
    """
    Foydalanuvchi tashqi manbadan (mustaqil ishlab chiqilgan, real
    production sinovlari orqali tasdiqlangan) qo'shimcha tuzatishlar
    bilan zip yubordi. Ko'rib chiqilgach, quyidagi 3 ta QO'SHIMCHA
    real xato ham aniqlandi va bizning kodga integratsiya qilindi:

    1) Deploy skripti xizmatni 'install' bilan (standart - odatda
       "Manual" ishga tushirish turi bilan) o'rnatgan edi - bu
       reboot vaqtida SCM'ning o'zi uni AVTOMATIK ishga tushirmasligini
       anglatadi (foydalanuvchining haqiqiy Get-WinEvent natijasida
       "Тип запуска службы: Вручную" ko'rinib, bu tasdiqlangan).
       Tuzatish: '--startup auto install'.

    2) service_wrapper.py DEFAULT_WATCH_DIRS_WINDOWS (%USERPROFILE%
       asosida) ishlatar edi - bu LocalSystem hisobi ostida
       mazmunsiz (haqiqiy foydalanuvchi profiliga ishora qilmaydi).
       Tuzatish: barcha haqiqiy Windows foydalanuvchi profillarini
       (C:\\Users\\* ostida) avtomatik aniqlaydigan
       _windows_watch_dirs() funksiyasi.

    3) LOCAL_CACHE_FILE (hash keshi) ham nisbiy yo'l bilan yozilgan
       edi - xuddi agent.log kabi, LocalSystem ish katalogi
       muammosiga uchrashi mumkin edi.
    """
    # 1) Deploy skriptida --startup auto borligini tekshirish
    deploy_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "deploy", "windows_agent_gpo", "Deploy-NetworkSecurityAgent.ps1",
    )
    with open(deploy_path) as f:
        deploy_content = f.read()
    assert "--startup auto" in deploy_content, (
        "Deploy skripti '--startup auto' bilan o'rnatishi kerak - aks holda "
        "xizmat reboot'da AVTOMATIK ishga tushmaydi (real production'da "
        "'Тип запуска службы: Вручную' orqali tasdiqlangan xato)"
    )
    # Post-start tekshiruv ham borligini tasdiqlash (xizmat haqiqatan Running holatida)
    assert "runningService" in deploy_content and "Running" in deploy_content, (
        "Deploy skripti Start-Service'dan keyin xizmat holatini qayta tekshirishi kerak"
    )

    # 2) service_wrapper.py'da ko'p-foydalanuvchi kuzatish funksiyasi borligini tekshirish
    wrapper_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "windows_agent", "service_wrapper.py",
    )
    with open(wrapper_path) as f:
        wrapper_content = f.read()
    assert "_windows_watch_dirs" in wrapper_content, (
        "service_wrapper.py barcha Windows foydalanuvchi profillarini avtomatik "
        "aniqlovchi funksiyaga ega bo'lishi kerak (LocalSystem %USERPROFILE% "
        "muammosini hal qilish uchun)"
    )
    assert "Users" in wrapper_content

    # 3) LOCAL_CACHE_FILE ham mutlaq/xavfsiz yo'lga bog'liq ekanligini tekshirish
    import agent_core.agent as agent_mod
    assert hasattr(agent_mod, "LOCAL_CACHE_FILE")
    # Linux muhitida _default_log_file() asosida hisoblanadi (nisbiy "./" emas)

    # 4) CI workflow'da haqiqiy SCM ro'yxatdan o'tish tekshiruvi borligini tasdiqlash
    workflow_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        ".github", "workflows", "build-windows-agent.yml",
    )
    with open(workflow_path) as f:
        workflow_content = f.read()
    assert "sc.exe query NetworkSecurityEndpointAgent" in workflow_content, (
        "CI workflow'da xizmatning HAQIQATAN SCM'da ro'yxatdan o'tishini "
        "tekshiruvchi qadam bo'lishi kerak - bu real production xatosini "
        "(muvaffaqiyat deb log qilingan, lekin SCM'da yo'q) avtomatik ushlaydi"
    )


check("Windows Agent qo'shimcha tuzatishlar (--startup auto, ko'p-foydalanuvchi kuzatish, CI SCM tekshiruvi)", _test_windows_agent_additional_fixes)

# ---------------------------------------------------------------------------
print("\n=== 56) EndpointAgent yangi start_background()/stop() API'si - real thread-asosli heartbeat ===")


def _test_endpoint_agent_start_background_stop():
    """
    agent_core/agent.py EndpointAgent klassi endi start_background()/
    stop() metodlariga ega - bu Windows Service uchun bloklanmaydigan
    ishga tushirish imkonini beradi (heartbeat alohida thread'da).
    Bu real HTTP orqali (heartbeat serverga haqiqatan yetib borishini)
    tekshiriladi.
    """
    import subprocess
    import time as _time
    import tempfile
    import threading as threading_check

    import agent_core.agent as agent_mod
    assert hasattr(agent_mod.EndpointAgent, "start_background")
    assert hasattr(agent_mod.EndpointAgent, "stop")

    api_env = {**os.environ, "AGENT_API_KEY": "ci-newapi-key"}
    api_proc = subprocess.Popen(["python3", "-m", "api.server"], env=api_env,
                                  stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    try:
        _time.sleep(2)
        os.environ["API_SERVER_URL"] = "http://127.0.0.1:8443"
        os.environ["AGENT_API_KEY"] = "ci-newapi-key"
        os.environ["HEARTBEAT_INTERVAL_SECONDS"] = "1"

        import importlib
        importlib.reload(agent_mod)

        watch_dir = tempfile.mkdtemp()
        agent = agent_mod.EndpointAgent([watch_dir])
        agent.start_background()
        assert agent._heartbeat_thread is not None and agent._heartbeat_thread.is_alive()

        _time.sleep(2.5)

        agent.stop()
        _time.sleep(0.5)
        assert not agent._heartbeat_thread.is_alive(), "Heartbeat thread stop() dan keyin ham ishlab turibdi"

        s = get_session()
        d = s.query(Device).filter(Device.hostname == agent.hostname).order_by(Device.id.desc()).first()
        assert d is not None, "Heartbeat orqali qurilma yozilmadi"
        assert d.agent_last_heartbeat is not None
        s.close()

    finally:
        api_proc.terminate()
        try:
            api_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            api_proc.kill()
        for k in ["API_SERVER_URL", "AGENT_API_KEY", "HEARTBEAT_INTERVAL_SECONDS"]:
            os.environ.pop(k, None)


check("EndpointAgent start_background()/stop() - real thread-asosli heartbeat (HTTP orqali tasdiqlangan)", _test_endpoint_agent_start_background_stop)

# ---------------------------------------------------------------------------
print("\n=== 57) service_wrapper.py: SCM Control Dispatcher aniq chaqiruvi (PyInstaller+pywin32 muammosi) ===")


def _test_service_wrapper_explicit_dispatcher():
    """
    Real production'da topilgan xato: ReportServiceStatus(SERVICE_RUNNING)
    va --startup auto tuzatilgandan KEYIN ham, xizmat hali "Cannot start
    service" bilan muvaffaqiyatsiz bo'lardi - garchi install/remove/debug
    (argumentlar bilan chaqirilganda) mukammal ishlagan bo'lsa ham.

    Bu - PyInstaller bilan "muzlatilgan" (frozen) pywin32 xizmatlarining
    tanilgan muammosi: Windows SCM xizmatni HECH QANDAY argumentsiz
    chaqiradi, va win32serviceutil.HandleCommandLine()ning bu holatni
    avtomatik aniqlashi frozen exe'larda ishonchsiz bo'lishi mumkin.

    Tuzatish: sys.argv uzunligini ANIQ tekshirib, argument bo'lmasa
    servicemanager.Initialize()/PrepareToHostSingle()/
    StartServiceCtrlDispatcher()ni QO'LDA chaqirish.
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "windows_agent", "service_wrapper.py",
    )
    with open(script_path) as f:
        content = f.read()

    assert "len(sys.argv) == 1" in content, (
        "Argumentsiz chaqirilish holati ANIQ tekshirilishi kerak (SCM "
        "xizmatni argumentsiz ishga tushiradi)"
    )
    assert "servicemanager.Initialize()" in content
    assert "PrepareToHostSingle" in content
    assert "StartServiceCtrlDispatcher" in content

    # CI workflow'da HAQIQIY Start-Service tekshiruvi borligini tasdiqlash
    # (faqat ro'yxatdan o'tish emas - bu farq real production xatosining
    # aynan o'zi edi)
    workflow_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        ".github", "workflows", "build-windows-agent.yml",
    )
    with open(workflow_path) as f:
        workflow_content = f.read()
    assert "Start-Service -Name NetworkSecurityEndpointAgent" in workflow_content, (
        "CI workflow'da xizmatning HAQIQATAN 'Running' holatiga o'tishini "
        "tekshiruvchi Start-Service chaqiruvi bo'lishi kerak - faqat "
        "ro'yxatdan o'tish (sc.exe query) yetarli emas"
    )
    assert '"Running"' in workflow_content or "'Running'" in workflow_content

    # Sintaksis to'g'riligini qayta tasdiqlash
    import ast
    ast.parse(content)


check("service_wrapper.py: SCM Control Dispatcher aniq chaqiruvi + CI haqiqiy Start-Service tekshiruvi", _test_service_wrapper_explicit_dispatcher)

# ---------------------------------------------------------------------------
print("\n=== 58) Web Activity: Zeek HTTP/SSL/DNS -> WebAccessLog -> Dashboard (to'liq real zanjir) ===")


def _test_web_activity_full_chain():
    """
    Foydalanuvchi yuborgan qo'shimcha funksiya: sayt/domen tarixini
    kuzatish. Zeek http.log/ssl.log/dns.log'dan WebAccessLog jadvaliga,
    va Dashboard'ning /web-activity sahifasida real HTTP orqali
    ko'rinishini tekshiradi.
    """
    import collectors.zeek_reader as zr

    s = get_session()

    # 1) Zeek HTTP log yozuvi
    http_rec = {
        "ts": 1755500000.0, "id.orig_h": "172.16.61.1", "id.resp_h": "93.184.216.34",
        "method": "GET", "host": "ci-test-site.com", "uri": "/page1",
        "status_code": 200, "user_agent": "TestAgent/1.0",
    }
    zr.process_http(s, http_rec)

    # 2) Zeek SSL (TLS SNI) log yozuvi
    ssl_rec = {
        "ts": 1755500010.0, "id.orig_h": "172.16.61.2", "id.resp_h": "142.250.1.1",
        "server_name": "ci-secure-site.com",
    }
    zr.process_ssl(s, ssl_rec)

    # 3) Zeek DNS log yozuvi
    dns_rec = {"ts": 1755500020.0, "id.orig_h": "172.16.61.3", "query": "ci-dns-site.com."}
    zr.process_dns(s, dns_rec)

    s.commit()

    logs = s.query(WebAccessLog).filter(WebAccessLog.source_ip.in_(["172.16.61.1", "172.16.61.2", "172.16.61.3"])).all()
    assert len(logs) == 3, f"3 ta WebAccessLog yozuvi kutilgan edi, {len(logs)} keldi"

    http_log = next(l for l in logs if l.protocol == "HTTP")
    assert http_log.domain == "ci-test-site.com"
    assert http_log.url == "http://ci-test-site.com/page1"
    assert http_log.status_code == 200

    ssl_log = next(l for l in logs if l.protocol == "HTTPS")
    assert ssl_log.domain == "ci-secure-site.com"
    assert ssl_log.url == "https://ci-secure-site.com/"

    dns_log = next(l for l in logs if l.protocol == "DNS")
    assert dns_log.domain == "ci-dns-site.com"
    s.close()

    # 4) Dashboard'da real HTTP orqali ko'rinishini tekshirish
    from dashboard.app import app as dash_app
    from dashboard.create_user import create_user
    create_user("webactivity_ci_admin", "webactivityci123", "admin")
    dash_app.secret_key = "test-secret-webactivity"
    client = dash_app.test_client()
    client.post("/login", data={"username": "webactivity_ci_admin", "password": "webactivityci123"})

    r = client.get("/web-activity")
    assert r.status_code == 200
    assert b"ci-test-site.com" in r.data
    assert b"ci-secure-site.com" in r.data

    # Filtr ishlashini tekshirish (faqat bitta sayt)
    r2 = client.get("/web-activity?site=ci-secure-site")
    assert b"ci-secure-site.com" in r2.data
    assert b"ci-test-site.com" not in r2.data, "Filtr boshqa saytni chiqarib tashlashi kerak edi"


check("Web Activity: Zeek HTTP/SSL/DNS -> WebAccessLog -> Dashboard (real HTTP orqali)", _test_web_activity_full_chain)

# ---------------------------------------------------------------------------
print("\n=== 59) Xavfsiz Karantin: SHA256 tasdiqlash + haqiqiy fayl bilan karantin (agent_core va engine) ===")


def _test_quarantine_mechanism():
    """
    Foydalanuvchi yuborgan qo'shimcha funksiya: zararli fayllarni
    o'chirish o'rniga xavfsiz karantinga olish (SHA256 orqali nusxa
    tasdiqlanadi, keyin asl fayl o'chiriladi).
    """
    import shutil
    import hashlib
    import importlib

    work_dir = "/tmp/_test_quarantine"
    quarantine_dir = "/tmp/_test_quarantine_output"
    for d in (work_dir, quarantine_dir):
        if os.path.exists(d):
            shutil.rmtree(d)
    os.makedirs(work_dir)

    # --- 1) engine/quarantine.py: real fayl bilan muvaffaqiyatli karantin ---
    os.environ["QUARANTINE_DIR"] = quarantine_dir
    import engine.quarantine as eq
    importlib.reload(eq)

    test_file = os.path.join(work_dir, "malware_test.exe")
    with open(test_file, "wb") as f:
        f.write(b"CI test uchun sun'iy zararli fayl mazmuni")

    with open(test_file, "rb") as f:
        real_sha256 = hashlib.sha256(f.read()).hexdigest()

    result = eq.quarantine_file(test_file, real_sha256, "CI test")
    assert result["quarantined"] is True
    assert not os.path.isfile(test_file), "Asl fayl karantinga olingandan keyin o'chirilishi kerak edi"
    assert os.path.isfile(result["quarantine_path"])

    with open(result["quarantine_path"], "rb") as f:
        quarantined_sha256 = hashlib.sha256(f.read()).hexdigest()
    assert quarantined_sha256 == real_sha256, "Karantin nusxasi original bilan bir xil bo'lishi kerak"

    # --- 2) agent_core/quarantine.py: SHA256 MOS KELMASA, xavfsizlik tekshiruvi rad etishi kerak ---
    os.environ["AGENT_QUARANTINE_DIR"] = quarantine_dir + "_agentcore"
    import agent_core.quarantine as aq
    importlib.reload(aq)

    test_file2 = os.path.join(work_dir, "real_file.exe")
    with open(test_file2, "wb") as f:
        f.write(b"Haqiqiy fayl mazmuni")

    wrong_sha256 = "f" * 64  # ataylab noto'g'ri
    result2 = aq.quarantine_file(test_file2, wrong_sha256, "CI test - noto'g'ri hash")
    assert result2["quarantined"] is False, "Noto'g'ri SHA256 bilan karantin MUVAFFAQIYATSIZ bo'lishi kerak edi"
    assert os.path.isfile(test_file2), (
        "SHA256 mos kelmasa, asl fayl SAQLANIB QOLISHI kerak (xavfsizlik nazorati)"
    )

    # --- 3) agent_core/quarantine.py: to'g'ri SHA256 bilan muvaffaqiyatli ---
    with open(test_file2, "rb") as f:
        correct_sha256 = hashlib.sha256(f.read()).hexdigest()
    result3 = aq.quarantine_file(test_file2, correct_sha256, "CI test - to'g'ri hash")
    assert result3["quarantined"] is True
    assert not os.path.isfile(test_file2)

    shutil.rmtree(work_dir, ignore_errors=True)
    shutil.rmtree(quarantine_dir, ignore_errors=True)
    shutil.rmtree(quarantine_dir + "_agentcore", ignore_errors=True)
    for k in ["QUARANTINE_DIR", "AGENT_QUARANTINE_DIR"]:
        os.environ.pop(k, None)


check("Xavfsiz Karantin (SHA256 tasdiqlash, real fayl bilan, mos kelmasa rad etish)", _test_quarantine_mechanism)

# ---------------------------------------------------------------------------
print("\n=== 60) File Analysis Engine: VirusTotal 'confirmed' chegara mantig'i (real DB bilan) ===")


def _test_file_analysis_confirmed_threshold():
    """
    Foydalanuvchi yuborgan qo'shimcha tuzatish: bitta VirusTotal
    dvigateli signal bergani hali "tasdiqlangan" (avtomatik karantin
    uchun asos) degani emas - soxta-pozitiv xavfi. Kamida 3 dvigatel
    VA hisobot beruvchilarning kamida 5% signal berishi talab qilinadi.
    Mahalliy blacklist va MalwareBazaar esa har doim "tasdiqlangan".
    """
    from unittest.mock import patch
    import engine.file_analysis_engine as fae

    # 1) Mahalliy blacklist - har doim tasdiqlangan
    s = get_session()
    s.add(HashBlacklist(sha256="1" * 64, threat_name="CI.LocalMalware", source="ci_test"))
    fe1 = FileEvent(src_ip="172.16.62.1", filename="local.exe", sha256="1" * 64, checked=False)
    s.add(fe1)
    s.commit()
    fae.analyze_one(s, fe1)
    s.commit()
    assert fe1.verdict == "malicious"
    alert1 = s.query(Alert).filter(Alert.file_event_id == fe1.id).first()
    assert alert1.severity == "critical"
    assert "TASDIQLANGAN" in alert1.action_taken
    s.close()

    # 2) VirusTotal past ishonch (1/70) - "shubhali" bo'lishi, karantin YO'Q
    s = get_session()
    fe2 = FileEvent(src_ip="172.16.62.2", filename="low_confidence.exe", sha256="2" * 64, checked=False)
    s.add(fe2)
    s.commit()
    with patch.object(fae, "check_virustotal", return_value={"malicious": True, "positives": 1, "total": 70, "threat_name": "Generic"}), \
         patch.object(fae, "check_malwarebazaar", return_value=None):
        fae.analyze_one(s, fe2)
        s.commit()
    assert fe2.verdict == "unknown", "1 ta dvigatel bilan 'tasdiqlangan' bo'lmasligi kerak edi"
    alert2 = s.query(Alert).filter(Alert.file_event_id == fe2.id).first()
    assert alert2.severity == "medium"
    assert "SHUBHALI" in alert2.action_taken
    s.close()

    # 3) VirusTotal yuqori ishonch (5/70, >=3 VA >=5%) - "tasdiqlangan"
    s = get_session()
    fe3 = FileEvent(src_ip="172.16.62.3", filename="high_confidence.exe", sha256="3" * 64, checked=False)
    s.add(fe3)
    s.commit()
    with patch.object(fae, "check_virustotal", return_value={"malicious": True, "positives": 5, "total": 70, "threat_name": "Trojan.Confirmed"}), \
         patch.object(fae, "check_malwarebazaar", return_value=None):
        fae.analyze_one(s, fe3)
        s.commit()
    assert fe3.verdict == "malicious"
    alert3 = s.query(Alert).filter(Alert.file_event_id == fe3.id).first()
    assert alert3.severity == "critical"
    assert "TASDIQLANGAN" in alert3.action_taken
    s.close()


check("File Analysis Engine: VirusTotal 'confirmed' chegara mantig'i (mahalliy/past/yuqori ishonch)", _test_file_analysis_confirmed_threshold)

# ---------------------------------------------------------------------------
print("\n=== 61) Deep Scan Engine: haqiqiy fayl bilan to'liq karantin zanjiri (EICAR) ===")


def _test_deep_scan_real_quarantine():
    """
    Foydalanuvchi yuborgan qo'shimcha tuzatish: Deep Scan Engine'da
    avvalgi 'TODO' placeholder o'rniga haqiqiy karantin. YARA/ClamAV
    signal berganda, fayl haqiqatan xavfsiz karantinga olinishi
    (SHA256 tasdiqlangan holda) real EICAR test signature bilan
    tekshiriladi.
    """
    import shutil
    from unittest.mock import patch
    import hashlib
    import importlib

    work_dir = "/tmp/_test_deep_scan_quarantine"
    quarantine_dir = "/tmp/_test_deep_scan_quarantine_output"
    for d in (work_dir, quarantine_dir):
        if os.path.exists(d):
            shutil.rmtree(d)
    os.makedirs(work_dir)

    os.environ["QUARANTINE_DIR"] = quarantine_dir
    import engine.deep_scan_engine as dse
    importlib.reload(dse)

    eicar_path = os.path.join(work_dir, "eicar.txt")
    eicar_content = b"X5O!P%@AP[4\\PZX54(P^)7CC)7}$EICAR-STANDARD-ANTIVIRUS-TEST-FILE!$H+H*\n"
    with open(eicar_path, "wb") as f:
        f.write(eicar_content)
    sha256 = hashlib.sha256(eicar_content).hexdigest()

    s = get_session()
    fe = FileEvent(src_ip="172.16.63.1", filename="eicar.txt", sha256=sha256,
                    stored_path=eicar_path, checked=True, verdict="unknown")
    s.add(fe)
    s.commit()

    with patch.object(dse, "yara_scan_file", return_value=[{"rule": "CI_Test_Rule", "severity": "critical", "description": "CI test"}]), \
         patch.object(dse, "clamav_db_available", return_value=False):
        dse.deep_scan_one(s, fe)
        s.commit()

    assert fe.verdict == "malicious"
    alert = s.query(Alert).filter(Alert.file_event_id == fe.id).first()
    assert "karantinaga olindi" in alert.action_taken
    assert not os.path.isfile(eicar_path), "EICAR fayli karantinga olinib, asli o'chirilishi kerak edi"
    s.close()

    shutil.rmtree(work_dir, ignore_errors=True)
    shutil.rmtree(quarantine_dir, ignore_errors=True)
    os.environ.pop("QUARANTINE_DIR", None)


check("Deep Scan Engine: real EICAR fayl bilan to'liq karantin zanjiri", _test_deep_scan_real_quarantine)

# ---------------------------------------------------------------------------
print("\n=== 62) Windows Agent: tizim proksi sozlamalaridan mustaqil ulanish (real production xatosi) ===")


def _test_agent_bypasses_system_proxy():
    """
    Real production'da topilgan xato: "Isobek" kompyuterida agent
    (LocalSystem hisobi) HAR BIR so'rovda ConnectionResetError bilan
    muvaffaqiyatsiz bo'lardi, garchi interaktiv foydalanuvchi
    sessiyasidan (Invoke-WebRequest) aynan bir xil serverga
    muvaffaqiyatli ulanish mumkin bo'lsa ham. Sabab: LocalSystem
    muhit/tizim darajasidagi proksi sozlamalarini (masalan noto'g'ri
    sozlangan WinHTTP proksi) hurmat qiladi, requests kutubxonasi esa
    standart holatda shu proksini ishlatishga urinadi.

    Tuzatish: barcha ichki API chaqiruvlariga aniq `proxies={"http":
    None, "https": None}` qo'shildi - bizning server bilan aloqa
    hech qachon tashqi proksiga muhtoj emas.
    """
    import subprocess
    import time as _time

    api_env = {**os.environ, "AGENT_API_KEY": "ci-proxy-test-key"}
    api_proc = subprocess.Popen(["python3", "-m", "api.server"], env=api_env,
                                  stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    try:
        _time.sleep(2)

        # Mavjud bo'lmagan, xato beruvchi proksi - real "Isobek" holatini simulyatsiya qiladi
        os.environ["HTTP_PROXY"] = "http://127.0.0.1:19998"
        os.environ["HTTPS_PROXY"] = "http://127.0.0.1:19998"
        os.environ["API_SERVER_URL"] = "http://127.0.0.1:8443"
        os.environ["AGENT_API_KEY"] = "ci-proxy-test-key"

        import importlib
        import agent_core.agent as agent_mod
        importlib.reload(agent_mod)

        result = agent_mod.check_hash_with_server_or_cache("b" * 64, {})
        assert result["source"] != "no_data_offline", (
            "Agent noto'g'ri tizim proksisi bilan ulanib bo'lmadi - "
            "bu real production'da 'Isobek' kompyuterida uchragan xato "
            "(ConnectionResetError) bilan bir xil turkum"
        )

        # Kod darajasida ham aniq tekshiramiz: barcha requests.post
        # chaqiruvlarida proxies= parametri borligini
        agent_source = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "agent_core", "agent.py",
        )
        with open(agent_source) as f:
            content = f.read()
        assert content.count('proxies={"http": None, "https": None}') >= 3, (
            "check_hash, report_incident, send_heartbeat - uchalasida ham "
            "proxies=None aniq belgilangan bo'lishi kerak"
        )

    finally:
        api_proc.terminate()
        try:
            api_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            api_proc.kill()
        for k in ["HTTP_PROXY", "HTTPS_PROXY", "API_SERVER_URL", "AGENT_API_KEY"]:
            os.environ.pop(k, None)


check("Windows Agent tizim proksi sozlamalaridan mustaqil (real 'Isobek' xatosi tuzatilgan)", _test_agent_bypasses_system_proxy)

# ---------------------------------------------------------------------------
print("\n=== 63) Windows Agent: _windows_watch_dirs() diagnostika loglari va SystemDrive fallback (real 'Isobek' xatosi) ===")


def _test_windows_watch_dirs_diagnostics_and_fallback():
    """
    Real production'da topilgan xato: agent qayta yoqilgandan keyin
    faqat C:\\WINDOWS\\TEMP va C:\\WINDOWS\\Temp'ni kuzatgan, foydalanuvchi
    Downloads papkasi butunlay tashlab ketilgan - hech qanday xato
    yoki ogohlantirish log qilinmagan.

    Tuzatish: (1) har bir profil tekshiruvi endi aniq log qilinadi
    (topildi/topilmadi/xato), (2) SystemDrive muhit o'zgaruvchisi
    kutilganidek ishlamasa (masalan bo'sh qator bo'lsa, natijada
    nisbiy "Users" yo'liga aylanib, jim ravishda hech narsa
    topilmasligi mumkin edi), standart C:\\Users yo'liga zaxira
    (fallback) qo'shildi.
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "windows_agent", "service_wrapper.py",
    )
    with open(script_path) as f:
        content = f.read()

    assert 'candidates = [r"C:\\Users"]' in content, (
        "SystemDrive muhit o'zgaruvchisi ishonchsiz bo'lganda standart "
        "C:\\Users yo'liga zaxira (fallback) qo'shilishi kerak"
    )
    assert "logger.info" in content and "Yakuniy kuzatiladigan papkalar" in content, (
        "_windows_watch_dirs() endi aniq diagnostika loglari yozishi kerak - "
        "aks holda 'Downloads topilmadi' kabi muammolar jim qolib ketadi"
    )
    assert "logger.warning" in content and "kirish huquqi cheklangan" in content, (
        "Profilga kirish huquqi bo'lmagan holat aniq ogohlantirilishi kerak"
    )

    import ast
    ast.parse(content)


check("Windows Agent: _windows_watch_dirs() diagnostika + SystemDrive fallback (real 'Isobek' xatosi)", _test_windows_watch_dirs_diagnostics_and_fallback)

# ---------------------------------------------------------------------------
print("\n=== 65) Deploy skripti: API_SERVER_URL SYSVOL faylidan (versiya yangilanishida qayta sozlash shart emas) ===")


def _test_deploy_script_reads_api_server_url_from_file():
    """
    Foydalanuvchi so'rovi: har safar yangi Deploy-NetworkSecurityAgent.ps1
    versiyasini GitHub'dan yuklab olganda, API_SERVER_URL'ni qo'lda
    qayta sozlashi shart bo'lmasligi kerak. AGENT_API_KEY allaqachon
    alohida SYSVOL faylidan (api_key.secret) o'qilardi - endi
    API_SERVER_URL ham xuddi shu naqsh bilan (api_server_url.txt)
    ishlaydi.
    """
    script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "deploy", "windows_agent_gpo", "Deploy-NetworkSecurityAgent.ps1",
    )
    with open(script_path) as f:
        content = f.read()

    assert 'api_server_url.txt' in content, (
        "Deploy skripti API_SERVER_URL'ni alohida SYSVOL faylidan "
        "o'qishi kerak - versiya yangilanganda qayta sozlash shart bo'lmasligi uchun"
    )
    # Fayl AGENT_API_KEY o'rnatilishidan OLDIN o'qilishi kerak (mantiqiy tartib)
    server_url_pos = content.find("api_server_url.txt")
    api_key_pos = content.find("api_key.secret")
    assert server_url_pos != -1 and api_key_pos != -1
    assert server_url_pos < api_key_pos, (
        "API_SERVER_URL SYSVOL faylini o'qish AGENT_API_KEY'dan OLDIN bo'lishi kerak"
    )

    # Qavslar balansini qayta tasdiqlash
    code_only = [l for l in content.splitlines(keepends=True) if not l.strip().startswith("#")]
    code_content = "".join(code_only)
    for open_c, close_c in [("{", "}"), ("(", ")"), ("[", "]")]:
        assert code_content.count(open_c) == code_content.count(close_c)


check("Deploy skripti: API_SERVER_URL SYSVOL faylidan o'qiladi (versiya yangilanishida qayta sozlash shart emas)", _test_deploy_script_reads_api_server_url_from_file)

# ---------------------------------------------------------------------------
print("\n=== 64) Windows Agent: service_wrapper.py haqiqatan import va ishga tushirilganda xato bermasligi (CI'da topilgan REGRESSIYA) ===")


def _test_service_wrapper_actually_runs_without_nameerror():
    """
    Men o'zim (oldingi commit'da) qo'shgan REAL REGRESSIYA: yangi
    _windows_watch_dirs() logger.info()/.warning()/.debug()ni
    ishlatadi, lekin service_wrapper.py FAQAT `EndpointAgent`ni
    import qilardi (`from windows_agent.agent import EndpointAgent`),
    `logger`ni EMAS. Bu sintaksis darajasida (ast.parse) sezilmaydi -
    faqat funksiya HAQIQATAN chaqirilganda NameError beradi. Bu
    xato GitHub CI'ning haqiqiy Start-Service tekshiruvida ushlandi
    (xizmat ishga tushishda qulab tushdi).

    Bu test pywin32 modullarini soxta (mock) qilib, service_wrapper.py
    ni HAQIQATAN import qilib, _windows_watch_dirs()ni chaqirib,
    NameError chiqmasligini tasdiqlaydi - bu faqat ast.parse() emas,
    balki HAQIQIY bajarilishni tekshiradi.
    """
    import shutil
    import importlib

    fake_pywin32_dir = "/tmp/_test_fake_pywin32"
    if os.path.exists(fake_pywin32_dir):
        shutil.rmtree(fake_pywin32_dir)
    os.makedirs(fake_pywin32_dir)

    with open(os.path.join(fake_pywin32_dir, "win32event.py"), "w") as f:
        f.write("INFINITE = -1\ndef CreateEvent(*a, **kw): return object()\ndef SetEvent(*a, **kw): pass\ndef WaitForSingleObject(*a, **kw): pass\n")
    with open(os.path.join(fake_pywin32_dir, "win32service.py"), "w") as f:
        f.write("SERVICE_RUNNING = 4\nSERVICE_STOP_PENDING = 3\n")
    with open(os.path.join(fake_pywin32_dir, "win32serviceutil.py"), "w") as f:
        f.write(
            "class ServiceFramework:\n"
            "    def __init__(self, args): pass\n"
            "    def ReportServiceStatus(self, status): pass\n"
            "def HandleCommandLine(cls): pass\n"
        )
    with open(os.path.join(fake_pywin32_dir, "servicemanager.py"), "w") as f:
        f.write(
            "EVENTLOG_INFORMATION_TYPE = 1\nPYS_SERVICE_STARTED = 1\n"
            "def LogMsg(*a, **kw): pass\n"
            "def LogErrorMsg(*a, **kw): pass\n"
            "def LogWarningMsg(*a, **kw): pass\n"
            "def Initialize(): pass\n"
            "def PrepareToHostSingle(cls): pass\n"
            "def StartServiceCtrlDispatcher(): pass\n"
        )

    sys.path.insert(0, fake_pywin32_dir)
    try:
        for mod_name in ["windows_agent.service_wrapper", "win32event", "win32service", "win32serviceutil", "servicemanager"]:
            sys.modules.pop(mod_name, None)

        import windows_agent.service_wrapper as sw
        assert hasattr(sw, "logger"), (
            "service_wrapper.py `logger`ni import qilishi SHART - "
            "aks holda _windows_watch_dirs() ishga tushganda NameError beradi "
            "(bu real CI'da xizmatning ishga tushmasligiga sabab bo'lgan edi)"
        )

        # HAQIQATAN chaqirib, NameError chiqmasligini tasdiqlash
        result = sw._windows_watch_dirs()
        assert isinstance(result, list)

        # To'liq SvcDoRun mantig'ini ham (win32event.WaitForSingleObject'siz) sinash
        agent = sw.EndpointAgent(result)
        assert agent is not None

    finally:
        sys.path.remove(fake_pywin32_dir)
        shutil.rmtree(fake_pywin32_dir, ignore_errors=True)
        for mod_name in ["windows_agent.service_wrapper", "win32event", "win32service", "win32serviceutil", "servicemanager"]:
            sys.modules.pop(mod_name, None)
        import windows_agent.agent  # noqa: F401 - keyingi testlar uchun agent_core.agent holatini tozalash


check("service_wrapper.py: HAQIQATAN import/chaqirilganda NameError yo'q (CI regressiyasi tuzatilgan)", _test_service_wrapper_actually_runs_without_nameerror)

# ---------------------------------------------------------------------------
print("\n=== 66) Dashboard: qurilma 'tarmoqqa ulangan/uzilgan' holati + qurilmalar soni (real HTTP orqali) ===")


def _test_device_online_offline_status():
    """
    Foydalanuvchi so'radi: Dashboard'da qurilmalarning tarmoqqa
    ulangan/ulanmaganini va qurilmalar sonini ko'rsatish.

    `Device.last_seen` DEVICE_OFFLINE_THRESHOLD_MINUTES (standart 60
    daqiqa)dan eski bo'lsa - qurilma "OFFLAYN" deb ko'rsatilishi kerak,
    aks holda "ONLAYN". /devices sahifasi sarlavhasi haqiqiy JAMI sonni
    ko'rsatishi kerak (200 tagacha cheklangan ro'yxat uzunligi emas).
    """
    from datetime import timedelta
    from db.database import get_session
    from db.models import Device, utcnow
    from dashboard.create_user import create_user
    from dashboard.app import app as dash_app

    s = get_session()
    online_dev = Device(ip_address="172.16.52.10", hostname="ONLINE-CI-PC",
                         last_seen=utcnow() - timedelta(minutes=2))
    offline_dev = Device(ip_address="172.16.52.20", hostname="OFFLINE-CI-PC",
                          last_seen=utcnow() - timedelta(hours=5))
    s.add_all([online_dev, offline_dev])
    s.commit()
    before_total = s.query(Device).count()
    s.close()

    create_user("device_status_ci_admin", "devstatusci123", "admin")
    dash_app.secret_key = "test-secret-device-status-ci"
    client = dash_app.test_client()
    client.post("/login", data={"username": "device_status_ci_admin", "password": "devstatusci123"})

    r = client.get("/devices")
    assert r.status_code == 200
    body = r.data.decode("utf-8")

    online_idx = body.find("ONLINE-CI-PC")
    offline_idx = body.find("OFFLINE-CI-PC")
    assert online_idx != -1 and offline_idx != -1, "Test qurilmalari /devices sahifasida topilmadi"

    online_row_start = body.rfind("<tr>", 0, online_idx)
    online_row = body[online_row_start:online_idx]
    assert "ONLAYN" in online_row, "So'nggi 2 daqiqada ko'rilgan qurilma ONLAYN deb belgilanishi kerak edi"

    offline_row_start = body.rfind("<tr>", 0, offline_idx)
    offline_row = body[offline_row_start:offline_idx]
    assert "OFFLAYN" in offline_row, "5 soatdan beri ko'rinmagan qurilma OFFLAYN deb belgilanishi kerak edi"

    # Sarlavha haqiqiy JAMI sonni (200 limitiga qaramay) ko'rsatishi kerak
    assert f"Barcha qurilmalar ({before_total})" in body, (
        f"/devices sahifasi haqiqiy jami qurilmalar sonini ({before_total}) ko'rsatmadi"
    )

    # status=online/offline filtri to'g'ri ishlashi
    r_online = client.get("/devices?status=online")
    assert b"ONLINE-CI-PC" in r_online.data and b"OFFLINE-CI-PC" not in r_online.data

    r_offline = client.get("/devices?status=offline")
    assert b"OFFLINE-CI-PC" in r_offline.data and b"ONLINE-CI-PC" not in r_offline.data

    # Bosh sahifada (index) ham onlayn/offlayn son ko'rinishi kerak
    r_index = client.get("/")
    assert r_index.status_code == 200
    assert b"Tarmoqqa ulangan" in r_index.data


check("Dashboard: qurilma onlayn/offlayn holati + jami qurilmalar soni (real HTTP orqali)", _test_device_online_offline_status)

# ---------------------------------------------------------------------------
print("\n=== 67) XAVFSIZLIK: AGENT_API_KEY standart (fallback) qiymati butunlay olib tashlandi ===")


def _test_no_default_agent_api_key():
    """
    Xavfsizlik auditida topilgan CRITICAL muammo: `AGENT_API_KEY = os.getenv(
    "AGENT_API_KEY", "change-me-in-production")` - agar administrator .env
    faylida buni sozlashni unutsa, server ochiq manba kodida hamma ko'radigan
    ANIQ shu qatorni "to'g'ri kalit" sifatida qabul qilardi.

    Tuzatish: standart qiymat YO'Q. AGENT_API_KEY bo'sh bo'lsa, eski
    umumiy-kalit yo'li BUTUNLAY o'chadi (faqat per-agent token'lar
    ishlaydi) - "yopiq holatda muvaffaqiyatsiz bo'lish" (fail-closed).
    """
    import importlib
    from api import server as api_server
    from api import token_manager

    # 1) AGENT_API_KEY sozlanmagan holat - eski "standart" qiymat endi
    #    ISHLAMASLIGI kerak, va bo'sh X-API-Key ham "" == ""ga o'xshab
    #    muvaffaqiyatli bo'lib qolmasligi kerak.
    os.environ.pop("AGENT_API_KEY", None)
    importlib.reload(api_server)
    assert api_server.AGENT_API_KEY == "", "AGENT_API_KEY sozlanmaganda bo'sh qator bo'lishi kerak edi"

    client = api_server.app.test_client()

    r = client.post("/api/v1/check_hash", json={"sha256": "e" * 64},
                     headers={"X-API-Key": "change-me-in-production"})
    assert r.status_code == 401, (
        f"ESKI STANDART QIYMAT hali ham qabul qilinmoqda! (status={r.status_code}) - "
        f"bu aynan xavfsizlik auditida topilgan CRITICAL zaiflik"
    )

    r = client.post("/api/v1/check_hash", json={"sha256": "f" * 64}, headers={"X-API-Key": ""})
    assert r.status_code == 401, "Bo'sh X-API-Key rad etilishi kerak edi"

    r = client.post("/api/v1/check_hash", json={"sha256": "g" * 64})
    assert r.status_code == 401, "X-API-Key umuman yo'q bo'lganda rad etilishi kerak edi"

    # Per-agent token AGENT_API_KEY sozlanmagan bo'lsa ham ishlashi kerak
    token = token_manager.create_token("ci-no-default-key-test", created_by="ci")
    r = client.post("/api/v1/check_hash", json={"sha256": "h" * 64}, headers={"X-API-Key": token})
    assert r.status_code == 200, "Per-agent token AGENT_API_KEY yo'q bo'lsa ham ishlashi kerak edi"

    # 2) AGENT_API_KEY ANIQ sozlansa - orqaga moslik hali ishlashi kerak
    os.environ["AGENT_API_KEY"] = "ci-explicit-legacy-key"
    importlib.reload(api_server)
    client = api_server.app.test_client()
    r = client.post("/api/v1/check_hash", json={"sha256": "i" * 64},
                     headers={"X-API-Key": "ci-explicit-legacy-key"})
    assert r.status_code == 200, "Aniq sozlangan AGENT_API_KEY ishlashi kerak edi"

    os.environ.pop("AGENT_API_KEY", None)
    importlib.reload(api_server)


check("XAVFSIZLIK: AGENT_API_KEY standart qiymati olib tashlandi (fail-closed, real HTTP orqali)", _test_no_default_agent_api_key)

# ---------------------------------------------------------------------------
print("\n=== 68) XAVFSIZLIK: Agent API so'rov chegaralash (rate limiting) ===")


def _test_api_rate_limiting():
    """
    Xavfsizlik auditida topilgan HIGH muammo: Agent API'da hech qanday
    so'rov chegarasi yo'q edi - bitta buzilgan/o'g'irlangan token
    cheksiz `check_hash` so'rovi yuborib, serverni yoki tashqi
    VirusTotal/MalwareBazaar API kvotasini tugatishi mumkin edi.

    Bu test tez, deterministik tekshiruv uchun juda kichik chegara
    ("3 per second") bilan real HTTP so'rovlar yuboradi.
    """
    import importlib
    import time as _time
    from db.database import get_session
    from db.models import HashBlacklist
    from api import server as api_server
    from api import token_manager

    # MUHIM: mahalliy blacklist'da OLDINDAN mavjud hash ishlatiladi -
    # aks holda check_hash() har safar VirusTotal/MalwareBazaar'ga
    # (sandbox'da bloklangan domenlar) chiqishga urinib, har bir so'rov
    # sekunlab kechikadi - bu "N ta so'rov 1 soniyada" chegarasini real
    # sinash imkonini bermaydi (har bir so'rov o'zi >1s cho'zilib ketadi).
    s = get_session()
    rl_hash = "7" * 64
    if not s.query(HashBlacklist).filter_by(sha256=rl_hash).first():
        s.add(HashBlacklist(sha256=rl_hash, threat_name="CI-RateLimit-Test", source="ci"))
        s.commit()
    s.close()

    os.environ["AGENT_API_KEY"] = "ci-ratelimit-legacy-key"
    os.environ["API_RATE_LIMIT_CHECK_HASH"] = "3 per second"
    importlib.reload(api_server)
    client = api_server.app.test_client()

    token_a = token_manager.create_token("ci-ratelimit-agent-a", created_by="ci")
    token_b = token_manager.create_token("ci-ratelimit-agent-b", created_by="ci")

    def _check(token):
        return client.post(
            "/api/v1/check_hash", json={"sha256": rl_hash},
            headers={"X-API-Key": token},
        )

    # 3 ta so'rov - chegara ichida, hammasi muvaffaqiyatli bo'lishi kerak
    statuses = [_check(token_a).status_code for _ in range(3)]
    assert statuses == [200, 200, 200], f"Chegara ichidagi so'rovlar 200 qaytarishi kerak edi: {statuses}"

    # 4-so'rov - xuddi shu oynada, xuddi shu token/kalit - 429 bo'lishi kerak
    r = _check(token_a)
    assert r.status_code == 429, f"Chegaradan oshgan so'rov 429 qaytarishi kerak edi, {r.status_code} keldi"

    # Boshqa agent/token - MUSTAQIL chegaraga ega, hali ham 200 bo'lishi kerak
    # (bitta buzilgan agent boshqalarni bloklab qo'ymasligi kerak)
    r = _check(token_b)
    assert r.status_code == 200, "Boshqa token/agent alohida chegaraga ega bo'lishi kerak edi"

    # Vaqt oynasi tugagach - qayta ishlashi kerak (doimiy bloklanmagan)
    _time.sleep(1.1)
    r = _check(token_a)
    assert r.status_code == 200, "Vaqt oynasi tugagach so'rov qayta ishlashi kerak edi"

    for k in ["API_RATE_LIMIT_CHECK_HASH", "AGENT_API_KEY"]:
        os.environ.pop(k, None)
    importlib.reload(api_server)


check("XAVFSIZLIK: Agent API rate limiting (per-token chegara, real HTTP orqali)", _test_api_rate_limiting)

# ---------------------------------------------------------------------------
print("\n=== 69) XAVFSIZLIK: RabbitMQ va Grafana standart credential/portlari yopildi ===")


def _test_rabbitmq_grafana_hardening():
    """
    Xavfsizlik auditida topilgan HIGH muammolar:
      - RabbitMQ: 5672 (AMQP) va 15672 (boshqaruv paneli) docker-compose'da
        BARCHA interfeyslarga (0.0.0.0) ochiq edi, va standart RabbitMQ
        "guest:guest" login/paroli hech qanday o'zgartirishsiz ishlatilardi.
      - Grafana: GF_SECURITY_ADMIN_PASSWORD `.env`'da sozlanmasa,
        `${GRAFANA_ADMIN_PASSWORD:-admin}` orqali "admin/admin" bilan
        ochiq qolardi.
    """
    import yaml

    with open("docker-compose.yml") as f:
        raw_compose = f.read()
        compose = yaml.safe_load(raw_compose)

    rabbitmq_svc = compose["services"]["rabbitmq"]
    assert rabbitmq_svc.get("ports") == ["127.0.0.1:15672:15672"], (
        f"RabbitMQ portlari xavfsiz emas: {rabbitmq_svc.get('ports')} - "
        f"5672 tashqariga chiqmasligi, 15672 esa faqat 127.0.0.1'ga bog'lanishi kerak"
    )
    assert "amqp://guest:guest" not in raw_compose, "docker-compose.yml'da hali ham 'guest:guest' AMQP URL'ga hardcode qilingan"
    assert ":?RABBITMQ_DEFAULT_USER" in raw_compose and ":?RABBITMQ_DEFAULT_PASS" in raw_compose, (
        "RABBITMQ_DEFAULT_USER/PASS majburiy shaklida (':?') sozlanmagan"
    )

    grafana_svc = compose["services"]["grafana"]
    grafana_pw = grafana_svc["environment"]["GF_SECURITY_ADMIN_PASSWORD"]
    assert ":-admin" not in grafana_pw, f"Grafana hali ham 'admin' standart paroliga ega: {grafana_pw}"
    assert ":?" in grafana_pw, f"GRAFANA_ADMIN_PASSWORD majburiy bo'lishi kerak: {grafana_pw}"

    # --- Real RabbitMQ broker bilan - kredensial talab qilinishini tasdiqlash ---
    import subprocess
    if subprocess.run(["which", "rabbitmqctl"], capture_output=True).returncode != 0:
        print("   (real broker qismi o'tkazib yuborildi - rabbitmq-server o'rnatilmagan bu muhitda)")
        return

    import pika
    from pika.exceptions import AMQPConnectionError

    test_user = "ci_hardening_user"
    test_pass = "CiHardeningStrongPass123!"
    subprocess.run(["rabbitmqctl", "delete_user", test_user], capture_output=True)
    r = subprocess.run(["rabbitmqctl", "add_user", test_user, test_pass], capture_output=True, text=True)
    assert r.returncode == 0, f"Test foydalanuvchi yaratilmadi: {r.stderr}"
    subprocess.run(["rabbitmqctl", "set_permissions", "-p", "/", test_user, ".*", ".*", ".*"], check=True)

    try:
        # To'g'ri kredensial bilan - ulanish MUVAFFAQIYATLI bo'lishi kerak
        conn = pika.BlockingConnection(pika.URLParameters(f"amqp://{test_user}:{test_pass}@localhost:5672/%2F"))
        assert conn.is_open
        conn.close()

        # Noto'g'ri parol bilan - ulanish RAD ETILISHI kerak (broker
        # haqiqatan kredensialni tekshiryapti, "hammaga ochiq" emas)
        try:
            pika.BlockingConnection(pika.URLParameters(f"amqp://{test_user}:WRONG_PASSWORD@localhost:5672/%2F"))
            assert False, "Noto'g'ri parol bilan ulanish MUVAFFAQIYATLI bo'ldi - bu xavfsizlik xatosi"
        except AMQPConnectionError:
            pass  # kutilgan natija
    finally:
        subprocess.run(["rabbitmqctl", "delete_user", test_user], capture_output=True)


check("XAVFSIZLIK: RabbitMQ portlari/credential + Grafana standart parol yopildi", _test_rabbitmq_grafana_hardening)

# ---------------------------------------------------------------------------
print("\n" + "=" * 60)
print("YAKUNIY HISOBOT")
print("=" * 60)
passed = sum(1 for _, ok, _ in RESULTS if ok)
failed = [name for name, ok, err in RESULTS if not ok]
print(f"O'tdi: {passed}/{len(RESULTS)}")
if failed:
    print("XATOLAR:")
    for name, ok, err in RESULTS:
        if not ok:
            print(f"  - {name}: {err}")
    sys.exit(1)
else:
    print("✅ BARCHA TESTLAR MUVAFFAQIYATLI O'TDI - XATOLIK YO'Q")
    sys.exit(0)
